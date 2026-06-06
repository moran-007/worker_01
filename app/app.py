from __future__ import annotations

import shutil
import base64
import csv
import json
import mimetypes
import re
import uuid
import zipfile
from decimal import Decimal
from io import BytesIO, StringIO
from collections import defaultdict
from datetime import date, datetime, timedelta
from functools import wraps
from pathlib import Path
from urllib.parse import urlparse

from flask import (
    Flask,
    abort,
    current_app,
    flash,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

import config
from database import DatabaseIntegrityError, get_db, init_app as init_database_app
from permissions import (
    ASSIGNABLE_ROLES,
    ROLE_LABELS,
    endpoint_permissions,
    has_any_permission,
    has_permission,
    matches_legacy_roles,
    normalize_role,
    permission_catalog,
    permissions_for_role,
    planned_role_options,
    role_label,
    role_options,
)


WEEKDAY_CHOICES = [
    (1, "周一"),
    (2, "周二"),
    (3, "周三"),
    (4, "周四"),
    (5, "周五"),
    (6, "周六"),
    (7, "周日"),
]
WEEKDAY_MAP = dict(WEEKDAY_CHOICES)

ATTENDANCE_STATUS = ["未确认", "已到", "迟到", "请假", "缺勤", "补签", "试听", "停课"]
DEDUCT_STATUS = {"已到", "迟到", "补签"}

CLASS_STATUS = {
    "active": "上课中",
    "pending": "未开班",
    "paused": "暂停",
    "closed": "已结课",
    "archived": "已归档",
}
LESSON_STATUS = {
    "planned": "待签到",
    "completed": "已完成",
    "cancelled": "已取消",
}
PERSON_STATUS = {
    "active": "启用",
    "inactive": "停用",
}
LESSON_DETAIL_FIELDS = (
    "teaching_content",
    "learning_goal",
    "class_performance",
    "homework",
    "next_plan",
    "materials",
)
PHONE_PATTERN = re.compile(r"^(1[3-9]\d{9}|0\d{2,3}-?\d{7,8})$")
SAFE_METHODS = {"GET", "HEAD", "OPTIONS"}


def create_app() -> Flask:
    app = Flask(__name__)
    frontend_dist = config.BASE_DIR / "frontend" / "dist"
    vue_exact_routes = {
        "/",
        "/login",
        "/student",
        "/parent",
        "/students",
        "/teachers",
        "/classes",
        "/lesson-types",
        "/course-presets",
        "/lessons",
        "/lessons/generate",
        "/statistics",
        "/users",
        "/logs",
        "/database",
    }
    vue_detail_route_pattern = re.compile(r"^/lessons/\d+/(detail|attendance)$")
    app.config.update(
        SECRET_KEY=config.SECRET_KEY,
        DATABASE=str(config.DATABASE_PATH),
        DATABASE_ENGINE=config.DATABASE_ENGINE,
        MYSQL={
            "host": config.MYSQL_HOST,
            "port": config.MYSQL_PORT,
            "user": config.MYSQL_USER,
            "password": config.MYSQL_PASSWORD,
            "database": config.MYSQL_DATABASE,
        },
        DATA_DIR=config.DATA_DIR,
        EXPORT_DIR=config.EXPORT_DIR,
        BACKUP_DIR=config.BACKUP_DIR,
        UPLOAD_DIR=config.UPLOAD_DIR,
        COURSEWARE_UPLOAD_DIR=config.COURSEWARE_UPLOAD_DIR,
        SCRATCH_UPLOAD_DIR=config.SCRATCH_UPLOAD_DIR,
        MATERIAL_UPLOAD_DIR=config.MATERIAL_UPLOAD_DIR,
        MAX_CONTENT_LENGTH=config.UPLOAD_MAX_BYTES,
        ALLOWED_UPLOAD_EXTENSIONS=config.ALLOWED_UPLOAD_EXTENSIONS,
        HYDRO_BASE_URL=config.HYDRO_BASE_URL,
        HYDRO_INTEGRATION_MODE=config.HYDRO_INTEGRATION_MODE,
        FRONTEND_BASE_URL=config.FRONTEND_BASE_URL,
        SCRATCH_EDITOR_URL=config.SCRATCH_EDITOR_URL,
        SCRATCH_OJ_API_URL=config.SCRATCH_OJ_API_URL,
        SCRATCH_EDITOR_PROJECT_DIR=config.SCRATCH_EDITOR_PROJECT_DIR,
    )
    init_database_app(app)

    def serve_vue_shell():
        index_path = frontend_dist / "index.html"
        if index_path.exists():
            return send_from_directory(frontend_dist, "index.html")
        target = config.FRONTEND_BASE_URL.rstrip("/") + request.path
        if request.query_string:
            target += "?" + request.query_string.decode("utf-8", errors="ignore")
        return redirect(target)

    def is_vue_shell_request() -> bool:
        if request.method != "GET":
            return False
        path = request.path.rstrip("/") or "/"
        if path.startswith(("/api/", "/static/", "/assets/")):
            return False
        if path == "/scratch/editor":
            return False
        if path.startswith(("/exports/", "/backups/")) and path.endswith("/download"):
            return False
        if path in vue_exact_routes:
            return True
        return True

    def is_classic_frontend_write_request() -> bool:
        if request.method in SAFE_METHODS:
            return False
        path = request.path.rstrip("/") or "/"
        return not path.startswith("/api/") and path != "/scratch/editor"

    @app.get("/assets/<path:filename>")
    def frontend_assets(filename: str):
        assets_dir = frontend_dist / "assets"
        if assets_dir.exists():
            return send_from_directory(assets_dir, filename)
        abort(404)

    @app.before_request
    def load_current_user():
        if is_classic_frontend_write_request():
            return jsonify({"ok": False, "message": "经典 Flask 页面入口已禁用，请使用 Vue 前端。"}), 410

        user_id = session.get("user_id")
        g.current_user = None
        if user_id:
            user_row = fetch_one(
                """
                SELECT u.*, t.name AS teacher_name, s.name AS student_name
                FROM users u
                LEFT JOIN teachers t ON t.id = u.teacher_id
                LEFT JOIN students s ON s.id = u.student_id
                WHERE u.id = ?
                """,
                (user_id,),
            )
            g.current_user = dict(user_row) if user_row is not None else None
            if g.current_user is None or g.current_user["status"] != "active":
                session.clear()
                g.current_user = None
            else:
                g.current_user["permissions"] = db_permissions_for_role(g.current_user["role"])

        if is_vue_shell_request():
            return serve_vue_shell()

        if request.endpoint in {"login", "api_login", "static", "frontend_assets"}:
            return None
        if g.current_user is None:
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "message": "请先登录。"}), 401
            next_url = request.full_path if request.query_string else request.path
            return redirect(url_for("login", next=next_url))
        if normalize_role(g.current_user["role"]) == "readonly" and request.method not in SAFE_METHODS:
            abort(403)
        required_permissions = endpoint_permissions(request.endpoint, request.method)
        if required_permissions and not has_any_permission(g.current_user, required_permissions):
            abort(403)
        return None

    @app.context_processor
    def inject_globals():
        def can(permission: str) -> bool:
            return has_permission(g.get("current_user"), permission)

        return {
            "weekday_choices": WEEKDAY_CHOICES,
            "weekday_map": WEEKDAY_MAP,
            "attendance_status": ATTENDANCE_STATUS,
            "class_status": CLASS_STATUS,
            "lesson_status": LESSON_STATUS,
            "person_status": PERSON_STATUS,
            "role_labels": ROLE_LABELS,
            "role_options": db_role_options(),
            "normalize_role": normalize_role,
            "can": can,
            "current_user": g.get("current_user"),
            "today": date.today().isoformat(),
        }

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            user = fetch_one(
                """
                SELECT u.*, t.name AS teacher_name, s.name AS student_name
                FROM users u
                LEFT JOIN teachers t ON t.id = u.teacher_id
                LEFT JOIN students s ON s.id = u.student_id
                WHERE u.username = ? AND u.status = 'active'
                """,
                (username,),
            )
            if user and check_password_hash(user["password_hash"], password):
                user = dict(user)
                user["permissions"] = db_permissions_for_role(user["role"])
                session.clear()
                session["user_id"] = user["id"]
                session["role"] = user["role"]
                get_db().execute(
                    "UPDATE users SET last_login_at = ? WHERE id = ?",
                    (now_text(), user["id"]),
                )
                get_db().commit()
                log_operation("登录系统", "user", user["id"], username)
                next_url = request.args.get("next")
                if should_honor_login_next(user, next_url):
                    return redirect(next_url)
                return redirect(default_landing_url(user))
            flash("账号或密码错误，或账号已停用。", "error")
        return render_template("login.html")

    @app.get("/logout")
    def logout():
        if g.current_user:
            log_operation("退出登录", "user", g.current_user["id"], g.current_user["username"])
        session.clear()
        flash("已退出登录。", "success")
        return redirect(url_for("login"))

    @app.post("/api/auth/login")
    def api_login():
        data = request.get_json(silent=True) or request.form
        username = (data.get("username") or "").strip()
        password = data.get("password") or ""
        user = fetch_one(
            """
            SELECT u.*, t.name AS teacher_name, s.name AS student_name
            FROM users u
            LEFT JOIN teachers t ON t.id = u.teacher_id
            LEFT JOIN students s ON s.id = u.student_id
            WHERE u.username = ? AND u.status = 'active'
            """,
            (username,),
        )
        if user and check_password_hash(user["password_hash"], password):
            user = dict(user)
            user["permissions"] = db_permissions_for_role(user["role"])
            session.clear()
            session["user_id"] = user["id"]
            session["role"] = user["role"]
            get_db().execute(
                "UPDATE users SET last_login_at = ? WHERE id = ?",
                (now_text(), user["id"]),
            )
            get_db().commit()
            log_operation("API 登录系统", "user", user["id"], username)
            return jsonify({"ok": True, "user": public_user(user)})
        return jsonify({"ok": False, "message": "账号或密码错误，或账号已停用。"}), 401

    @app.post("/api/auth/logout")
    def api_logout():
        if g.current_user:
            log_operation("API 退出登录", "user", g.current_user["id"], g.current_user["username"])
        session.clear()
        return jsonify({"ok": True})

    @app.get("/api/me")
    def api_me():
        return jsonify({"user": public_user(g.current_user)})

    @app.get("/api/permissions")
    def api_permissions():
        return jsonify(
            {
                "roles": db_role_options(),
                "planned_roles": planned_role_options(),
                "permissions": db_permission_catalog(),
            }
        )

    @app.post("/api/permissions")
    def api_upsert_permission():
        data = request.get_json(silent=True) or {}
        permission_key = (data.get("key") or data.get("permission_key") or "").strip()
        if not permission_key or not re.match(r"^[a-z][a-z0-9_.:-]*$", permission_key):
            return api_error("权限标识格式不正确。")
        label = (data.get("label") or "").strip()
        if not label:
            return api_error("请填写权限名称。")
        now = now_text()
        get_db().execute(
            """
            INSERT INTO permission_definitions
                (permission_key, label, description, category, is_system, status,
                 created_at, updated_at)
            VALUES (?, ?, ?, ?, 0, ?, ?, ?)
            ON CONFLICT(permission_key) DO UPDATE SET
                label = excluded.label,
                description = excluded.description,
                category = excluded.category,
                status = excluded.status,
                updated_at = excluded.updated_at
            """,
            (
                permission_key,
                label,
                (data.get("description") or "").strip(),
                (data.get("category") or "自定义").strip(),
                data.get("status") or "active",
                now,
                now,
            ),
        )
        get_db().commit()
        log_operation("维护权限点", "permission", None, permission_key)
        return jsonify({"ok": True, "permission": get_permission_definition(permission_key)})

    @app.put("/api/permissions/roles/<role_code>")
    def api_update_role_permissions(role_code: str):
        role_code = normalize_role(role_code)
        data = request.get_json(silent=True) or {}
        ok, message = save_role_permissions(role_code, data.get("permissions", []))
        if not ok:
            return api_error(message)
        return jsonify({"ok": True, "role": role_code, "permissions": list(db_permissions_for_role(role_code))})

    @app.get("/permissions")
    def permission_matrix():
        roles = db_role_options(active_only=False)
        permissions = db_permission_catalog()
        categories = []
        grouped = {}
        for permission in permissions:
            category = permission["category"] or "未分类"
            if category not in grouped:
                categories.append(category)
                grouped[category] = []
            grouped[category].append(permission)
        return render_template(
            "permission_matrix.html",
            roles=roles,
            permissions=permissions,
            categories=categories,
            grouped_permissions=grouped,
        )

    @app.post("/permissions/roles")
    def create_role():
        result = save_role_definition(request.form)
        if result[0]:
            flash("身份已创建。", "success")
        else:
            flash(result[1], "error")
        return redirect(url_for("permission_matrix"))

    @app.post("/permissions/roles/<role_code>")
    def update_role(role_code: str):
        data = request.form
        if data.get("action") == "save_permissions":
            permission_keys = request.form.getlist("permissions")
            ok, message = save_role_permissions(role_code, permission_keys)
        else:
            ok, message = update_role_definition(role_code, data)
        flash(message, "success" if ok else "error")
        return redirect(url_for("permission_matrix"))

    @app.post("/api/permissions/roles")
    def api_create_role():
        data = request.get_json(silent=True) or {}
        ok, message_or_role = save_role_definition(data)
        if not ok:
            return api_error(message_or_role)
        return jsonify({"ok": True, "role": message_or_role})

    @app.put("/api/permissions/roles/<role_code>/definition")
    def api_update_role(role_code: str):
        data = request.get_json(silent=True) or {}
        ok, message_or_role = update_role_definition(role_code, data)
        if not ok:
            return api_error(message_or_role)
        return jsonify({"ok": True, "role": message_or_role})

    @app.get("/student")
    def student_portal():
        if normalize_role(g.current_user["role"]) != "student":
            return redirect(url_for("dashboard"))
        if not g.current_user.get("student_id"):
            return render_template(
                "student_portal.html",
                portal=empty_student_portal_context(),
                preview_mode=False,
            )
        portal = redact_student_hours(build_student_portal_context(g.current_user["student_id"]))
        return render_template("student_portal.html", portal=portal, preview_mode=False)

    @app.get("/api/student/me")
    def api_student_me():
        if normalize_role(g.current_user["role"]) != "student":
            return api_error("当前账号不是学生账号。", 403)
        if not g.current_user.get("student_id"):
            return jsonify({"ok": True, "portal": empty_student_portal_context()})
        portal = redact_student_hours(build_student_portal_context(g.current_user["student_id"]))
        return jsonify({"ok": True, "portal": portal})

    @app.post("/student/scratch/tasks/<int:lesson_id>/<int:template_id>/start")
    def student_start_scratch_task(lesson_id: int, template_id: int):
        user = g.current_user
        if normalize_role(user["role"]) != "student" or not user.get("student_id"):
            abort(403)
        try:
            work = create_student_scratch_work(lesson_id, template_id, user["student_id"], user["id"])
        except ValueError as error:
            flash(str(error), "error")
            return redirect(url_for("student_portal"))
        flash("Scratch 作品已创建，可以开始编辑。", "success")
        return redirect(url_for("scratch_editor", work_id=work["id"]))

    @app.get("/parent")
    def parent_portal():
        if normalize_role(g.current_user["role"]) != "parent":
            return redirect(url_for("dashboard"))
        if not g.current_user.get("student_id"):
            return render_template(
                "parent_portal.html",
                portal=empty_parent_portal_context(),
                preview_mode=False,
            )
        portal = build_parent_portal_context(g.current_user["student_id"])
        return render_template("parent_portal.html", portal=portal, preview_mode=False)

    @app.get("/api/parent/me")
    def api_parent_me():
        if normalize_role(g.current_user["role"]) != "parent":
            return api_error("当前账号不是家长账号。", 403)
        if not g.current_user.get("student_id"):
            return jsonify({"ok": True, "portal": empty_parent_portal_context()})
        return jsonify({"ok": True, "portal": build_parent_portal_context(g.current_user["student_id"])})

    @app.get("/students/<int:student_id>/portal")
    def student_portal_preview(student_id: int):
        ensure_student_access(student_id)
        portal = redact_student_hours(build_student_portal_context(student_id))
        return render_template("student_portal.html", portal=portal, preview_mode=True)

    @app.get("/api/students/<int:student_id>/portal")
    def api_student_portal(student_id: int):
        ensure_student_access(student_id)
        return jsonify({"ok": True, "portal": redact_student_hours(build_student_portal_context(student_id))})

    @app.get("/students/<int:student_id>/parent")
    def parent_portal_preview(student_id: int):
        ensure_student_access(student_id)
        portal = build_parent_portal_context(student_id)
        return render_template("parent_portal.html", portal=portal, preview_mode=True)

    @app.get("/api/students/<int:student_id>/parent")
    def api_parent_portal(student_id: int):
        ensure_student_access(student_id)
        return jsonify({"ok": True, "portal": build_parent_portal_context(student_id)})

    @app.get("/api/uploads/assets")
    def api_uploaded_assets():
        filters = {
            "asset_type": request.args.get("asset_type", "").strip(),
            "usage_scope": request.args.get("usage_scope", "").strip(),
            "status": request.args.get("status", "").strip(),
        }
        return jsonify({"items": rows_to_dicts(query_uploaded_assets(filters))})

    @app.post("/api/uploads/assets")
    def api_upload_asset():
        upload = request.files.get("file")
        if upload is None or not upload.filename:
            return api_error("请上传文件。")
        try:
            asset = save_uploaded_asset(
                upload,
                usage_scope=request.form.get("usage_scope") or "courseware",
                asset_type=request.form.get("asset_type") or "",
                metadata={
                    "title": request.form.get("title", "").strip(),
                    "source": request.form.get("source", "").strip(),
                },
            )
        except ValueError as error:
            return api_error(str(error))
        log_operation("上传资源", "asset", asset["id"], asset["original_filename"])
        return jsonify({"ok": True, "asset": asset})

    @app.get("/api/lessons/<int:lesson_id>/assets")
    def api_lesson_assets(lesson_id: int):
        lesson = get_lesson(lesson_id)
        if lesson is None:
            abort(404)
        ensure_lesson_access(lesson)
        return jsonify({"items": rows_to_dicts(query_lesson_assets(lesson_id))})

    @app.post("/api/lessons/<int:lesson_id>/assets")
    def api_upload_lesson_asset(lesson_id: int):
        lesson = get_lesson(lesson_id)
        if lesson is None:
            abort(404)
        ensure_lesson_access(lesson)
        try:
            item = create_lesson_asset(lesson_id, request)
        except ValueError as error:
            return api_error(str(error))
        log_operation("上传课次课件", "lesson", lesson_id, item["title"] or item["original_filename"])
        return jsonify({"ok": True, "asset": item})

    @app.delete("/api/lessons/<int:lesson_id>/assets/<int:lesson_asset_id>")
    def api_delete_lesson_asset(lesson_id: int, lesson_asset_id: int):
        lesson = get_lesson(lesson_id)
        if lesson is None:
            abort(404)
        ensure_lesson_access(lesson)
        if not deactivate_lesson_asset(lesson_id, lesson_asset_id):
            abort(404)
        log_operation("移除课次课件", "lesson", lesson_id, str(lesson_asset_id))
        return jsonify({"ok": True})

    @app.get("/api/uploads/assets/<int:asset_id>/download")
    def api_download_asset(asset_id: int):
        asset = get_uploaded_asset(asset_id)
        if asset is None:
            abort(404)
        if not can_access_uploaded_asset(asset):
            abort(403)
        return send_asset_file(asset)

    @app.get("/api/scratch/editor/config")
    def api_scratch_editor_config():
        return jsonify(scratch_editor_config())

    @app.get("/scratch/editor")
    def scratch_editor():
        work_id = parse_int(request.args.get("work_id"))
        template_id = parse_int(request.args.get("template_id"))
        work = get_scratch_work(work_id) if work_id else None
        if work is not None:
            ensure_scratch_work_access(work, write=False)
        template = get_scratch_template(template_id) if template_id else None
        if template_id and template is None:
            abort(404)
        if template is not None:
            ensure_scratch_template_preview_access(template["id"])
        project_url = ""
        if work and work.get("asset_url"):
            project_url = absolute_app_url(work["asset_url"])
        elif template and template.get("asset_url"):
            project_url = absolute_app_url(template["asset_url"])
        title = work["title"] if work else (template["title"] if template else "Scratch 编辑器")
        editor_context = {
            "workId": str(work_id or ""),
            "templateId": str(template_id or ""),
            "title": title,
            "projectUrl": project_url,
            "editorUrl": scratch_editor_config()["editor_url"],
            "editorOrigin": origin_from_url(scratch_editor_config()["editor_url"]),
            "saveEndpoint": url_for("api_editor_save_scratch_work", work_id=work["id"]) if work else "",
            "fallbackSaveEndpoint": url_for("api_save_scratch_work", work_id=work["id"]) if work else "",
            "submitEndpoint": url_for("api_submit_scratch_work", work_id=work["id"]) if work else "",
        }
        return render_template(
            "scratch_editor.html",
            config=scratch_editor_config(),
            work=work,
            template=template,
            title=title,
            work_id=str(work_id or ""),
            template_id=str(template_id or ""),
            project_url=project_url,
            editor_context=editor_context,
        )

    @app.get("/api/scratch/templates")
    def api_scratch_templates():
        filters = {
            "q": request.args.get("q", "").strip(),
            "status": request.args.get("status", "active").strip(),
        }
        return jsonify({"items": rows_to_dicts(query_scratch_templates(filters))})

    @app.post("/api/scratch/templates")
    def api_create_scratch_template():
        try:
            template = create_scratch_template(request)
        except ValueError as error:
            return api_error(str(error))
        log_operation("新增 Scratch 模板", "scratch_template", template["id"], template["title"])
        return jsonify({"ok": True, "template": template})

    @app.put("/api/scratch/templates/<int:template_id>")
    def api_update_scratch_template(template_id: int):
        data = request.get_json(silent=True) or {}
        try:
            template = update_scratch_template(template_id, data)
        except ValueError as error:
            return api_error(str(error))
        if template is None:
            abort(404)
        log_operation("更新 Scratch 模板", "scratch_template", template_id, template["title"])
        return jsonify({"ok": True, "template": template})

    @app.get("/api/lessons/<int:lesson_id>/scratch/templates")
    def api_lesson_scratch_templates(lesson_id: int):
        lesson = get_lesson(lesson_id)
        if lesson is None:
            abort(404)
        ensure_lesson_access(lesson)
        return jsonify({"lesson": dict(lesson), "items": rows_to_dicts(query_lesson_scratch_templates(lesson_id))})

    @app.post("/api/lessons/<int:lesson_id>/scratch/templates")
    def api_bind_lesson_scratch_template(lesson_id: int):
        lesson = get_lesson(lesson_id)
        if lesson is None:
            abort(404)
        ensure_lesson_access(lesson)
        data = request.get_json(silent=True) or request.form
        template_id = parse_int(data.get("template_id"))
        if not template_id:
            return api_error("请选择 Scratch 模板。")
        try:
            binding = bind_lesson_scratch_template(lesson_id, template_id, data.get("bind_note", ""))
        except ValueError as error:
            return api_error(str(error))
        log_operation("绑定课次 Scratch 模板", "lesson", lesson_id, str(template_id))
        return jsonify({"ok": True, "binding": binding})

    @app.post("/api/lessons/<int:lesson_id>/scratch/assignments")
    def api_create_lesson_scratch_assignment(lesson_id: int):
        lesson = get_lesson(lesson_id)
        if lesson is None:
            abort(404)
        ensure_lesson_access(lesson)
        try:
            assignment = create_lesson_scratch_assignment(lesson_id, request)
        except ValueError as error:
            return api_error(str(error))
        log_operation("发布 Scratch 作业", "lesson", lesson_id, str(assignment["template_id"]))
        return jsonify({"ok": True, "assignment": assignment})

    @app.put("/api/lessons/<int:lesson_id>/scratch/templates/<int:template_id>/assignment")
    def api_update_lesson_scratch_assignment(lesson_id: int, template_id: int):
        lesson = get_lesson(lesson_id)
        if lesson is None:
            abort(404)
        ensure_lesson_access(lesson)
        try:
            assignment = update_lesson_scratch_assignment(
                lesson_id,
                template_id,
                request.get_json(silent=True) or request.form,
            )
        except ValueError as error:
            return api_error(str(error))
        if assignment is None:
            abort(404)
        log_operation("更新 Scratch 作业", "lesson", lesson_id, str(template_id))
        return jsonify({"ok": True, "assignment": assignment})

    @app.delete("/api/lessons/<int:lesson_id>/scratch/templates/<int:template_id>")
    def api_unbind_lesson_scratch_template(lesson_id: int, template_id: int):
        lesson = get_lesson(lesson_id)
        if lesson is None:
            abort(404)
        ensure_lesson_access(lesson)
        get_db().execute(
            """
            UPDATE lesson_scratch_templates
            SET status = 'inactive'
            WHERE lesson_id = ? AND template_id = ?
            """,
            (lesson_id, template_id),
        )
        get_db().commit()
        log_operation("取消绑定课次 Scratch 模板", "lesson", lesson_id, str(template_id))
        return jsonify({"ok": True})

    @app.get("/api/lessons/<int:lesson_id>/scratch/works")
    def api_lesson_scratch_works(lesson_id: int):
        lesson = get_lesson(lesson_id)
        if lesson is None:
            abort(404)
        ensure_lesson_access(lesson)
        return jsonify({"lesson": dict(lesson), "items": rows_to_dicts(query_scratch_works({"lesson_id": lesson_id}))})

    @app.get("/api/student/scratch/works")
    def api_student_scratch_works():
        user = g.current_user
        if normalize_role(user["role"]) != "student" or not user.get("student_id"):
            return api_error("当前账号不是学生账号。", 403)
        return jsonify({"items": rows_to_dicts(query_student_scratch_tasks(user["student_id"]))})

    @app.post("/api/lessons/<int:lesson_id>/scratch/templates/<int:template_id>/work")
    def api_create_scratch_work(lesson_id: int, template_id: int):
        user = g.current_user
        if normalize_role(user["role"]) != "student" or not user.get("student_id"):
            return api_error("只有学生账号可以复制模板为个人作品。", 403)
        try:
            work = create_student_scratch_work(lesson_id, template_id, user["student_id"], user["id"])
        except ValueError as error:
            return api_error(str(error))
        log_operation("复制 Scratch 模板为学生作品", "scratch_work", work["id"], work["title"])
        return jsonify({"ok": True, "work": work})

    @app.put("/api/scratch/works/<int:work_id>/save")
    def api_save_scratch_work(work_id: int):
        work = get_scratch_work(work_id)
        if work is None:
            abort(404)
        ensure_scratch_work_access(work, write=True)
        try:
            updated = save_scratch_work_progress(work, request)
        except ValueError as error:
            return api_error(str(error))
        log_operation("保存 Scratch 作品", "scratch_work", work_id, updated["title"])
        return jsonify({"ok": True, "work": updated})

    @app.post("/api/scratch/works/<int:work_id>/editor-save")
    def api_editor_save_scratch_work(work_id: int):
        work = get_scratch_work(work_id)
        if work is None:
            abort(404)
        ensure_scratch_work_access(work, write=True)
        try:
            updated = save_scratch_work_from_editor(work, request.get_json(silent=True) or {})
        except ValueError as error:
            return api_error(str(error))
        log_operation("编辑器保存 Scratch 作品", "scratch_work", work_id, updated["title"])
        return jsonify({"ok": True, "work": updated})

    @app.post("/api/scratch/works/<int:work_id>/submit")
    def api_submit_scratch_work(work_id: int):
        work = get_scratch_work(work_id)
        if work is None:
            abort(404)
        ensure_scratch_work_access(work, write=True)
        updated = submit_scratch_work(work_id, request.get_json(silent=True) or request.form)
        log_operation("提交 Scratch 作品", "scratch_work", work_id, updated["title"])
        return jsonify({"ok": True, "work": updated})

    @app.get("/api/scratch/works/<int:work_id>/judge-runs")
    def api_scratch_work_judge_runs(work_id: int):
        work = get_scratch_work(work_id)
        if work is None:
            abort(404)
        ensure_scratch_work_access(work, write=False)
        return jsonify({"items": rows_to_dicts(query_scratch_judge_runs(work_id))})

    @app.post("/api/scratch/works/<int:work_id>/judge")
    def api_run_scratch_work_judge(work_id: int):
        work = get_scratch_work(work_id)
        if work is None:
            abort(404)
        ensure_scratch_work_review_access(work)
        try:
            result = run_scratch_work_judge(work)
        except ValueError as error:
            return api_error(str(error))
        log_operation("Scratch 自动测评", "scratch_work", work_id, result["status"])
        return jsonify({"ok": True, "result": result, "work": get_scratch_work(work_id)})

    @app.put("/api/scratch/works/<int:work_id>/review")
    def api_review_scratch_work(work_id: int):
        work = get_scratch_work(work_id)
        if work is None:
            abort(404)
        ensure_scratch_work_review_access(work)
        data = request.get_json(silent=True) or {}
        updated = review_scratch_work(work_id, data)
        log_operation("点评 Scratch 作品", "scratch_work", work_id, updated["title"])
        return jsonify({"ok": True, "work": updated})

    @app.post("/api/scratch/test-points/generate")
    def api_generate_scratch_test_points():
        try:
            config_payload = generate_judge_config_from_spec(request.get_json(silent=True) or {})
        except ValueError as error:
            return api_error(str(error))
        return jsonify({"ok": True, "judge_config": config_payload})

    @app.get("/api/scratch/material-categories")
    def api_scratch_material_categories():
        rows = fetch_all(
            """
            SELECT *
            FROM scratch_material_categories
            WHERE status = 'active'
            ORDER BY sort_order, id
            """
        )
        return jsonify({"items": rows_to_dicts(rows)})

    @app.get("/api/scratch/materials")
    def api_scratch_materials():
        rows = fetch_all(
            """
            SELECT sm.*, c.name AS category_name, a.original_filename, a.public_path
            FROM scratch_materials sm
            LEFT JOIN scratch_material_categories c ON c.id = sm.category_id
            LEFT JOIN uploaded_assets a ON a.id = sm.asset_id
            WHERE sm.status = 'active'
            ORDER BY sm.created_at DESC, sm.id DESC
            """
        )
        return jsonify({"items": rows_to_dicts(rows), "reserved": True})

    @app.get("/api/integrations/accounts")
    def api_external_account_bindings():
        rows = fetch_all(
            """
            SELECT b.*, u.username, u.display_name
            FROM external_account_bindings b
            JOIN users u ON u.id = b.user_id
            ORDER BY b.provider, b.created_at DESC
            """
        )
        return jsonify({"items": rows_to_dicts(rows)})

    @app.post("/api/integrations/accounts")
    def api_upsert_external_account_binding():
        data = request.get_json(silent=True) or {}
        user_id = parse_int(data.get("user_id"))
        provider = (data.get("provider") or "hydro").strip().lower()
        external_user_id = (data.get("external_user_id") or "").strip()
        if user_id is None or not external_user_id:
            return api_error("请选择本地账号并填写外部账号 ID。")
        if fetch_one("SELECT id FROM users WHERE id = ?", (user_id,)) is None:
            return api_error("本地账号不存在。", 404)
        now = now_text()
        metadata = data.get("metadata") or {}
        if not isinstance(metadata, str):
            metadata = json.dumps(metadata, ensure_ascii=False)
        try:
            get_db().execute(
                """
                INSERT INTO external_account_bindings
                    (user_id, provider, external_user_id, external_username, display_name,
                     status, metadata_json, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(provider, external_user_id) DO UPDATE SET
                    user_id = excluded.user_id,
                    external_username = excluded.external_username,
                    display_name = excluded.display_name,
                    status = excluded.status,
                    metadata_json = excluded.metadata_json,
                    updated_at = excluded.updated_at
                """,
                (
                    user_id,
                    provider,
                    external_user_id,
                    (data.get("external_username") or "").strip(),
                    (data.get("display_name") or "").strip(),
                    data.get("status") or "active",
                    metadata,
                    now,
                    now,
                ),
            )
            get_db().commit()
        except DatabaseIntegrityError:
            return api_error("该用户已绑定同类型外部系统账号。")
        log_operation("维护外部账号绑定", "external_account", user_id, f"{provider}:{external_user_id}")
        return jsonify({"ok": True})

    @app.post("/api/me/password")
    def api_change_my_password():
        data = request.get_json(silent=True) or {}
        current_password = data.get("current_password", "")
        new_password = data.get("new_password", "").strip()
        confirm_password = data.get("confirm_password", "").strip()
        user = fetch_one("SELECT * FROM users WHERE id = ?", (g.current_user["id"],))
        if user is None:
            abort(404)
        if not check_password_hash(user["password_hash"], current_password):
            return jsonify({"ok": False, "message": "当前密码不正确。"}), 400
        if len(new_password) < 6:
            return jsonify({"ok": False, "message": "新密码至少需要 6 位。"}), 400
        if new_password != confirm_password:
            return jsonify({"ok": False, "message": "两次输入的新密码不一致。"}), 400
        get_db().execute(
            "UPDATE users SET password_hash = ? WHERE id = ?",
            (generate_password_hash(new_password), user["id"]),
        )
        get_db().commit()
        log_operation("自助修改密码", "user", user["id"], user["username"])
        return jsonify({"ok": True})

    @app.get("/api/dashboard")
    def api_dashboard():
        today_value = date.today().isoformat()
        lesson_filters = apply_teacher_scope({"date": today_value})
        lessons = query_lessons(lesson_filters)
        teacher_filters = apply_teacher_scope({})
        status_counts = count_attendance_by_status(today_value, today_value, teacher_filters)
        expected_total = sum(row["expected_count"] for row in lessons)
        arrived_total = sum(
            status_counts.get(status, 0) for status in ("已到", "迟到", "补签", "试听")
        )
        summary = {
            "lesson_count": len(lessons),
            "pending_count": sum(1 for row in lessons if row["status"] == "planned"),
            "completed_count": sum(1 for row in lessons if row["status"] == "completed"),
            "expected_total": expected_total,
            "arrived_total": arrived_total,
            "leave_total": status_counts.get("请假", 0),
            "absent_total": status_counts.get("缺勤", 0),
        }
        return jsonify(
            {
                "summary": summary,
                "today_lessons": rows_to_dicts(lessons),
                "upcoming_lessons": rows_to_dicts(
                    query_upcoming_lessons(today_value, filters=teacher_filters)
                ),
                "low_hour_students": rows_to_dicts(query_low_hour_students(filters=teacher_filters)),
                "consecutive_absences": query_consecutive_absences(filters=teacher_filters),
                "recent_logs": rows_to_dicts(query_operation_logs({}, limit=8))
                if has_permission(g.current_user, "logs.view")
                else [],
            }
        )

    @app.get("/api/students")
    def api_students():
        filters = {
            "q": request.args.get("q", "").strip(),
            "class_id": request.args.get("class_id", "").strip(),
            "status": request.args.get("status", "active").strip(),
        }
        return jsonify({"items": rows_to_dicts(query_students(apply_teacher_scope(filters)))})

    @app.get("/api/teachers")
    def api_teachers():
        filters = {
            "q": request.args.get("q", "").strip(),
            "status": request.args.get("status", "active").strip(),
        }
        if g.current_user["role"] == "teacher":
            filters["teacher_id"] = str(g.current_user["teacher_id"] or 0)
        return jsonify({"items": rows_to_dicts(query_teachers(filters))})

    @app.get("/api/classes")
    def api_classes():
        filters = {
            "q": request.args.get("q", "").strip(),
            "status": request.args.get("status", "active").strip(),
        }
        return jsonify({"items": rows_to_dicts(query_classes(apply_teacher_scope(filters)))})

    @app.get("/api/lessons")
    def api_lessons():
        range_info = resolve_date_range(request.args)
        filters = {
            "date": request.args.get("date", "").strip(),
            "start_date": range_info["start_date"],
            "end_date": range_info["end_date"],
            "class_id": request.args.get("class_id", "").strip(),
            "teacher_id": request.args.get("teacher_id", "").strip(),
            "lesson_type_id": request.args.get("lesson_type_id", "").strip(),
            "status": request.args.get("status", "").strip(),
        }
        return jsonify(
            {
                "items": rows_to_dicts(query_lessons(apply_teacher_scope(filters))),
                "range": range_info,
            }
        )

    @app.get("/api/statistics")
    def api_statistics():
        range_info = resolve_date_range(request.args)
        filters = {
            "class_id": request.args.get("class_id", "").strip(),
            "teacher_id": request.args.get("teacher_id", "").strip(),
            "lesson_type_id": request.args.get("lesson_type_id", "").strip(),
            "student_id": request.args.get("student_id", "").strip(),
            "attendance_status": request.args.get("attendance_status", "").strip(),
        }
        filters = apply_teacher_scope(filters)
        rows = query_attendance_rows(range_info["start_date"], range_info["end_date"], filters)
        lesson_rows = query_lessons_for_statistics(
            range_info["start_date"], range_info["end_date"], filters
        )
        summary, class_stats, teacher_stats = build_statistics(rows, lesson_rows)
        return jsonify(
            {
                "range": range_info,
                "summary": summary,
                "class_stats": class_stats,
                "teacher_stats": teacher_stats,
                "rows": rows_to_dicts(rows),
            }
        )

    @app.get("/api/database/meta")
    @require_roles("admin", "staff")
    def api_database_meta():
        if app.config.get("DATABASE_ENGINE") == "mysql":
            database_name = app.config["MYSQL"]["database"]
            tables = rows_to_dicts(
                fetch_all(
                    """
                    SELECT
                        TABLE_NAME AS name,
                        CASE
                            WHEN TABLE_TYPE = 'BASE TABLE' THEN 'table'
                            ELSE 'view'
                        END AS type
                    FROM INFORMATION_SCHEMA.TABLES
                    WHERE TABLE_SCHEMA = ?
                      AND TABLE_TYPE IN ('BASE TABLE', 'VIEW')
                    ORDER BY type, name
                    """,
                    (database_name,),
                )
            )
            foreign_keys = {}
            for item in tables:
                if item["type"] == "table":
                    foreign_keys[item["name"]] = rows_to_dicts(
                        fetch_all(
                            """
                            SELECT
                                CONSTRAINT_NAME AS id,
                                TABLE_NAME AS `table`,
                                COLUMN_NAME AS `from`,
                                REFERENCED_TABLE_NAME AS ref_table,
                                REFERENCED_COLUMN_NAME AS `to`
                            FROM INFORMATION_SCHEMA.KEY_COLUMN_USAGE
                            WHERE TABLE_SCHEMA = ?
                              AND TABLE_NAME = ?
                              AND REFERENCED_TABLE_NAME IS NOT NULL
                            ORDER BY CONSTRAINT_NAME, ORDINAL_POSITION
                            """,
                            (database_name, item["name"]),
                        )
                    )
            return jsonify({"objects": tables, "foreign_keys": foreign_keys})

        tables = rows_to_dicts(
            fetch_all(
                """
                SELECT name, type
                FROM sqlite_master
                WHERE type IN ('table', 'view')
                  AND name NOT LIKE 'sqlite_%'
                ORDER BY type, name
                """
            )
        )
        foreign_keys = {}
        for item in tables:
            if item["type"] == "table":
                foreign_keys[item["name"]] = rows_to_dicts(
                    fetch_all(f"PRAGMA foreign_key_list({item['name']})")
                )
        return jsonify({"objects": tables, "foreign_keys": foreign_keys})

    @app.get("/api/lesson-types")
    def api_lesson_types():
        rows = fetch_all("SELECT * FROM lesson_types ORDER BY id")
        return jsonify({"items": rows_to_dicts(rows)})

    @app.get("/api/course-presets")
    def api_course_presets():
        filters = {
            "q": request.args.get("q", "").strip(),
            "category": request.args.get("category", "").strip(),
            "stage": request.args.get("stage", "").strip(),
            "status": request.args.get("status", "active").strip(),
        }
        return jsonify({"items": rows_to_dicts(query_course_presets(filters))})

    @app.get("/api/course-presets/import-template")
    @require_roles("admin", "staff", "teacher")
    def api_download_course_preset_template():
        file_path = export_course_preset_import_template()
        return send_file(file_path, as_attachment=True, download_name=file_path.name)

    @app.post("/api/course-presets")
    @require_roles("admin", "staff")
    def api_create_course_preset():
        data = request.get_json(silent=True) or {}
        if error := validate_course_preset_payload(data):
            return api_error(error)
        try:
            db = get_db()
            cursor = db.execute(
                """
                INSERT INTO course_presets
                    (category, stage, lesson_no, course_name, lesson_type_id,
                     default_hours, status, remark, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                course_preset_values(data) + (now_text(),),
            )
            db.commit()
        except DatabaseIntegrityError:
            return jsonify({"ok": False, "message": "同一大类、阶段、节次和课程名称已存在。"}), 400
        log_operation("API 新增预设课程", "course_preset", cursor.lastrowid, data.get("course_name", "").strip())
        return jsonify({"ok": True, "id": cursor.lastrowid})

    @app.put("/api/course-presets/<int:preset_id>")
    @require_roles("admin", "staff")
    def api_update_course_preset(preset_id: int):
        data = request.get_json(silent=True) or {}
        preset = fetch_one("SELECT * FROM course_presets WHERE id = ?", (preset_id,))
        if preset is None:
            abort(404)
        if error := validate_course_preset_payload(data):
            return api_error(error)
        try:
            get_db().execute(
                """
                UPDATE course_presets
                SET category = ?, stage = ?, lesson_no = ?, course_name = ?,
                    lesson_type_id = ?, default_hours = ?, status = ?, remark = ?
                WHERE id = ?
                """,
                course_preset_values(data) + (preset_id,),
            )
            get_db().commit()
        except DatabaseIntegrityError:
            return jsonify({"ok": False, "message": "同一大类、阶段、节次和课程名称已存在。"}), 400
        log_operation("API 编辑预设课程", "course_preset", preset_id, preset["course_name"])
        return jsonify({"ok": True})

    @app.post("/api/course-presets/import")
    @require_roles("admin", "staff")
    def api_import_course_presets():
        upload = request.files.get("file")
        if upload is None or not upload.filename:
            return jsonify({"ok": False, "message": "请上传 Excel 或 CSV 文件。"}), 400
        try:
            rows = read_course_preset_import_rows(upload)
            result = import_course_presets(rows)
        except ValueError as error:
            return jsonify({"ok": False, "message": str(error)}), 400
        log_operation(
            "API 批量导入预设课程",
            "course_preset",
            None,
            f"新增 {result['created']}，更新 {result['updated']}，跳过 {result['skipped']}",
        )
        return jsonify({"ok": True, **result})

    @app.post("/api/lessons")
    @require_roles("admin", "staff")
    def api_create_lesson():
        data = request.get_json(silent=True) or {}
        lesson_date = (data.get("lesson_date") or "").strip()
        preset = get_course_preset(parse_int(data.get("course_preset_id")))
        if error := validate_lesson_payload(data, preset):
            return api_error(error)
        lesson_type_id = parse_int(data.get("lesson_type_id")) or (
            preset["lesson_type_id"] if preset else None
        )
        lesson_hours = parse_float(
            data.get("lesson_hours"),
            float(preset["default_hours"] or 1) if preset else 1,
        )
        topic = ((data.get("lesson_topic") or "").strip() or (preset["course_name"] if preset else ""))
        db = get_db()
        cursor = db.execute(
            """
            INSERT INTO lessons
                (class_id, lesson_date, weekday, start_time, end_time, teacher_id,
                 lesson_type_id, course_preset_id, lesson_topic, lesson_hours,
                 classroom, status, remark, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'planned', ?, ?)
            """,
            (
                parse_int(data.get("class_id")),
                lesson_date,
                parse_date(lesson_date).isoweekday() if parse_date(lesson_date) else None,
                (data.get("start_time") or "").strip(),
                (data.get("end_time") or "").strip(),
                resolve_lesson_teacher_id(data.get("teacher_id")),
                lesson_type_id,
                preset["id"] if preset else None,
                topic,
                lesson_hours,
                (data.get("classroom") or "").strip(),
                (data.get("remark") or "").strip(),
                now_text(),
            ),
        )
        db.commit()
        log_operation("API 新增课次", "lesson", cursor.lastrowid, topic)
        return jsonify({"ok": True, "id": cursor.lastrowid})

    @app.post("/api/lessons/<int:lesson_id>/cancel")
    @require_roles("admin", "staff")
    def api_cancel_lesson(lesson_id: int):
        lesson = get_lesson(lesson_id)
        if lesson is None:
            abort(404)
        ensure_lesson_access(lesson)
        get_db().execute("UPDATE lessons SET status = 'cancelled' WHERE id = ?", (lesson_id,))
        get_db().commit()
        log_operation("API 取消课次", "lesson", lesson_id, lesson["class_name"])
        return jsonify({"ok": True})

    @app.post("/api/students")
    @require_roles("admin", "staff")
    def api_create_student():
        data = request.get_json(silent=True) or {}
        if error := validate_student_payload(data):
            return api_error(error)
        db = get_db()
        cursor = db.execute(
            """
            INSERT INTO students
                (name, gender, age, phone, parent_name, parent_phone, school,
                 purchased_hours, gift_hours, status, remark, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                data.get("name", "").strip(),
                data.get("gender", "").strip(),
                parse_int(data.get("age")),
                data.get("phone", "").strip(),
                data.get("parent_name", "").strip(),
                data.get("parent_phone", "").strip(),
                data.get("school", "").strip(),
                parse_float(data.get("purchased_hours")),
                parse_float(data.get("gift_hours")),
                data.get("status", "active"),
                data.get("remark", "").strip(),
                now_text(),
            ),
        )
        db.commit()
        log_operation("API 新增学生", "student", cursor.lastrowid, data.get("name", "").strip())
        return jsonify({"ok": True, "id": cursor.lastrowid})

    @app.put("/api/students/<int:student_id>")
    @require_roles("admin", "staff")
    def api_update_student(student_id: int):
        data = request.get_json(silent=True) or {}
        student = fetch_one("SELECT * FROM students WHERE id = ?", (student_id,))
        if student is None:
            abort(404)
        if error := validate_student_payload(data):
            return api_error(error)
        get_db().execute(
            """
            UPDATE students
            SET name = ?, gender = ?, age = ?, phone = ?, parent_name = ?,
                parent_phone = ?, school = ?, purchased_hours = ?, gift_hours = ?,
                status = ?, remark = ?
            WHERE id = ?
            """,
            (
                data.get("name", "").strip(),
                data.get("gender", "").strip(),
                parse_int(data.get("age")),
                data.get("phone", "").strip(),
                data.get("parent_name", "").strip(),
                data.get("parent_phone", "").strip(),
                data.get("school", "").strip(),
                parse_float(data.get("purchased_hours")),
                parse_float(data.get("gift_hours")),
                data.get("status", "active"),
                data.get("remark", "").strip(),
                student_id,
            ),
        )
        get_db().commit()
        log_operation("API 编辑学生", "student", student_id, student["name"])
        return jsonify({"ok": True})

    @app.post("/api/students/<int:student_id>/deactivate")
    @require_roles("admin", "staff")
    def api_deactivate_student(student_id: int):
        student = fetch_one("SELECT * FROM students WHERE id = ?", (student_id,))
        if student is None:
            abort(404)
        get_db().execute("UPDATE students SET status = 'inactive' WHERE id = ?", (student_id,))
        get_db().commit()
        log_operation("API 停用学生", "student", student_id, student["name"])
        return jsonify({"ok": True})

    @app.post("/api/teachers")
    @require_roles("admin", "staff")
    def api_create_teacher():
        data = request.get_json(silent=True) or {}
        if error := validate_teacher_payload(data):
            return api_error(error)
        db = get_db()
        cursor = db.execute(
            """
            INSERT INTO teachers (name, phone, subject, status, remark, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                data.get("name", "").strip(),
                data.get("phone", "").strip(),
                data.get("subject", "").strip(),
                data.get("status", "active"),
                data.get("remark", "").strip(),
                now_text(),
            ),
        )
        db.commit()
        log_operation("API 新增教师", "teacher", cursor.lastrowid, data.get("name", "").strip())
        return jsonify({"ok": True, "id": cursor.lastrowid})

    @app.put("/api/teachers/<int:teacher_id>")
    @require_roles("admin", "staff")
    def api_update_teacher(teacher_id: int):
        data = request.get_json(silent=True) or {}
        teacher = fetch_one("SELECT * FROM teachers WHERE id = ?", (teacher_id,))
        if teacher is None:
            abort(404)
        if error := validate_teacher_payload(data):
            return api_error(error)
        get_db().execute(
            """
            UPDATE teachers
            SET name = ?, phone = ?, subject = ?, status = ?, remark = ?
            WHERE id = ?
            """,
            (
                data.get("name", "").strip(),
                data.get("phone", "").strip(),
                data.get("subject", "").strip(),
                data.get("status", "active"),
                data.get("remark", "").strip(),
                teacher_id,
            ),
        )
        get_db().commit()
        log_operation("API 编辑教师", "teacher", teacher_id, teacher["name"])
        return jsonify({"ok": True})

    @app.post("/api/teachers/<int:teacher_id>/deactivate")
    @require_roles("admin", "staff")
    def api_deactivate_teacher(teacher_id: int):
        teacher = fetch_one("SELECT * FROM teachers WHERE id = ?", (teacher_id,))
        if teacher is None:
            abort(404)
        get_db().execute("UPDATE teachers SET status = 'inactive' WHERE id = ?", (teacher_id,))
        get_db().commit()
        log_operation("API 停用教师", "teacher", teacher_id, teacher["name"])
        return jsonify({"ok": True})

    @app.post("/api/classes")
    @require_roles("admin", "staff")
    def api_create_class():
        data = request.get_json(silent=True) or {}
        if error := validate_class_payload(data):
            return api_error(error)
        db = get_db()
        cursor = db.execute(
            """
            INSERT INTO classes
                (class_name, course_name, class_type, teacher_id, default_weekday,
                 default_start_time, default_end_time, capacity, start_date, end_date,
                 status, remark, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                data.get("class_name", "").strip(),
                data.get("course_name", "").strip(),
                data.get("class_type", "").strip(),
                parse_int(data.get("teacher_id")),
                parse_int(data.get("default_weekday")),
                data.get("default_start_time", "").strip(),
                data.get("default_end_time", "").strip(),
                parse_int(data.get("capacity")),
                data.get("start_date", "").strip(),
                data.get("end_date", "").strip(),
                data.get("status", "active"),
                data.get("remark", "").strip(),
                now_text(),
            ),
        )
        db.commit()
        log_operation("API 新建班级", "class", cursor.lastrowid, data.get("class_name", "").strip())
        return jsonify({"ok": True, "id": cursor.lastrowid})

    @app.put("/api/classes/<int:class_id>")
    @require_roles("admin", "staff")
    def api_update_class(class_id: int):
        data = request.get_json(silent=True) or {}
        class_row = fetch_one("SELECT * FROM classes WHERE id = ?", (class_id,))
        if class_row is None:
            abort(404)
        if error := validate_class_payload(data):
            return api_error(error)
        get_db().execute(
            """
            UPDATE classes
            SET class_name = ?, course_name = ?, class_type = ?, teacher_id = ?,
                default_weekday = ?, default_start_time = ?, default_end_time = ?,
                capacity = ?, start_date = ?, end_date = ?, status = ?, remark = ?
            WHERE id = ?
            """,
            (
                data.get("class_name", "").strip(),
                data.get("course_name", "").strip(),
                data.get("class_type", "").strip(),
                parse_int(data.get("teacher_id")),
                parse_int(data.get("default_weekday")),
                data.get("default_start_time", "").strip(),
                data.get("default_end_time", "").strip(),
                parse_int(data.get("capacity")),
                data.get("start_date", "").strip(),
                data.get("end_date", "").strip(),
                data.get("status", "active"),
                data.get("remark", "").strip(),
                class_id,
            ),
        )
        get_db().commit()
        log_operation("API 编辑班级", "class", class_id, class_row["class_name"])
        return jsonify({"ok": True})

    @app.post("/api/classes/<int:class_id>/archive")
    @require_roles("admin", "staff")
    def api_archive_class(class_id: int):
        class_row = fetch_one("SELECT * FROM classes WHERE id = ?", (class_id,))
        if class_row is None:
            abort(404)
        get_db().execute("UPDATE classes SET status = 'archived' WHERE id = ?", (class_id,))
        get_db().commit()
        log_operation("API 归档班级", "class", class_id, class_row["class_name"])
        return jsonify({"ok": True})

    @app.post("/api/classes/<int:class_id>/lessons/generate")
    @require_roles("admin", "staff")
    def api_generate_class_lessons(class_id: int):
        class_row = fetch_one("SELECT * FROM classes WHERE id = ?", (class_id,))
        if class_row is None:
            abort(404)
        data = request.get_json(silent=True) or {}
        start = parse_date(data.get("start_date"))
        end = parse_date(data.get("end_date"))
        weekday = parse_int(data.get("weekday"))
        preset = get_course_preset(parse_int(data.get("course_preset_id")))
        lesson_topic = (data.get("lesson_topic") or "").strip() or (
            preset["course_name"] if preset else ""
        )
        if error := first_error(
            "请选择开始日期。" if start is None else None,
            "请选择结束日期。" if end is None else None,
            validate_date_range_values(
                (data.get("start_date") or "").strip(),
                (data.get("end_date") or "").strip(),
            ),
            "请选择上课星期。" if weekday not in range(1, 8) else None,
            required_text(data, "start_time", "开始时间"),
            required_text(data, "end_time", "结束时间"),
            validate_time_range_values(
                (data.get("start_time") or "").strip(),
                (data.get("end_time") or "").strip(),
            ),
            "请选择上课老师。" if resolve_lesson_teacher_id(data.get("teacher_id")) is None else None,
            "请选择预设课程或填写课程主题。" if not lesson_topic else None,
            validate_positive_number(
                {"lesson_hours": data.get("lesson_hours") or (preset["default_hours"] if preset else None)},
                "lesson_hours",
                "课时数量",
            ),
        ):
            return api_error(error)
        lesson_type_id = parse_int(data.get("lesson_type_id")) or (
            preset["lesson_type_id"] if preset else None
        )
        lesson_type = fetch_one("SELECT * FROM lesson_types WHERE id = ?", (lesson_type_id,))
        lesson_hours = parse_float(
            data.get("lesson_hours"),
            float(preset["default_hours"] or 1) if preset else 1,
        )
        if lesson_hours <= 0 and lesson_type:
            lesson_hours = float(lesson_type["default_hours"] or 1)
        teacher_id = resolve_lesson_teacher_id(data.get("teacher_id"))
        created_count = 0
        skipped_count = 0
        current = start
        while current <= end:
            if current.isoweekday() == weekday:
                exists = fetch_one(
                    """
                    SELECT id FROM lessons
                    WHERE class_id = ? AND lesson_date = ? AND start_time = ?
                    """,
                    (class_id, current.isoformat(), (data.get("start_time") or "").strip()),
                )
                if exists is None:
                    get_db().execute(
                        """
                        INSERT INTO lessons
                            (class_id, lesson_date, weekday, start_time, end_time,
                             teacher_id, lesson_type_id, course_preset_id, lesson_topic,
                             lesson_hours, classroom, status, remark, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'planned', ?, ?)
                        """,
                        (
                            class_id,
                            current.isoformat(),
                            weekday,
                            (data.get("start_time") or "").strip(),
                            (data.get("end_time") or "").strip(),
                            teacher_id,
                            lesson_type_id,
                            preset["id"] if preset else None,
                            lesson_topic,
                            lesson_hours,
                            (data.get("classroom") or "").strip(),
                            (data.get("remark") or "").strip(),
                            now_text(),
                        ),
                    )
                    created_count += 1
                else:
                    skipped_count += 1
            current += timedelta(days=1)
        get_db().commit()
        log_operation(
            "API 批量生成课次",
            "class",
            class_id,
            f"{class_row['class_name']} 新增 {created_count} 节，跳过 {skipped_count} 节",
        )
        return jsonify({"ok": True, "created": created_count, "skipped": skipped_count})

    @app.get("/api/classes/<int:class_id>/students")
    def api_class_students(class_id: int):
        class_row = fetch_one("SELECT * FROM classes WHERE id = ?", (class_id,))
        if class_row is None:
            abort(404)
        ensure_class_access(class_row)
        members = fetch_all(
            """
            SELECT cs.*, s.name, s.phone, s.parent_phone, s.school
            FROM class_students cs
            JOIN students s ON s.id = cs.student_id
            WHERE cs.class_id = ? AND cs.status = 'active'
            ORDER BY s.name
            """,
            (class_id,),
        )
        return jsonify({"items": rows_to_dicts(members)})

    @app.post("/api/classes/<int:class_id>/students")
    @require_roles("admin", "staff")
    def api_add_class_student(class_id: int):
        if fetch_one("SELECT id FROM classes WHERE id = ?", (class_id,)) is None:
            abort(404)
        data = request.get_json(silent=True) or {}
        student_id = parse_int(data.get("student_id"))
        if not student_id:
            return jsonify({"ok": False, "message": "请选择学生。"}), 400
        get_db().execute(
            """
            INSERT INTO class_students (class_id, student_id, join_date, leave_date, status)
            VALUES (?, ?, ?, NULL, 'active')
            ON CONFLICT(class_id, student_id) DO UPDATE SET
                join_date = excluded.join_date,
                leave_date = NULL,
                status = 'active'
            """,
            (class_id, student_id, data.get("join_date") or date.today().isoformat()),
        )
        get_db().commit()
        log_operation("API 学生加入班级", "class", class_id, f"student_id={student_id}")
        return jsonify({"ok": True})

    @app.delete("/api/classes/<int:class_id>/students/<int:student_id>")
    @require_roles("admin", "staff")
    def api_remove_class_student(class_id: int, student_id: int):
        get_db().execute(
            """
            UPDATE class_students
            SET status = 'inactive', leave_date = ?
            WHERE class_id = ? AND student_id = ?
            """,
            (date.today().isoformat(), class_id, student_id),
        )
        get_db().commit()
        log_operation("API 学生移出班级", "class", class_id, f"student_id={student_id}")
        return jsonify({"ok": True})

    @app.post("/api/lesson-types")
    @require_roles("admin", "staff")
    def api_create_lesson_type():
        data = request.get_json(silent=True) or {}
        if error := validate_lesson_type_payload(data):
            return api_error(error)
        try:
            db = get_db()
            cursor = db.execute(
                """
                INSERT INTO lesson_types
                    (type_name, default_hours, count_in_statistics, remark, created_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    data.get("type_name", "").strip(),
                    parse_float(data.get("default_hours"), 1),
                    1 if data.get("count_in_statistics", True) else 0,
                    data.get("remark", "").strip(),
                    now_text(),
                ),
            )
            db.commit()
        except DatabaseIntegrityError:
            return jsonify({"ok": False, "message": "课程类型名称已存在。"}), 400
        log_operation("API 新增课程类型", "lesson_type", cursor.lastrowid, data.get("type_name", "").strip())
        return jsonify({"ok": True, "id": cursor.lastrowid})

    @app.put("/api/lesson-types/<int:type_id>")
    @require_roles("admin", "staff")
    def api_update_lesson_type(type_id: int):
        data = request.get_json(silent=True) or {}
        lesson_type = fetch_one("SELECT * FROM lesson_types WHERE id = ?", (type_id,))
        if lesson_type is None:
            abort(404)
        if error := validate_lesson_type_payload(data):
            return api_error(error)
        try:
            get_db().execute(
                """
                UPDATE lesson_types
                SET type_name = ?, default_hours = ?, count_in_statistics = ?, remark = ?
                WHERE id = ?
                """,
                (
                    data.get("type_name", "").strip(),
                    parse_float(data.get("default_hours"), 1),
                    1 if data.get("count_in_statistics", True) else 0,
                    data.get("remark", "").strip(),
                    type_id,
                ),
            )
            get_db().commit()
        except DatabaseIntegrityError:
            return jsonify({"ok": False, "message": "课程类型名称已存在。"}), 400
        log_operation("API 编辑课程类型", "lesson_type", type_id, lesson_type["type_name"])
        return jsonify({"ok": True})

    @app.get("/api/lessons/<int:lesson_id>/attendance")
    def api_lesson_attendance(lesson_id: int):
        lesson = get_lesson(lesson_id)
        if lesson is None:
            abort(404)
        ensure_lesson_access(lesson)
        rows = query_lesson_attendance(lesson_id, lesson["class_id"])
        return jsonify(
            {
                "lesson": dict(lesson),
                "items": rows_to_dicts(rows),
                "summary": summarize_attendance(rows),
                "statuses": ATTENDANCE_STATUS,
            }
        )

    @app.post("/api/lessons/<int:lesson_id>/attendance")
    def api_save_attendance(lesson_id: int):
        lesson = get_lesson(lesson_id)
        if lesson is None:
            abort(404)
        ensure_lesson_access(lesson)
        data = request.get_json(silent=True) or {}
        records = data.get("records") or []
        operator = (data.get("operator") or g.current_user["display_name"] or "系统").strip()
        save_attendance_records(lesson, records, operator)
        log_operation("API 保存签到", "lesson", lesson_id, lesson["class_name"])
        return jsonify({"ok": True})

    @app.get("/api/lessons/<int:lesson_id>/detail")
    def api_lesson_detail(lesson_id: int):
        lesson = get_lesson(lesson_id)
        if lesson is None:
            abort(404)
        ensure_lesson_access(lesson)
        return jsonify(
            {
                "lesson": dict(lesson),
                "detail": get_lesson_detail(lesson_id),
                "courseware_assets": rows_to_dicts(query_lesson_assets(lesson_id)),
                "scratch_assignments": rows_to_dicts(query_lesson_scratch_templates(lesson_id)),
                "scratch_works": rows_to_dicts(query_scratch_works({"lesson_id": lesson_id})),
                "scratch_templates": rows_to_dicts(query_scratch_templates({"status": "active"})),
            }
        )

    @app.put("/api/lessons/<int:lesson_id>/detail")
    @require_roles("admin", "staff", "teacher")
    def api_save_lesson_detail(lesson_id: int):
        lesson = get_lesson(lesson_id)
        if lesson is None:
            abort(404)
        ensure_lesson_access(lesson)
        data = request.get_json(silent=True) or {}
        if error := validate_lesson_detail_payload(data):
            return api_error(error)
        operator = (g.current_user["display_name"] or g.current_user["username"] or "系统").strip()
        detail = save_lesson_detail(lesson_id, data, operator)
        log_operation("API 保存课程详情", "lesson", lesson_id, lesson["class_name"])
        return jsonify({"ok": True, "detail": detail})

    @app.get("/api/users")
    @require_roles("admin")
    def api_users():
        rows = fetch_all(
            """
            SELECT u.id, u.username, u.display_name, u.role, u.status,
                   u.teacher_id, u.student_id, u.remark, u.last_login_at,
                   u.created_at, t.name AS teacher_name, s.name AS student_name
            FROM users u
            LEFT JOIN teachers t ON t.id = u.teacher_id
            LEFT JOIN students s ON s.id = u.student_id
            ORDER BY u.created_at DESC, u.id DESC
            """
        )
        return jsonify({"items": [public_account(row) for row in rows]})

    @app.post("/api/users")
    @require_roles("admin")
    def api_create_user():
        data = request.get_json(silent=True) or {}
        password = data.get("password", "").strip()
        if error := validate_user_payload(data, require_password=True):
            return api_error(error)
        try:
            db = get_db()
            cursor = db.execute(
                """
                INSERT INTO users
                    (username, password_hash, display_name, role, status, teacher_id,
                     student_id, remark, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    data.get("username", "").strip(),
                    generate_password_hash(password),
                    data.get("display_name", "").strip(),
                    normalize_role(data.get("role", "academic_manager")),
                    data.get("status", "active"),
                    parse_int(data.get("teacher_id")),
                    parse_int(data.get("student_id")),
                    data.get("remark", "").strip(),
                    now_text(),
                ),
            )
            db.commit()
        except DatabaseIntegrityError:
            return jsonify({"ok": False, "message": "用户名已存在。"}), 400
        log_operation("API 新增账号", "user", cursor.lastrowid, data.get("username", "").strip())
        return jsonify({"ok": True, "id": cursor.lastrowid})

    @app.put("/api/users/<int:user_id>")
    @require_roles("admin")
    def api_update_user(user_id: int):
        data = request.get_json(silent=True) or {}
        user = fetch_one("SELECT * FROM users WHERE id = ?", (user_id,))
        if user is None:
            abort(404)
        if error := validate_user_payload(data, require_password=False):
            return api_error(error)
        get_db().execute(
            """
            UPDATE users
            SET display_name = ?, role = ?, status = ?, teacher_id = ?, student_id = ?, remark = ?
            WHERE id = ?
            """,
            (
                data.get("display_name", "").strip(),
                normalize_role(data.get("role", "academic_manager")),
                data.get("status", "active"),
                parse_int(data.get("teacher_id")),
                parse_int(data.get("student_id")),
                data.get("remark", "").strip(),
                user_id,
            ),
        )
        password = data.get("password", "").strip()
        if password:
            get_db().execute(
                "UPDATE users SET password_hash = ? WHERE id = ?",
                (generate_password_hash(password), user_id),
            )
        get_db().commit()
        log_operation("API 编辑账号", "user", user_id, user["username"])
        return jsonify({"ok": True})

    @app.get("/api/logs")
    @require_roles("admin", "staff")
    def api_logs():
        filters = {
            "q": request.args.get("q", "").strip(),
            "start_date": request.args.get("start_date", "").strip(),
            "end_date": request.args.get("end_date", "").strip(),
        }
        return jsonify({"items": rows_to_dicts(query_operation_logs(filters))})

    @app.get("/api/logs/export")
    @require_roles("admin", "staff")
    def api_export_logs():
        filters = {
            "q": request.args.get("q", "").strip(),
            "start_date": request.args.get("start_date", "").strip(),
            "end_date": request.args.get("end_date", "").strip(),
        }
        rows = query_operation_logs(filters, limit=10000)
        file_path = export_operation_logs_workbook(rows, filters)
        return send_file(file_path, as_attachment=True, download_name=file_path.name)

    @app.route("/users", methods=["GET", "POST"])
    @require_roles("admin")
    def users():
        if request.method == "POST":
            form = request.form
            password = form.get("password", "").strip()
            if not password:
                flash("请设置初始密码。", "error")
                return redirect(url_for("users"))
            if error := validate_user_payload(form, require_password=True):
                flash(error, "error")
                return redirect(url_for("users"))
            try:
                get_db().execute(
                    """
                    INSERT INTO users
                        (username, password_hash, display_name, role, status, teacher_id,
                         student_id, remark, created_at)
                    VALUES (?, ?, ?, ?, 'active', ?, ?, ?, ?)
                    """,
                    (
                        form.get("username", "").strip(),
                        generate_password_hash(password),
                        form.get("display_name", "").strip(),
                        normalize_role(form.get("role", "academic_manager")),
                        parse_int(form.get("teacher_id")),
                        parse_int(form.get("student_id")),
                        form.get("remark", "").strip(),
                        now_text(),
                    ),
                )
                get_db().commit()
                log_operation("新增账号", "user", None, form.get("username", "").strip())
                flash("账号已创建。", "success")
            except DatabaseIntegrityError:
                flash("用户名已存在，请换一个。", "error")
            return redirect(url_for("users"))

        rows = fetch_all(
            """
            SELECT u.*, t.name AS teacher_name, s.name AS student_name,
                   rd.label AS role_label
            FROM users u
            LEFT JOIN teachers t ON t.id = u.teacher_id
            LEFT JOIN students s ON s.id = u.student_id
            LEFT JOIN role_definitions rd
              ON rd.role_code = CASE
                  WHEN u.role = 'admin' THEN 'super_admin'
                  WHEN u.role = 'staff' THEN 'academic_manager'
                  ELSE u.role
              END
            ORDER BY u.created_at DESC, u.id DESC
            """
        )
        return render_template("users.html", users=rows, teachers=get_active_teachers(), students=get_active_students())

    @app.route("/users/<int:user_id>/edit", methods=["GET", "POST"])
    @require_roles("admin")
    def edit_user(user_id: int):
        user = fetch_one("SELECT * FROM users WHERE id = ?", (user_id,))
        if user is None:
            abort(404)
        if request.method == "POST":
            form = request.form
            if error := validate_user_payload(form, require_password=False):
                flash(error, "error")
                return redirect(url_for("edit_user", user_id=user_id))
            params = [
                form.get("display_name", "").strip(),
                normalize_role(form.get("role", "academic_manager")),
                form.get("status", "active"),
                parse_int(form.get("teacher_id")),
                parse_int(form.get("student_id")),
                form.get("remark", "").strip(),
                user_id,
            ]
            get_db().execute(
                """
                UPDATE users
                SET display_name = ?, role = ?, status = ?, teacher_id = ?, student_id = ?, remark = ?
                WHERE id = ?
                """,
                tuple(params),
            )
            password = form.get("password", "").strip()
            if password:
                get_db().execute(
                    "UPDATE users SET password_hash = ? WHERE id = ?",
                    (generate_password_hash(password), user_id),
                )
            get_db().commit()
            log_operation("编辑账号", "user", user_id, user["username"])
            flash("账号已更新。", "success")
            return redirect(url_for("users"))
        return render_template("user_form.html", user=user, teachers=get_active_teachers(), students=get_active_students())

    @app.get("/logs")
    @require_roles("admin", "staff")
    def logs():
        filters = {
            "q": request.args.get("q", "").strip(),
            "start_date": request.args.get("start_date", "").strip(),
            "end_date": request.args.get("end_date", "").strip(),
        }
        rows = query_operation_logs(filters)
        return render_template("logs.html", logs=rows, filters=filters)

    @app.get("/exports/logs")
    @require_roles("admin", "staff")
    def export_logs():
        filters = {
            "q": request.args.get("q", "").strip(),
            "start_date": request.args.get("start_date", "").strip(),
            "end_date": request.args.get("end_date", "").strip(),
        }
        rows = query_operation_logs(filters, limit=10000)
        file_path = export_operation_logs_workbook(rows, filters)
        return send_file(file_path, as_attachment=True, download_name=file_path.name)

    @app.get("/")
    def dashboard():
        today_value = date.today().isoformat()
        lesson_filters = apply_teacher_scope({"date": today_value})
        lessons = query_lessons(lesson_filters)
        teacher_filters = apply_teacher_scope({})
        status_counts = count_attendance_by_status(today_value, today_value, teacher_filters)
        expected_total = sum(row["expected_count"] for row in lessons)
        arrived_total = sum(
            status_counts.get(status, 0) for status in ("已到", "迟到", "补签", "试听")
        )
        summary = {
            "lesson_count": len(lessons),
            "pending_count": sum(1 for row in lessons if row["status"] == "planned"),
            "completed_count": sum(1 for row in lessons if row["status"] == "completed"),
            "expected_total": expected_total,
            "arrived_total": arrived_total,
            "leave_total": status_counts.get("请假", 0),
            "absent_total": status_counts.get("缺勤", 0),
        }
        alerts = {
            "low_hours": query_low_hour_students(),
            "consecutive_absences": query_consecutive_absences(),
            "recent_logs": query_operation_logs({}, limit=8)
            if has_permission(g.current_user, "logs.view")
            else [],
        }
        upcoming_lessons = query_upcoming_lessons(today_value, filters=teacher_filters)
        return render_template(
            "dashboard.html",
            lessons=lessons,
            summary=summary,
            alerts=alerts,
            upcoming_lessons=upcoming_lessons,
        )

    @app.route("/students", methods=["GET", "POST"])
    def students():
        if request.method == "POST":
            ensure_roles("admin", "staff")
            form = request.form
            get_db().execute(
                """
                INSERT INTO students
                    (name, gender, age, phone, parent_name, parent_phone, school,
                     purchased_hours, gift_hours, status, remark, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'active', ?, ?)
                """,
                (
                    form.get("name", "").strip(),
                    form.get("gender", "").strip(),
                    parse_int(form.get("age")),
                    form.get("phone", "").strip(),
                    form.get("parent_name", "").strip(),
                    form.get("parent_phone", "").strip(),
                    form.get("school", "").strip(),
                    parse_float(form.get("purchased_hours")),
                    parse_float(form.get("gift_hours")),
                    form.get("remark", "").strip(),
                    now_text(),
                ),
            )
            get_db().commit()
            log_operation("新增学生", "student", None, form.get("name", "").strip())
            flash("学生已添加。", "success")
            return redirect(url_for("students"))

        filters = {
            "q": request.args.get("q", "").strip(),
            "class_id": request.args.get("class_id", "").strip(),
            "status": request.args.get("status", "active").strip(),
        }
        filters = apply_teacher_scope(filters)
        student_rows = query_students(filters)
        class_rows = get_active_classes()
        return render_template(
            "students.html",
            students=student_rows,
            classes=class_rows,
            filters=filters,
        )

    @app.route("/students/<int:student_id>/edit", methods=["GET", "POST"])
    @require_roles("admin", "staff")
    def edit_student(student_id: int):
        student = fetch_one("SELECT * FROM students WHERE id = ?", (student_id,))
        if student is None:
            abort(404)
        if request.method == "POST":
            form = request.form
            get_db().execute(
                """
                UPDATE students
                SET name = ?, gender = ?, age = ?, phone = ?, parent_name = ?,
                    parent_phone = ?, school = ?, purchased_hours = ?, gift_hours = ?,
                    status = ?, remark = ?
                WHERE id = ?
                """,
                (
                    form.get("name", "").strip(),
                    form.get("gender", "").strip(),
                    parse_int(form.get("age")),
                    form.get("phone", "").strip(),
                    form.get("parent_name", "").strip(),
                    form.get("parent_phone", "").strip(),
                    form.get("school", "").strip(),
                    parse_float(form.get("purchased_hours")),
                    parse_float(form.get("gift_hours")),
                    form.get("status", "active"),
                    form.get("remark", "").strip(),
                    student_id,
                ),
            )
            get_db().commit()
            log_operation("编辑学生", "student", student_id, student["name"])
            flash("学生信息已更新。", "success")
            return redirect(url_for("students"))
        return render_template("student_form.html", student=student)

    @app.post("/students/<int:student_id>/deactivate")
    @require_roles("admin", "staff")
    def deactivate_student(student_id: int):
        get_db().execute("UPDATE students SET status = 'inactive' WHERE id = ?", (student_id,))
        get_db().commit()
        log_operation("停用学生", "student", student_id)
        flash("学生已停用，历史记录仍会保留。", "success")
        return redirect(url_for("students"))

    @app.get("/students/<int:student_id>/history")
    def student_history(student_id: int):
        student = fetch_one("SELECT * FROM students WHERE id = ?", (student_id,))
        if student is None:
            abort(404)
        ensure_student_access(student_id)
        rows = fetch_all(
            """
            SELECT a.*, l.lesson_date, l.start_time, l.end_time, l.lesson_topic,
                   c.class_name, t.name AS teacher_name, lt.type_name
            FROM attendance a
            JOIN lessons l ON l.id = a.lesson_id
            JOIN classes c ON c.id = l.class_id
            LEFT JOIN teachers t ON t.id = l.teacher_id
            LEFT JOIN lesson_types lt ON lt.id = l.lesson_type_id
            WHERE a.student_id = ?
            ORDER BY l.lesson_date DESC, l.start_time DESC
            """,
            (student_id,),
        )
        consumed = sum(float(row["deduct_hours"] or 0) for row in rows)
        status_counts = defaultdict(int)
        for row in rows:
            status_counts[row["status"]] += 1
        total_hours = float(student["purchased_hours"] or 0) + float(student["gift_hours"] or 0)
        summary = {
            "total_hours": total_hours,
            "consumed": consumed,
            "remaining": total_hours - consumed,
            "status_counts": status_counts,
            "last_lesson": rows[0]["lesson_date"] if rows else "-",
        }
        return render_template(
            "student_history.html", student=student, rows=rows, summary=summary
        )

    @app.route("/teachers", methods=["GET", "POST"])
    @require_roles("admin", "staff")
    def teachers():
        if request.method == "POST":
            form = request.form
            get_db().execute(
                """
                INSERT INTO teachers (name, phone, subject, status, remark, created_at)
                VALUES (?, ?, ?, 'active', ?, ?)
                """,
                (
                    form.get("name", "").strip(),
                    form.get("phone", "").strip(),
                    form.get("subject", "").strip(),
                    form.get("remark", "").strip(),
                    now_text(),
                ),
            )
            get_db().commit()
            log_operation("新增教师", "teacher", None, form.get("name", "").strip())
            flash("教师已添加。", "success")
            return redirect(url_for("teachers"))

        filters = {
            "q": request.args.get("q", "").strip(),
            "status": request.args.get("status", "active").strip(),
        }
        rows = query_teachers(filters)
        return render_template("teachers.html", teachers=rows, filters=filters)

    @app.route("/teachers/<int:teacher_id>/edit", methods=["GET", "POST"])
    @require_roles("admin", "staff")
    def edit_teacher(teacher_id: int):
        teacher = fetch_one("SELECT * FROM teachers WHERE id = ?", (teacher_id,))
        if teacher is None:
            abort(404)
        if request.method == "POST":
            form = request.form
            get_db().execute(
                """
                UPDATE teachers
                SET name = ?, phone = ?, subject = ?, status = ?, remark = ?
                WHERE id = ?
                """,
                (
                    form.get("name", "").strip(),
                    form.get("phone", "").strip(),
                    form.get("subject", "").strip(),
                    form.get("status", "active"),
                    form.get("remark", "").strip(),
                    teacher_id,
                ),
            )
            get_db().commit()
            log_operation("编辑教师", "teacher", teacher_id, teacher["name"])
            flash("教师信息已更新。", "success")
            return redirect(url_for("teachers"))
        return render_template("teacher_form.html", teacher=teacher)

    @app.post("/teachers/<int:teacher_id>/deactivate")
    @require_roles("admin", "staff")
    def deactivate_teacher(teacher_id: int):
        get_db().execute("UPDATE teachers SET status = 'inactive' WHERE id = ?", (teacher_id,))
        get_db().commit()
        log_operation("停用教师", "teacher", teacher_id)
        flash("教师已停用。", "success")
        return redirect(url_for("teachers"))

    @app.route("/lesson-types", methods=["GET", "POST"])
    @require_roles("admin", "staff")
    def lesson_types():
        if request.method == "POST":
            form = request.form
            get_db().execute(
                """
                INSERT INTO lesson_types
                    (type_name, default_hours, count_in_statistics, remark, created_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    form.get("type_name", "").strip(),
                    parse_float(form.get("default_hours"), 1),
                    1 if form.get("count_in_statistics") == "on" else 0,
                    form.get("remark", "").strip(),
                    now_text(),
                ),
            )
            get_db().commit()
            log_operation("新增课程类型", "lesson_type", None, form.get("type_name", "").strip())
            flash("课程类型已添加。", "success")
            return redirect(url_for("lesson_types"))
        rows = fetch_all("SELECT * FROM lesson_types ORDER BY id")
        return render_template("lesson_types.html", lesson_types=rows)

    @app.route("/lesson-types/<int:type_id>/edit", methods=["GET", "POST"])
    @require_roles("admin", "staff")
    def edit_lesson_type(type_id: int):
        lesson_type = fetch_one("SELECT * FROM lesson_types WHERE id = ?", (type_id,))
        if lesson_type is None:
            abort(404)
        if request.method == "POST":
            form = request.form
            get_db().execute(
                """
                UPDATE lesson_types
                SET type_name = ?, default_hours = ?, count_in_statistics = ?, remark = ?
                WHERE id = ?
                """,
                (
                    form.get("type_name", "").strip(),
                    parse_float(form.get("default_hours"), 1),
                    1 if form.get("count_in_statistics") == "on" else 0,
                    form.get("remark", "").strip(),
                    type_id,
                ),
            )
            get_db().commit()
            log_operation("编辑课程类型", "lesson_type", type_id, lesson_type["type_name"])
            flash("课程类型已更新。", "success")
            return redirect(url_for("lesson_types"))
        return render_template("lesson_type_form.html", lesson_type=lesson_type)

    @app.route("/course-presets", methods=["GET", "POST"])
    @require_roles("admin", "staff")
    def course_presets():
        if request.method == "POST":
            form = request.form
            try:
                db = get_db()
                cursor = db.execute(
                    """
                    INSERT INTO course_presets
                        (category, stage, lesson_no, course_name, lesson_type_id,
                         default_hours, status, remark, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    course_preset_values(form) + (now_text(),),
                )
                db.commit()
            except DatabaseIntegrityError:
                flash("同一大类、阶段、节次和课程名称已存在。", "error")
                return redirect(url_for("course_presets"))
            log_operation("新增预设课程", "course_preset", cursor.lastrowid, form.get("course_name", "").strip())
            flash("预设课程已添加。", "success")
            return redirect(url_for("course_presets"))

        filters = {
            "q": request.args.get("q", "").strip(),
            "category": request.args.get("category", "").strip(),
            "stage": request.args.get("stage", "").strip(),
            "status": request.args.get("status", "active").strip(),
        }
        rows = query_course_presets(filters)
        categories = fetch_all(
            "SELECT DISTINCT category FROM course_presets WHERE category != '' ORDER BY category"
        )
        stages = fetch_all(
            "SELECT DISTINCT stage FROM course_presets WHERE stage != '' ORDER BY stage"
        )
        return render_template(
            "course_presets.html",
            course_presets=rows,
            filters=filters,
            categories=categories,
            stages=stages,
            lesson_types=fetch_all("SELECT * FROM lesson_types ORDER BY id"),
        )

    @app.get("/course-presets/import-template")
    @require_roles("admin", "staff", "teacher")
    def download_course_preset_template():
        file_path = export_course_preset_import_template()
        return send_file(file_path, as_attachment=True, download_name=file_path.name)

    @app.post("/course-presets/import")
    @require_roles("admin", "staff")
    def import_course_presets_view():
        upload = request.files.get("file")
        if upload is None or not upload.filename:
            flash("请上传 Excel 或 CSV 文件。", "error")
            return redirect(url_for("course_presets"))
        try:
            rows = read_course_preset_import_rows(upload)
            result = import_course_presets(rows)
        except ValueError as error:
            flash(str(error), "error")
            return redirect(url_for("course_presets"))
        log_operation(
            "批量导入预设课程",
            "course_preset",
            None,
            f"新增 {result['created']}，更新 {result['updated']}，跳过 {result['skipped']}",
        )
        flash(
            f"导入完成：新增 {result['created']} 条，更新 {result['updated']} 条，跳过 {result['skipped']} 条。",
            "success",
        )
        return redirect(url_for("course_presets"))

    @app.route("/course-presets/<int:preset_id>/edit", methods=["GET", "POST"])
    @require_roles("admin", "staff")
    def edit_course_preset(preset_id: int):
        preset = fetch_one("SELECT * FROM course_presets WHERE id = ?", (preset_id,))
        if preset is None:
            abort(404)
        if request.method == "POST":
            form = request.form
            try:
                get_db().execute(
                    """
                    UPDATE course_presets
                    SET category = ?, stage = ?, lesson_no = ?, course_name = ?,
                        lesson_type_id = ?, default_hours = ?, status = ?, remark = ?
                    WHERE id = ?
                    """,
                    course_preset_values(form) + (preset_id,),
                )
                get_db().commit()
            except DatabaseIntegrityError:
                flash("同一大类、阶段、节次和课程名称已存在。", "error")
                return redirect(url_for("edit_course_preset", preset_id=preset_id))
            log_operation("编辑预设课程", "course_preset", preset_id, preset["course_name"])
            flash("预设课程已更新。", "success")
            return redirect(url_for("course_presets"))
        return render_template(
            "course_preset_form.html",
            preset=preset,
            lesson_types=fetch_all("SELECT * FROM lesson_types ORDER BY id"),
        )

    @app.route("/classes", methods=["GET", "POST"])
    def classes():
        if request.method == "POST":
            ensure_roles("admin", "staff")
            form = request.form
            get_db().execute(
                """
                INSERT INTO classes
                    (class_name, course_name, class_type, teacher_id, default_weekday,
                     default_start_time, default_end_time, capacity, start_date, end_date,
                     status, remark, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    form.get("class_name", "").strip(),
                    form.get("course_name", "").strip(),
                    form.get("class_type", "").strip(),
                    parse_int(form.get("teacher_id")),
                    parse_int(form.get("default_weekday")),
                    form.get("default_start_time", "").strip(),
                    form.get("default_end_time", "").strip(),
                    parse_int(form.get("capacity")),
                    form.get("start_date", "").strip(),
                    form.get("end_date", "").strip(),
                    form.get("status", "active"),
                    form.get("remark", "").strip(),
                    now_text(),
                ),
            )
            get_db().commit()
            log_operation("新建班级", "class", None, form.get("class_name", "").strip())
            flash("班级已创建。", "success")
            return redirect(url_for("classes"))

        filters = {
            "q": request.args.get("q", "").strip(),
            "status": request.args.get("status", "active").strip(),
        }
        filters = apply_teacher_scope(filters)
        rows = query_classes(filters)
        teachers = get_active_teachers()
        return render_template("classes.html", classes=rows, teachers=teachers, filters=filters)

    @app.route("/classes/<int:class_id>/edit", methods=["GET", "POST"])
    @require_roles("admin", "staff")
    def edit_class(class_id: int):
        class_row = fetch_one("SELECT * FROM classes WHERE id = ?", (class_id,))
        if class_row is None:
            abort(404)
        if request.method == "POST":
            form = request.form
            get_db().execute(
                """
                UPDATE classes
                SET class_name = ?, course_name = ?, class_type = ?, teacher_id = ?,
                    default_weekday = ?, default_start_time = ?, default_end_time = ?,
                    capacity = ?, start_date = ?, end_date = ?, status = ?, remark = ?
                WHERE id = ?
                """,
                (
                    form.get("class_name", "").strip(),
                    form.get("course_name", "").strip(),
                    form.get("class_type", "").strip(),
                    parse_int(form.get("teacher_id")),
                    parse_int(form.get("default_weekday")),
                    form.get("default_start_time", "").strip(),
                    form.get("default_end_time", "").strip(),
                    parse_int(form.get("capacity")),
                    form.get("start_date", "").strip(),
                    form.get("end_date", "").strip(),
                    form.get("status", "active"),
                    form.get("remark", "").strip(),
                    class_id,
                ),
            )
            get_db().commit()
            log_operation("编辑班级", "class", class_id, class_row["class_name"])
            flash("班级信息已更新。", "success")
            return redirect(url_for("classes"))
        return render_template("class_form.html", class_row=class_row, teachers=get_active_teachers())

    @app.post("/classes/<int:class_id>/archive")
    @require_roles("admin", "staff")
    def archive_class(class_id: int):
        get_db().execute("UPDATE classes SET status = 'archived' WHERE id = ?", (class_id,))
        get_db().commit()
        log_operation("归档班级", "class", class_id)
        flash("班级已归档。", "success")
        return redirect(url_for("classes"))

    @app.route("/classes/<int:class_id>/students", methods=["GET", "POST"])
    def class_students(class_id: int):
        class_row = fetch_one("SELECT * FROM classes WHERE id = ?", (class_id,))
        if class_row is None:
            abort(404)
        ensure_class_access(class_row)
        if request.method == "POST":
            ensure_roles("admin", "staff")
            student_id = parse_int(request.form.get("student_id"))
            if student_id:
                get_db().execute(
                    """
                    INSERT INTO class_students
                        (class_id, student_id, join_date, leave_date, status)
                    VALUES (?, ?, ?, NULL, 'active')
                    ON CONFLICT(class_id, student_id) DO UPDATE SET
                        join_date = excluded.join_date,
                        leave_date = NULL,
                        status = 'active'
                    """,
                    (class_id, student_id, request.form.get("join_date") or date.today().isoformat()),
                )
                get_db().commit()
                log_operation("学生加入班级", "class", class_id, f"student_id={student_id}")
                flash("学生已加入班级。", "success")
            return redirect(url_for("class_students", class_id=class_id))

        members = fetch_all(
            """
            SELECT cs.*, s.name, s.phone, s.parent_phone, s.school
            FROM class_students cs
            JOIN students s ON s.id = cs.student_id
            WHERE cs.class_id = ? AND cs.status = 'active'
            ORDER BY s.name
            """,
            (class_id,),
        )
        available = []
        if has_permission(g.current_user, "class_roster.manage"):
            available = fetch_all(
                """
                SELECT s.*
                FROM students s
                WHERE s.status = 'active'
                  AND NOT EXISTS (
                      SELECT 1 FROM class_students cs
                      WHERE cs.class_id = ? AND cs.student_id = s.id AND cs.status = 'active'
                  )
                ORDER BY s.name
                """,
                (class_id,),
            )
        return render_template(
            "class_students.html", class_row=class_row, members=members, available=available
        )

    @app.post("/classes/<int:class_id>/students/<int:student_id>/remove")
    @require_roles("admin", "staff")
    def remove_class_student(class_id: int, student_id: int):
        get_db().execute(
            """
            UPDATE class_students
            SET status = 'inactive', leave_date = ?
            WHERE class_id = ? AND student_id = ?
            """,
            (date.today().isoformat(), class_id, student_id),
        )
        get_db().commit()
        log_operation("学生移出班级", "class", class_id, f"student_id={student_id}")
        flash("学生已从班级移出，历史签到不受影响。", "success")
        return redirect(url_for("class_students", class_id=class_id))

    @app.route("/classes/<int:class_id>/generate", methods=["GET", "POST"])
    @require_roles("admin", "staff")
    def generate_lessons(class_id: int):
        class_row = fetch_one("SELECT * FROM classes WHERE id = ?", (class_id,))
        if class_row is None:
            abort(404)
        lesson_types_rows = fetch_all("SELECT * FROM lesson_types ORDER BY id")
        if request.method == "POST":
            form = request.form
            start = parse_date(form.get("start_date"))
            end = parse_date(form.get("end_date"))
            weekday = parse_int(form.get("weekday"))
            if not start or not end or not weekday:
                flash("请填写开始日期、结束日期和星期。", "error")
                return redirect(url_for("generate_lessons", class_id=class_id))
            if end < start:
                flash("结束日期不能早于开始日期。", "error")
                return redirect(url_for("generate_lessons", class_id=class_id))

            preset = get_course_preset(parse_int(form.get("course_preset_id")))
            lesson_type_id = parse_int(form.get("lesson_type_id")) or (
                preset["lesson_type_id"] if preset else None
            )
            lesson_type = fetch_one(
                "SELECT * FROM lesson_types WHERE id = ?", (lesson_type_id,)
            )
            lesson_hours = parse_float(
                form.get("lesson_hours"),
                float(preset["default_hours"] or 1) if preset else 1,
            )
            if lesson_hours <= 0 and lesson_type:
                lesson_hours = float(lesson_type["default_hours"] or 1)
            lesson_topic = form.get("lesson_topic", "").strip() or (
                preset["course_name"] if preset else ""
            )
            teacher_id = resolve_lesson_teacher_id(form.get("teacher_id"))

            created_count = 0
            current = start
            while current <= end:
                if current.isoweekday() == weekday:
                    exists = fetch_one(
                        """
                        SELECT id FROM lessons
                        WHERE class_id = ? AND lesson_date = ? AND start_time = ?
                        """,
                        (class_id, current.isoformat(), form.get("start_time")),
                    )
                    if exists is None:
                        get_db().execute(
                            """
                            INSERT INTO lessons
                                (class_id, lesson_date, weekday, start_time, end_time,
                                 teacher_id, lesson_type_id, course_preset_id, lesson_topic,
                                 lesson_hours, classroom, status, remark, created_at)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'planned', ?, ?)
                            """,
                            (
                                class_id,
                                current.isoformat(),
                                weekday,
                                form.get("start_time", "").strip(),
                                form.get("end_time", "").strip(),
                                teacher_id,
                                lesson_type_id,
                                preset["id"] if preset else None,
                                lesson_topic,
                                lesson_hours,
                                form.get("classroom", "").strip(),
                                form.get("remark", "").strip(),
                                now_text(),
                            ),
                        )
                        created_count += 1
                current += timedelta(days=1)
            get_db().commit()
            log_operation("批量生成课次", "class", class_id, f"生成 {created_count} 节")
            flash(f"已生成 {created_count} 节课次。", "success")
            return redirect(url_for("lessons", class_id=class_id))

        return render_template(
            "class_generate.html",
            class_row=class_row,
            teachers=get_active_teachers(),
            lesson_types=lesson_types_rows,
            course_presets=query_course_presets({"status": "active"}),
        )

    @app.route("/lessons", methods=["GET", "POST"])
    def lessons():
        if request.method == "POST":
            ensure_roles("admin", "staff")
            form = request.form
            lesson_date = form.get("lesson_date", "").strip()
            preset = get_course_preset(parse_int(form.get("course_preset_id")))
            lesson_type_id = parse_int(form.get("lesson_type_id")) or (
                preset["lesson_type_id"] if preset else None
            )
            lesson_topic = form.get("lesson_topic", "").strip() or (
                preset["course_name"] if preset else ""
            )
            get_db().execute(
                """
                INSERT INTO lessons
                    (class_id, lesson_date, weekday, start_time, end_time, teacher_id,
                     lesson_type_id, course_preset_id, lesson_topic, lesson_hours,
                     classroom, status, remark, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'planned', ?, ?)
                """,
                (
                    parse_int(form.get("class_id")),
                    lesson_date,
                    parse_date(lesson_date).isoweekday() if parse_date(lesson_date) else None,
                    form.get("start_time", "").strip(),
                    form.get("end_time", "").strip(),
                    resolve_lesson_teacher_id(form.get("teacher_id")),
                    lesson_type_id,
                    preset["id"] if preset else None,
                    lesson_topic,
                    parse_float(form.get("lesson_hours"), float(preset["default_hours"] or 1) if preset else 1),
                    form.get("classroom", "").strip(),
                    form.get("remark", "").strip(),
                    now_text(),
                ),
            )
            get_db().commit()
            log_operation("新增课次", "lesson", None, lesson_topic)
            flash("课次已新增。", "success")
            return redirect(url_for("lessons", date=lesson_date))

        range_info = resolve_date_range(request.args)
        filters = {
            "date": request.args.get("date", "").strip(),
            "start_date": range_info["start_date"],
            "end_date": range_info["end_date"],
            "class_id": request.args.get("class_id", "").strip(),
            "teacher_id": request.args.get("teacher_id", "").strip(),
            "lesson_type_id": request.args.get("lesson_type_id", "").strip(),
            "status": request.args.get("status", "").strip(),
        }
        filters = apply_teacher_scope(filters)
        rows = query_lessons(filters)
        return render_template(
            "lessons.html",
            lessons=rows,
            filters=filters,
            classes=get_active_classes(),
            teachers=get_active_teachers(),
            lesson_types=fetch_all("SELECT * FROM lesson_types ORDER BY id"),
            course_presets=query_course_presets({"status": "active"}),
        )

    @app.route("/lessons/<int:lesson_id>/edit", methods=["GET", "POST"])
    @require_roles("admin", "staff")
    def edit_lesson(lesson_id: int):
        lesson = get_lesson(lesson_id)
        if lesson is None:
            abort(404)
        ensure_lesson_access(lesson)
        if request.method == "POST":
            form = request.form
            lesson_date = form.get("lesson_date", "").strip()
            preset = get_course_preset(parse_int(form.get("course_preset_id")))
            lesson_type_id = parse_int(form.get("lesson_type_id")) or (
                preset["lesson_type_id"] if preset else None
            )
            lesson_topic = form.get("lesson_topic", "").strip() or (
                preset["course_name"] if preset else ""
            )
            get_db().execute(
                """
                UPDATE lessons
                SET class_id = ?, lesson_date = ?, weekday = ?, start_time = ?, end_time = ?,
                    teacher_id = ?, lesson_type_id = ?, course_preset_id = ?, lesson_topic = ?,
                    lesson_hours = ?, classroom = ?, status = ?, remark = ?
                WHERE id = ?
                """,
                (
                    parse_int(form.get("class_id")),
                    lesson_date,
                    parse_date(lesson_date).isoweekday() if parse_date(lesson_date) else None,
                    form.get("start_time", "").strip(),
                    form.get("end_time", "").strip(),
                    resolve_lesson_teacher_id(form.get("teacher_id")),
                    lesson_type_id,
                    preset["id"] if preset else None,
                    lesson_topic,
                    parse_float(form.get("lesson_hours"), float(preset["default_hours"] or 1) if preset else 1),
                    form.get("classroom", "").strip(),
                    form.get("status", "planned"),
                    form.get("remark", "").strip(),
                    lesson_id,
                ),
            )
            get_db().commit()
            log_operation("编辑课次", "lesson", lesson_id, lesson["lesson_topic"])
            flash("课次已更新。", "success")
            return redirect(url_for("lessons", date=lesson_date))
        return render_template(
            "lesson_form.html",
            lesson=lesson,
            classes=get_active_classes(),
            teachers=get_active_teachers(),
            lesson_types=fetch_all("SELECT * FROM lesson_types ORDER BY id"),
            course_presets=query_course_presets({"status": "active"}),
        )

    @app.post("/lessons/<int:lesson_id>/cancel")
    @require_roles("admin", "staff")
    def cancel_lesson(lesson_id: int):
        lesson = get_lesson(lesson_id)
        if lesson is None:
            abort(404)
        ensure_lesson_access(lesson)
        get_db().execute("UPDATE lessons SET status = 'cancelled' WHERE id = ?", (lesson_id,))
        get_db().commit()
        log_operation("取消课次", "lesson", lesson_id)
        flash("课次已取消。", "success")
        return redirect(request.referrer or url_for("lessons"))

    @app.route("/lessons/<int:lesson_id>/attendance", methods=["GET", "POST"])
    def attendance(lesson_id: int):
        lesson = get_lesson(lesson_id)
        if lesson is None:
            abort(404)
        ensure_lesson_access(lesson)
        if request.method == "POST":
            save_attendance(lesson)
            log_operation("保存签到", "lesson", lesson_id, lesson["class_name"])
            flash("签到已保存。", "success")
            return redirect(url_for("attendance", lesson_id=lesson_id))
        rows = query_lesson_attendance(lesson_id, lesson["class_id"])
        summary = summarize_attendance(rows)
        return render_template("attendance.html", lesson=lesson, rows=rows, summary=summary)

    @app.route("/lessons/<int:lesson_id>/detail", methods=["GET", "POST"])
    def lesson_detail(lesson_id: int):
        lesson = get_lesson(lesson_id)
        if lesson is None:
            abort(404)
        ensure_lesson_access(lesson)
        if request.method == "POST":
            if error := validate_lesson_detail_payload(request.form):
                flash(error, "error")
                return redirect(url_for("lesson_detail", lesson_id=lesson_id))
            operator = (g.current_user["display_name"] or g.current_user["username"] or "系统").strip()
            save_lesson_detail(lesson_id, request.form, operator)
            log_operation("保存课程详情", "lesson", lesson_id, lesson["class_name"])
            flash("课程详情已保存。", "success")
            return redirect(url_for("lesson_detail", lesson_id=lesson_id))
        return render_template(
            "lesson_detail.html",
            lesson=lesson,
            detail=get_lesson_detail(lesson_id),
            scratch_templates=query_lesson_scratch_templates(lesson_id),
            scratch_works=query_scratch_works({"lesson_id": lesson_id}),
        )

    @app.get("/lessons/<int:lesson_id>/attendance/export")
    def export_lesson_attendance(lesson_id: int):
        lesson = get_lesson(lesson_id)
        if lesson is None:
            abort(404)
        ensure_lesson_access(lesson)
        rows = query_lesson_attendance(lesson_id, lesson["class_id"])
        file_path = export_lesson_workbook(lesson, rows)
        return send_file(file_path, as_attachment=True, download_name=file_path.name)

    @app.get("/statistics")
    def statistics():
        range_info = resolve_date_range(request.args)
        filters = {
            "class_id": request.args.get("class_id", "").strip(),
            "teacher_id": request.args.get("teacher_id", "").strip(),
            "lesson_type_id": request.args.get("lesson_type_id", "").strip(),
            "student_id": request.args.get("student_id", "").strip(),
            "attendance_status": request.args.get("attendance_status", "").strip(),
        }
        filters = apply_teacher_scope(filters)
        rows = query_attendance_rows(range_info["start_date"], range_info["end_date"], filters)
        lesson_rows = query_lessons_for_statistics(
            range_info["start_date"], range_info["end_date"], filters
        )
        summary, class_stats, teacher_stats = build_statistics(rows, lesson_rows)
        return render_template(
            "statistics.html",
            range_info=range_info,
            filters=filters,
            rows=rows,
            summary=summary,
            class_stats=class_stats,
            teacher_stats=teacher_stats,
            classes=get_active_classes(),
            teachers=get_active_teachers(),
            students=query_students(apply_teacher_scope({"status": "active"})),
            lesson_types=fetch_all("SELECT * FROM lesson_types ORDER BY id"),
        )

    @app.get("/exports/attendance")
    def export_attendance():
        range_info = resolve_date_range(request.args)
        filters = {
            "class_id": request.args.get("class_id", "").strip(),
            "teacher_id": request.args.get("teacher_id", "").strip(),
            "lesson_type_id": request.args.get("lesson_type_id", "").strip(),
            "student_id": request.args.get("student_id", "").strip(),
            "attendance_status": request.args.get("attendance_status", "").strip(),
        }
        filters = apply_teacher_scope(filters)
        rows = query_attendance_rows(range_info["start_date"], range_info["end_date"], filters)
        file_path = export_attendance_workbook(rows, range_info)
        return send_file(file_path, as_attachment=True, download_name=file_path.name)

    @app.get("/backups")
    def backups():
        pattern = "attendance_*.sql" if app.config.get("DATABASE_ENGINE") == "mysql" else "attendance_*.db"
        files = sorted(
            config.BACKUP_DIR.glob(pattern),
            key=lambda item: item.stat().st_mtime,
            reverse=True,
        )
        backup_files = [
            {
                "name": item.name,
                "size_kb": round(item.stat().st_size / 1024, 1),
                "modified": datetime.fromtimestamp(item.stat().st_mtime).strftime(
                    "%Y-%m-%d %H:%M:%S"
                ),
            }
            for item in files
        ]
        return render_template("backups.html", backup_files=backup_files)

    @app.post("/backups/create")
    def create_backup():
        if app.config.get("DATABASE_ENGINE") == "mysql":
            filename = f"attendance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql"
            target = Path(app.config["BACKUP_DIR"]) / filename
            export_mysql_backup(target, app.config["MYSQL"]["database"])
            log_operation("创建 MySQL 数据库备份", "backup", None, filename)
            flash("MySQL 数据库备份已创建。", "success")
            return redirect(url_for("backups"))

        source = Path(app.config["DATABASE"])
        if not source.exists():
            flash("数据库文件尚未创建。", "error")
            return redirect(url_for("backups"))
        filename = f"attendance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        target = Path(app.config["BACKUP_DIR"]) / filename
        shutil.copy2(source, target)
        log_operation("创建数据库备份", "backup", None, filename)
        flash("数据库备份已创建。", "success")
        return redirect(url_for("backups"))

    @app.get("/backups/<path:filename>/download")
    @require_roles("admin", "staff")
    def download_backup(filename: str):
        backup_file = resolve_backup_file(filename)
        return send_file(backup_file, as_attachment=True, download_name=backup_file.name)

    @app.post("/backups/<path:filename>/restore")
    @require_roles("admin")
    def restore_backup(filename: str):
        if app.config.get("DATABASE_ENGINE") == "mysql":
            flash("MySQL 模式暂不支持从页面直接恢复，请下载 SQL 备份后用 MySQL 工具导入。", "error")
            return redirect(url_for("backups"))

        backup_file = resolve_backup_file(filename)
        current_db = Path(app.config["DATABASE"])
        safety_name = f"before_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        if current_db.exists():
            shutil.copy2(current_db, Path(app.config["BACKUP_DIR"]) / safety_name)
        shutil.copy2(backup_file, current_db)
        log_operation("恢复数据库备份", "backup", None, backup_file.name)
        flash(f"已恢复备份 {backup_file.name}，并已额外保存恢复前备份 {safety_name}。", "success")
        return redirect(url_for("backups"))

    return app


def fetch_one(sql: str, params: tuple = ()):
    return get_db().execute(sql, params).fetchone()


def fetch_all(sql: str, params: tuple = ()):
    return get_db().execute(sql, params).fetchall()


def quote_mysql_identifier(identifier: str) -> str:
    return "`" + str(identifier).replace("`", "``") + "`"


def mysql_literal(value) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (int, float, Decimal)):
        return str(value)
    text = str(value).replace("\\", "\\\\").replace("'", "''")
    return f"'{text}'"


def export_mysql_backup(target: Path, database_name: str) -> Path:
    tables = rows_to_dicts(
        fetch_all(
            """
            SELECT TABLE_NAME AS name
            FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_SCHEMA = ? AND TABLE_TYPE = 'BASE TABLE'
            ORDER BY TABLE_NAME
            """,
            (database_name,),
        )
    )
    views = rows_to_dicts(
        fetch_all(
            """
            SELECT TABLE_NAME AS name
            FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_SCHEMA = ? AND TABLE_TYPE = 'VIEW'
            ORDER BY TABLE_NAME
            """,
            (database_name,),
        )
    )

    with target.open("w", encoding="utf-8", newline="\n") as file:
        file.write(f"-- MySQL backup for {database_name}\n")
        file.write(f"-- Created at {now_text()}\n\n")
        file.write("SET FOREIGN_KEY_CHECKS=0;\n\n")

        for view in views:
            file.write(f"DROP VIEW IF EXISTS {quote_mysql_identifier(view['name'])};\n")
        for table in reversed(tables):
            file.write(f"DROP TABLE IF EXISTS {quote_mysql_identifier(table['name'])};\n")
        file.write("\n")

        for table in tables:
            name = table["name"]
            create_row = fetch_one(f"SHOW CREATE TABLE {quote_mysql_identifier(name)}")
            file.write(create_row["Create Table"] + ";\n\n")

        for table in tables:
            name = table["name"]
            rows = rows_to_dicts(fetch_all(f"SELECT * FROM {quote_mysql_identifier(name)}"))
            if not rows:
                continue
            columns = list(rows[0].keys())
            column_sql = ", ".join(quote_mysql_identifier(column) for column in columns)
            file.write(f"INSERT INTO {quote_mysql_identifier(name)} ({column_sql}) VALUES\n")
            values = []
            for row in rows:
                values.append(
                    "(" + ", ".join(mysql_literal(row[column]) for column in columns) + ")"
                )
            file.write(",\n".join(values))
            file.write(";\n\n")

        for view in views:
            name = view["name"]
            create_row = fetch_one(f"SHOW CREATE VIEW {quote_mysql_identifier(name)}")
            create_sql = create_row["Create View"]
            file.write(create_sql + ";\n\n")

        file.write("SET FOREIGN_KEY_CHECKS=1;\n")
    return target


def now_text() -> str:
    return datetime.now().isoformat(timespec="seconds")


def parse_int(value):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def parse_float(value, default: float = 0) -> float:
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def parse_bool(value, default: bool = False) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "on", "y"}


def parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def api_error(message: str, status: int = 400):
    return jsonify({"ok": False, "message": message}), status


def required_text(data, key: str, label: str) -> str | None:
    if not str(data.get(key) or "").strip():
        return f"请填写{label}。"
    return None


def required_int(data, key: str, label: str) -> str | None:
    if parse_int(data.get(key)) is None:
        return f"请选择{label}。"
    return None


def validate_phone_field(data, key: str, label: str, required: bool = False) -> str | None:
    value = str(data.get(key) or "").strip()
    if not value:
        return f"请填写{label}。" if required else None
    if not PHONE_PATTERN.match(value):
        return f"{label}格式不正确。"
    return None


def validate_positive_number(data, key: str, label: str) -> str | None:
    value = parse_float(data.get(key), None)
    if value is None or value <= 0:
        return f"{label}必须大于 0。"
    return None


def validate_non_negative_number(data, key: str, label: str) -> str | None:
    value = parse_float(data.get(key), None)
    if value is None or value < 0:
        return f"{label}不能小于 0。"
    return None


def validate_date_range_values(start_date: str, end_date: str) -> str | None:
    if start_date and parse_date(start_date) is None:
        return "开始日期格式不正确。"
    if end_date and parse_date(end_date) is None:
        return "结束日期格式不正确。"
    if start_date and end_date and end_date < start_date:
        return "结束日期不能早于开始日期。"
    return None


def validate_time_range_values(start_time: str, end_time: str) -> str | None:
    if start_time and end_time and end_time <= start_time:
        return "结束时间必须晚于开始时间。"
    return None


def first_error(*messages: str | None) -> str | None:
    return next((message for message in messages if message), None)


def default_deduct_hours(status: str, lesson_hours: float, count_in_statistics: int) -> float:
    if not count_in_statistics:
        return 0
    return float(lesson_hours or 0) if status in DEDUCT_STATUS else 0


def course_preset_values(data) -> tuple:
    return (
        (data.get("category") or "").strip(),
        (data.get("stage") or "").strip(),
        parse_int(data.get("lesson_no")) or 1,
        (data.get("course_name") or "").strip(),
        parse_int(data.get("lesson_type_id")),
        parse_float(data.get("default_hours"), 1),
        data.get("status", "active"),
        (data.get("remark") or "").strip(),
    )


def validate_course_preset_payload(data) -> str | None:
    return first_error(
        required_text(data, "category", "课程大类"),
        required_text(data, "stage", "阶段"),
        required_int(data, "lesson_no", "第几节课"),
        required_text(data, "course_name", "具体课程名称"),
        validate_positive_number(data, "default_hours", "默认课时"),
    )


def validate_student_payload(data) -> str | None:
    return first_error(
        required_text(data, "name", "学生姓名"),
        validate_phone_field(data, "phone", "联系电话"),
        validate_phone_field(data, "parent_phone", "家长电话", required=True),
        validate_non_negative_number(data, "purchased_hours", "购买课时"),
        validate_non_negative_number(data, "gift_hours", "赠送课时"),
    )


def validate_teacher_payload(data) -> str | None:
    return first_error(
        required_text(data, "name", "教师姓名"),
        validate_phone_field(data, "phone", "手机号", required=True),
    )


def validate_class_payload(data) -> str | None:
    return first_error(
        required_text(data, "class_name", "班级名称"),
        required_text(data, "course_name", "课程方向"),
        validate_non_negative_number(data, "capacity", "容量")
        if data.get("capacity") not in (None, "")
        else None,
        validate_date_range_values(
            (data.get("start_date") or "").strip(),
            (data.get("end_date") or "").strip(),
        ),
        validate_time_range_values(
            (data.get("default_start_time") or "").strip(),
            (data.get("default_end_time") or "").strip(),
        ),
    )


def validate_lesson_payload(data, preset=None) -> str | None:
    lesson_date = (data.get("lesson_date") or "").strip()
    topic = (data.get("lesson_topic") or "").strip() or (preset["course_name"] if preset else "")
    return first_error(
        required_int(data, "class_id", "班级"),
        "上课日期格式不正确。" if not parse_date(lesson_date) else None,
        required_text(data, "start_time", "开始时间"),
        required_text(data, "end_time", "结束时间"),
        validate_time_range_values(
            (data.get("start_time") or "").strip(),
            (data.get("end_time") or "").strip(),
        ),
        "请选择上课老师。" if resolve_lesson_teacher_id(data.get("teacher_id")) is None else None,
        "请选择预设课程或填写课程主题。" if not topic else None,
        validate_positive_number(
            {"lesson_hours": data.get("lesson_hours") or (preset["default_hours"] if preset else None)},
            "lesson_hours",
            "课时数量",
        ),
    )


def validate_lesson_type_payload(data) -> str | None:
    return first_error(
        required_text(data, "type_name", "类型名称"),
        validate_positive_number(data, "default_hours", "默认课时"),
    )


def validate_user_payload(data, require_password: bool = False) -> str | None:
    password = (data.get("password") or "").strip()
    role = normalize_role((data.get("role") or "").strip())
    if require_password and not password:
        return "请设置初始密码。"
    if password and len(password) < 6:
        return "密码至少需要 6 位。"
    if role not in db_role_codes():
        return "请选择可分配的角色。"
    if role == "teacher" and not parse_int(data.get("teacher_id")):
        return "老师账号必须关联教师。"
    if role in {"student", "parent"} and not parse_int(data.get("student_id")):
        return "学生/家长账号必须关联学生。"
    return first_error(
        required_text(data, "username", "登录账号") if require_password else None,
        required_text(data, "display_name", "显示名称"),
        required_text(data, "role", "角色"),
    )


def validate_lesson_detail_payload(data) -> str | None:
    return first_error(
        required_text(data, "teaching_content", "本节课内容"),
        required_text(data, "learning_goal", "教学目标"),
    )


COURSE_PRESET_IMPORT_HEADERS = {
    "课程大类": "category",
    "大类": "category",
    "某某课": "category",
    "category": "category",
    "阶段": "stage",
    "stage": "stage",
    "第几节课": "lesson_no",
    "节次": "lesson_no",
    "课程序号": "lesson_no",
    "lesson_no": "lesson_no",
    "具体课程名称": "course_name",
    "课程名称": "course_name",
    "内容": "course_name",
    "标题": "course_name",
    "content": "course_name",
    "title": "course_name",
    "course_name": "course_name",
    "课程类型": "lesson_type_name",
    "默认课程类型": "lesson_type_name",
    "lesson_type": "lesson_type_name",
    "默认课时": "default_hours",
    "课时": "default_hours",
    "default_hours": "default_hours",
    "状态": "status",
    "status": "status",
    "备注": "remark",
    "remark": "remark",
}

COURSE_PRESET_IMPORT_ORDER = (
    "category",
    "stage",
    "lesson_no",
    "course_name",
    "lesson_type_name",
    "default_hours",
    "status",
    "remark",
)
COURSE_PRESET_REQUIRED_IMPORT_HEADERS = {"category", "stage", "lesson_no", "course_name"}


def normalize_import_header(value) -> str:
    key = str(value or "").strip().lstrip("\ufeff")
    return COURSE_PRESET_IMPORT_HEADERS.get(key) or COURSE_PRESET_IMPORT_HEADERS.get(key.lower(), key)


def normalize_import_status(value) -> str:
    text = str(value or "active").strip()
    if text in {"停用", "禁用", "inactive", "disabled"}:
        return "inactive"
    return "active"


def import_row_has_value(row) -> bool:
    return any(str(cell or "").strip() for cell in row)


def import_row_looks_like_data(row) -> bool:
    values = [str(cell or "").strip() for cell in row]
    return (
        len(values) >= 4
        and bool(values[0])
        and bool(values[1])
        and parse_int(values[2]) is not None
        and bool(values[3])
    )


def map_course_preset_import_rows(rows) -> list[dict]:
    clean_rows = [list(row or []) for row in rows if import_row_has_value(row or [])]
    if not clean_rows:
        raise ValueError("导入文件为空。")

    headers = [normalize_import_header(cell) for cell in clean_rows[0]]
    if COURSE_PRESET_REQUIRED_IMPORT_HEADERS.issubset(set(headers)):
        return [
            {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            for row in clean_rows[1:]
        ]

    data_rows = clean_rows if import_row_looks_like_data(clean_rows[0]) else clean_rows[1:]
    return [
        {
            COURSE_PRESET_IMPORT_ORDER[index]: value
            for index, value in enumerate(row)
            if index < len(COURSE_PRESET_IMPORT_ORDER)
        }
        for row in data_rows
    ]


def read_course_preset_import_rows(upload) -> list[dict]:
    filename = upload.filename.lower()
    if filename.endswith((".xlsx", ".xlsm")):
        upload.stream.seek(0)
        workbook = load_workbook(upload.stream, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        return map_course_preset_import_rows(rows)
    if filename.endswith(".csv"):
        upload.stream.seek(0)
        raw_content = upload.stream.read()
        try:
            text = raw_content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw_content.decode("gb18030")
        csv_rows = list(csv.reader(StringIO(text)))
        return map_course_preset_import_rows(csv_rows)
    raise ValueError("仅支持 .xlsx、.xlsm 或 .csv 文件。")


def normalize_course_preset_import_row(row: dict, lesson_type_ids: dict) -> dict | None:
    category = str(row.get("category") or "").strip()
    stage = str(row.get("stage") or "").strip()
    course_name = str(row.get("course_name") or "").strip()
    if not category and not stage and not course_name:
        return None
    lesson_no = parse_int(row.get("lesson_no"))
    if not category or not stage or not lesson_no or not course_name:
        raise ValueError("导入失败：课程大类、阶段、第几节课、具体课程名称均为必填。")
    lesson_type_name = str(row.get("lesson_type_name") or "").strip()
    return {
        "category": category,
        "stage": stage,
        "lesson_no": lesson_no,
        "course_name": course_name,
        "lesson_type_id": lesson_type_ids.get(lesson_type_name) if lesson_type_name else None,
        "default_hours": parse_float(row.get("default_hours"), 1),
        "status": normalize_import_status(row.get("status")),
        "remark": str(row.get("remark") or "").strip(),
    }


def import_course_presets(rows: list[dict]) -> dict:
    if not rows:
        raise ValueError("导入文件没有可读取的数据。")
    db = get_db()
    lesson_type_ids = {
        row["type_name"]: row["id"] for row in fetch_all("SELECT id, type_name FROM lesson_types")
    }
    created = 0
    updated = 0
    skipped = 0
    timestamp = now_text()
    for raw_row in rows:
        row = normalize_course_preset_import_row(raw_row, lesson_type_ids)
        if row is None:
            skipped += 1
            continue
        exists = fetch_one(
            """
            SELECT id FROM course_presets
            WHERE category = ? AND stage = ? AND lesson_no = ? AND course_name = ?
            """,
            (row["category"], row["stage"], row["lesson_no"], row["course_name"]),
        )
        db.execute(
            """
            INSERT INTO course_presets
                (category, stage, lesson_no, course_name, lesson_type_id,
                 default_hours, status, remark, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(category, stage, lesson_no, course_name) DO UPDATE SET
                lesson_type_id = excluded.lesson_type_id,
                default_hours = excluded.default_hours,
                status = excluded.status,
                remark = excluded.remark
            """,
            (
                row["category"],
                row["stage"],
                row["lesson_no"],
                row["course_name"],
                row["lesson_type_id"],
                row["default_hours"],
                row["status"],
                row["remark"],
                timestamp,
            ),
        )
        if exists:
            updated += 1
        else:
            created += 1
    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped}


def export_course_preset_import_template() -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "预设课程导入模板"
    append_header(
        sheet,
        [
            "课程大类",
            "阶段",
            "第几节课",
            "具体课程名称",
            "课程类型",
            "默认课时",
            "状态",
            "备注",
        ],
    )
    sheet.append(["Scratch", "入门阶段", 1, "角色移动与坐标", "常规课", 1, "启用", "示例，可删除"])
    sheet.append(["Python", "基础阶段", 2, "变量与输入输出", "常规课", 1.5, "启用", ""])
    sheet.freeze_panes = "A2"
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
    autosize(sheet)
    filename = f"course_preset_import_template_{safe_stamp()}.xlsx"
    file_path = config.EXPORT_DIR / filename
    workbook.save(file_path)
    return file_path


def get_course_preset(preset_id):
    if not preset_id:
        return None
    return fetch_one("SELECT * FROM course_presets WHERE id = ?", (preset_id,))


def query_course_presets(filters: dict):
    conditions = []
    params = []
    if filters.get("q"):
        conditions.append(
            "(cp.category LIKE ? OR cp.stage LIKE ? OR cp.course_name LIKE ? OR cp.remark LIKE ?)"
        )
        term = f"%{filters['q']}%"
        params.extend([term, term, term, term])
    if filters.get("category"):
        conditions.append("cp.category = ?")
        params.append(filters["category"])
    if filters.get("stage"):
        conditions.append("cp.stage = ?")
        params.append(filters["stage"])
    if filters.get("status"):
        conditions.append("cp.status = ?")
        params.append(filters["status"])
    where_sql = "WHERE " + " AND ".join(conditions) if conditions else ""
    return fetch_all(
        f"""
        SELECT cp.*, lt.type_name AS lesson_type_name
        FROM course_presets cp
        LEFT JOIN lesson_types lt ON lt.id = cp.lesson_type_id
        {where_sql}
        ORDER BY cp.category, cp.stage, cp.lesson_no, cp.course_name
        """,
        tuple(params),
    )


def query_students(filters: dict):
    conditions = []
    params = []
    join_params = []
    class_join = "LEFT JOIN classes c ON c.id = cs.class_id"
    if filters.get("teacher_id"):
        class_join = "LEFT JOIN classes c ON c.id = cs.class_id AND c.teacher_id = ?"
        join_params.append(filters["teacher_id"])
    if filters.get("q"):
        conditions.append("(s.name LIKE ? OR s.parent_phone LIKE ? OR s.phone LIKE ?)")
        term = f"%{filters['q']}%"
        params.extend([term, term, term])
    if filters.get("status"):
        conditions.append("s.status = ?")
        params.append(filters["status"])
    if filters.get("class_id"):
        conditions.append(
            """
            EXISTS (
                SELECT 1 FROM class_students cs2
                WHERE cs2.student_id = s.id AND cs2.class_id = ? AND cs2.status = 'active'
            )
            """
        )
        params.append(filters["class_id"])
    if filters.get("teacher_id"):
        conditions.append(
            """
            EXISTS (
                SELECT 1
                FROM class_students cs3
                JOIN classes c3 ON c3.id = cs3.class_id
                WHERE cs3.student_id = s.id
                  AND cs3.status = 'active'
                  AND c3.teacher_id = ?
            )
            """
        )
        params.append(filters["teacher_id"])
    where_sql = "WHERE " + " AND ".join(conditions) if conditions else ""
    return fetch_all(
        f"""
        SELECT s.*,
               COALESCE(GROUP_CONCAT(c.class_name, '、'), '') AS class_names,
               (
                   SELECT COALESCE(SUM(a.deduct_hours), 0)
                   FROM attendance a
                   WHERE a.student_id = s.id
               ) AS consumed_hours
        FROM students s
        LEFT JOIN class_students cs ON cs.student_id = s.id AND cs.status = 'active'
        {class_join}
        {where_sql}
        GROUP BY s.id
        ORDER BY s.created_at DESC, s.id DESC
        """,
        tuple(join_params + params),
    )


def query_teachers(filters: dict):
    conditions = []
    params = []
    if filters.get("q"):
        conditions.append("(t.name LIKE ? OR t.phone LIKE ? OR t.subject LIKE ?)")
        term = f"%{filters['q']}%"
        params.extend([term, term, term])
    if filters.get("status"):
        conditions.append("t.status = ?")
        params.append(filters["status"])
    if filters.get("teacher_id"):
        conditions.append("t.id = ?")
        params.append(filters["teacher_id"])
    where_sql = "WHERE " + " AND ".join(conditions) if conditions else ""
    return fetch_all(
        f"""
        SELECT t.*,
               COUNT(DISTINCT l.id) AS lesson_count,
               COALESCE(SUM(CASE WHEN l.status != 'cancelled' THEN l.lesson_hours ELSE 0 END), 0)
                   AS lesson_hours
        FROM teachers t
        LEFT JOIN lessons l ON l.teacher_id = t.id
        {where_sql}
        GROUP BY t.id
        ORDER BY t.created_at DESC, t.id DESC
        """,
        tuple(params),
    )


def query_classes(filters: dict):
    conditions = []
    params = []
    if filters.get("q"):
        conditions.append("(c.class_name LIKE ? OR c.course_name LIKE ? OR c.class_type LIKE ?)")
        term = f"%{filters['q']}%"
        params.extend([term, term, term])
    if filters.get("status"):
        conditions.append("c.status = ?")
        params.append(filters["status"])
    if filters.get("teacher_id"):
        conditions.append("c.teacher_id = ?")
        params.append(filters["teacher_id"])
    where_sql = "WHERE " + " AND ".join(conditions) if conditions else ""
    return fetch_all(
        f"""
        SELECT c.*, t.name AS teacher_name,
               COUNT(DISTINCT CASE WHEN cs.status = 'active' THEN cs.student_id END) AS student_count,
               COUNT(DISTINCT CASE WHEN l.status != 'cancelled' THEN l.id END) AS lesson_count
        FROM classes c
        LEFT JOIN teachers t ON t.id = c.teacher_id
        LEFT JOIN class_students cs ON cs.class_id = c.id
        LEFT JOIN lessons l ON l.class_id = c.id
        {where_sql}
        GROUP BY c.id
        ORDER BY c.created_at DESC, c.id DESC
        """,
        tuple(params),
    )


def query_lessons(filters: dict):
    conditions = []
    params = []
    if filters.get("start_date") and filters.get("end_date"):
        conditions.append("l.lesson_date BETWEEN ? AND ?")
        params.extend([filters["start_date"], filters["end_date"]])
    elif filters.get("date"):
        conditions.append("l.lesson_date = ?")
        params.append(filters["date"])
    if filters.get("class_id"):
        conditions.append("l.class_id = ?")
        params.append(filters["class_id"])
    if filters.get("teacher_id"):
        conditions.append("l.teacher_id = ?")
        params.append(filters["teacher_id"])
    if filters.get("lesson_type_id"):
        conditions.append("l.lesson_type_id = ?")
        params.append(filters["lesson_type_id"])
    if filters.get("status"):
        conditions.append("l.status = ?")
        params.append(filters["status"])
    where_sql = "WHERE " + " AND ".join(conditions) if conditions else ""
    return fetch_all(
        f"""
        SELECT l.*, c.class_name, c.course_name, t.name AS teacher_name,
               lt.type_name, lt.count_in_statistics,
               cp.category AS course_category,
               cp.stage AS course_stage,
               cp.lesson_no AS course_lesson_no,
               cp.course_name AS preset_course_name,
               (
                   SELECT COUNT(*)
                   FROM class_students cs
                   JOIN students s ON s.id = cs.student_id
                   WHERE cs.class_id = l.class_id AND cs.status = 'active' AND s.status = 'active'
               ) AS expected_count,
               (
                   SELECT COUNT(*)
                   FROM attendance a
                   WHERE a.lesson_id = l.id
                     AND a.status IN ('已到', '迟到', '补签', '试听')
               ) AS arrived_count,
               (
                   SELECT COUNT(*)
                   FROM attendance a
                   WHERE a.lesson_id = l.id AND a.status = '请假'
               ) AS leave_count,
               (
                   SELECT COUNT(*)
                   FROM attendance a
                   WHERE a.lesson_id = l.id AND a.status = '缺勤'
               ) AS absent_count,
               CASE
                   WHEN EXISTS (SELECT 1 FROM lesson_details ld WHERE ld.lesson_id = l.id) THEN 1
                   ELSE 0
               END AS has_lesson_detail,
               (
                   SELECT ld.updated_at
                   FROM lesson_details ld
                   WHERE ld.lesson_id = l.id
               ) AS detail_updated_at
        FROM lessons l
        JOIN classes c ON c.id = l.class_id
        LEFT JOIN teachers t ON t.id = l.teacher_id
        LEFT JOIN lesson_types lt ON lt.id = l.lesson_type_id
        LEFT JOIN course_presets cp ON cp.id = l.course_preset_id
        {where_sql}
        ORDER BY l.lesson_date DESC, l.start_time, l.id DESC
        """,
        tuple(params),
    )


def get_lesson(lesson_id: int):
    return fetch_one(
        """
        SELECT l.*, c.class_name, c.course_name, t.name AS teacher_name,
               lt.type_name, lt.count_in_statistics,
               cp.category AS course_category,
               cp.stage AS course_stage,
               cp.lesson_no AS course_lesson_no,
               cp.course_name AS preset_course_name
        FROM lessons l
        JOIN classes c ON c.id = l.class_id
        LEFT JOIN teachers t ON t.id = l.teacher_id
        LEFT JOIN lesson_types lt ON lt.id = l.lesson_type_id
        LEFT JOIN course_presets cp ON cp.id = l.course_preset_id
        WHERE l.id = ?
        """,
        (lesson_id,),
    )


def current_access_scope(user=None) -> dict:
    user = user or g.get("current_user")
    role = normalize_role(user["role"]) if user else ""
    return {
        "role": role,
        "teacher_id": user.get("teacher_id") if user else None,
        "student_id": user.get("student_id") if user else None,
        "is_global": role not in {"teacher", "student", "parent"},
    }


def student_in_teacher_scope(student_id: int, teacher_id: int | None) -> bool:
    if not teacher_id:
        return False
    allowed = fetch_one(
        """
        SELECT 1
        FROM class_students cs
        JOIN classes c ON c.id = cs.class_id
        WHERE cs.student_id = ?
          AND cs.status = 'active'
          AND c.teacher_id = ?
        """,
        (student_id, teacher_id),
    )
    return allowed is not None


def ensure_lesson_access(lesson):
    scope = current_access_scope()
    if scope["role"] == "teacher":
        if not scope["teacher_id"] or lesson["teacher_id"] != scope["teacher_id"]:
            abort(403)


def ensure_class_access(class_row):
    scope = current_access_scope()
    if scope["role"] == "teacher":
        if not scope["teacher_id"] or class_row["teacher_id"] != scope["teacher_id"]:
            abort(403)


def ensure_student_access(student_id: int):
    scope = current_access_scope()
    if scope["role"] in {"student", "parent"}:
        if not scope["student_id"] or int(scope["student_id"]) != int(student_id):
            abort(403)
    if scope["role"] == "teacher" and not student_in_teacher_scope(student_id, scope["teacher_id"]):
        abort(403)


def empty_student_portal_context() -> dict:
    return {
        "student": None,
        "summary": {
            "total_hours": 0,
            "consumed_hours": 0,
            "remaining_hours": 0,
            "attendance_count": 0,
            "arrived_count": 0,
            "leave_count": 0,
            "absent_count": 0,
            "attendance_rate": 0,
        },
        "classes": [],
        "upcoming_lessons": [],
        "recent_attendance": [],
        "scratch_tasks": [],
        "scratch_summary": {
            "work_count": 0,
            "submitted_count": 0,
            "passed_count": 0,
            "public_count": 0,
            "judge_count": 0,
            "total_points": 0,
        },
        "student_features": [],
    }


def empty_parent_portal_context() -> dict:
    context = empty_student_portal_context()
    context.update(
        {
            "report_cards": [],
            "guardian_notes": [],
        }
    )
    return context


STUDENT_HOUR_FIELDS = (
    "purchased_hours",
    "gift_hours",
    "total_hours",
    "consumed_hours",
    "remaining_hours",
)


def can_view_student_hours(user=None) -> bool:
    user = user or g.get("current_user")
    return has_permission(user, "student_portal.hours.view") or has_permission(user, "students.view")


def redact_student_hours(context: dict) -> dict:
    if can_view_student_hours():
        return context
    redacted = dict(context)
    summary = dict(redacted.get("summary") or {})
    for key in ("total_hours", "consumed_hours", "remaining_hours"):
        summary[key] = None
    redacted["summary"] = summary
    student = redacted.get("student")
    if student:
        student_row = dict(student)
        for key in STUDENT_HOUR_FIELDS:
            if key in student_row:
                student_row[key] = None
        redacted["student"] = student_row
    return redacted


def query_student_scratch_summary(student_id: int) -> dict:
    works = fetch_one(
        """
        SELECT
            COUNT(*) AS work_count,
            SUM(CASE WHEN status IN ('submitted', 'reviewed') THEN 1 ELSE 0 END) AS submitted_count,
            SUM(CASE WHEN judge_status = 'passed' THEN 1 ELSE 0 END) AS passed_count,
            SUM(CASE WHEN visibility = 'public' THEN 1 ELSE 0 END) AS public_count
        FROM scratch_works
        WHERE student_id = ?
        """,
        (student_id,),
    )
    judge_runs = fetch_one(
        """
        SELECT COUNT(*) AS judge_count
        FROM scratch_judge_runs
        WHERE student_id = ?
        """,
        (student_id,),
    )
    points = fetch_one(
        """
        SELECT COALESCE(SUM(points), 0) AS total_points
        FROM student_points_ledger
        WHERE student_id = ?
        """,
        (student_id,),
    )
    works = dict(works) if works else {}
    judge_runs = dict(judge_runs) if judge_runs else {}
    points = dict(points) if points else {}
    return {
        "work_count": int(works.get("work_count") or 0),
        "submitted_count": int(works.get("submitted_count") or 0),
        "passed_count": int(works.get("passed_count") or 0),
        "public_count": int(works.get("public_count") or 0),
        "judge_count": int(judge_runs.get("judge_count") or 0),
        "total_points": int(points.get("total_points") or 0),
    }


def query_student_courseware_assets(student_id: int):
    return fetch_all(
        """
        SELECT la.id AS lesson_asset_id, la.lesson_id, la.title, la.note,
               a.id AS asset_id, a.original_filename, a.public_path, a.mime_type, a.file_size,
               l.lesson_date, l.start_time, l.end_time, l.lesson_topic,
               c.class_name, c.course_name
        FROM lesson_assets la
        JOIN uploaded_assets a ON a.id = la.asset_id
        JOIN lessons l ON l.id = la.lesson_id
        JOIN classes c ON c.id = l.class_id
        JOIN class_students cs ON cs.class_id = c.id
        WHERE cs.student_id = ?
          AND cs.status = 'active'
          AND la.status = 'active'
          AND a.status = 'uploaded'
          AND l.status != 'cancelled'
        ORDER BY l.lesson_date DESC, la.created_at DESC, la.id DESC
        LIMIT 8
        """,
        (student_id,),
    )


def build_student_portal_context(student_id: int) -> dict:
    student = fetch_one(
        """
        SELECT *
        FROM v_students_summary
        WHERE student_id = ?
        """,
        (student_id,),
    )
    if student is None:
        student = fetch_one("SELECT * FROM students WHERE id = ?", (student_id,))
    if student is None:
        abort(404)

    classes = rows_to_dicts(
        fetch_all(
            """
            SELECT c.id, c.class_name, c.course_name, c.class_type,
                   t.name AS teacher_name, cs.join_date
            FROM class_students cs
            JOIN classes c ON c.id = cs.class_id
            LEFT JOIN teachers t ON t.id = c.teacher_id
            WHERE cs.student_id = ? AND cs.status = 'active' AND c.status != 'archived'
            ORDER BY c.class_name
            """,
            (student_id,),
        )
    )
    upcoming_lessons = rows_to_dicts(
        fetch_all(
            """
            SELECT l.id, l.lesson_date, l.start_time, l.end_time, l.lesson_topic,
                   l.classroom, l.status, c.class_name, c.course_name,
                   t.name AS teacher_name, cp.course_name AS preset_course_name,
                   cp.category AS course_category, cp.stage AS course_stage
            FROM lessons l
            JOIN classes c ON c.id = l.class_id
            JOIN class_students cs ON cs.class_id = c.id
            LEFT JOIN teachers t ON t.id = l.teacher_id
            LEFT JOIN course_presets cp ON cp.id = l.course_preset_id
            WHERE cs.student_id = ?
              AND cs.status = 'active'
              AND l.status != 'cancelled'
              AND l.lesson_date >= ?
            ORDER BY l.lesson_date, l.start_time
            LIMIT 8
            """,
            (student_id, date.today().isoformat()),
        )
    )
    recent_attendance = rows_to_dicts(
        fetch_all(
            """
            SELECT a.status, a.deduct_hours, a.remark, l.lesson_date,
                   l.start_time, l.end_time, l.lesson_topic,
                   c.class_name, t.name AS teacher_name,
                   cp.course_name AS preset_course_name
            FROM attendance a
            JOIN lessons l ON l.id = a.lesson_id
            JOIN classes c ON c.id = l.class_id
            LEFT JOIN teachers t ON t.id = l.teacher_id
            LEFT JOIN course_presets cp ON cp.id = l.course_preset_id
            WHERE a.student_id = ?
            ORDER BY l.lesson_date DESC, l.start_time DESC
            LIMIT 8
            """,
            (student_id,),
        )
    )
    scratch_tasks = rows_to_dicts(query_student_scratch_tasks(student_id))
    scratch_summary = query_student_scratch_summary(student_id)
    courseware_assets = rows_to_dicts(query_student_courseware_assets(student_id))
    row = dict(student)
    if "student_name" not in row:
        row["student_name"] = row.get("name")
    if "class_names" not in row:
        row["class_names"] = "、".join(item["class_name"] for item in classes)
    total_hours = float(row.get("total_hours") or ((row.get("purchased_hours") or 0) + (row.get("gift_hours") or 0)))
    consumed_hours = float(row.get("consumed_hours") or 0)
    arrived_count = int(row.get("arrived_count") or 0)
    attendance_count = int(row.get("attendance_count") or 0)
    summary = {
        "total_hours": round(total_hours, 1),
        "consumed_hours": round(consumed_hours, 1),
        "remaining_hours": round(total_hours - consumed_hours, 1),
        "attendance_count": attendance_count,
        "arrived_count": arrived_count,
        "leave_count": int(row.get("leave_count") or 0),
        "absent_count": int(row.get("absent_count") or 0),
        "attendance_rate": round(arrived_count * 100 / attendance_count, 1) if attendance_count else 0,
    }
    return {
        "student": row,
        "summary": summary,
        "classes": classes,
        "upcoming_lessons": upcoming_lessons,
        "recent_attendance": recent_attendance,
        "courseware_assets": courseware_assets,
        "scratch_tasks": scratch_tasks,
        "scratch_summary": scratch_summary,
        "student_features": [
            {"title": "上课", "label": "近期课程", "value": len(upcoming_lessons), "hint": "查看自己的上课安排和任务入口"},
            {"title": "课件", "label": "可下载资料", "value": len(courseware_assets), "hint": "查看老师为本人班级课次上传的课件"},
            {"title": "作品", "label": "Scratch 模板", "value": scratch_summary["work_count"], "hint": "进入模板作品并保存自己的课堂作品"},
            {"title": "积分", "label": "成长积分", "value": scratch_summary["total_points"], "hint": "来自作品提交和老师点评等课堂行为"},
        ],
    }


def build_parent_portal_context(student_id: int) -> dict:
    context = build_student_portal_context(student_id)
    summary = context["summary"]
    upcoming_count = len(context["upcoming_lessons"])
    recent_count = len(context["recent_attendance"])
    context["report_cards"] = [
        {"title": "课时余额", "value": summary["remaining_hours"], "hint": "仅展示孩子本人的课时余额"},
        {"title": "近期排课", "value": upcoming_count, "hint": "只展示孩子已加入班级的课程"},
        {"title": "出勤率", "value": f'{summary["attendance_rate"]}%', "hint": "根据孩子本人学习记录计算"},
        {"title": "学习记录", "value": recent_count, "hint": "最近课程与签到状态"},
    ]
    context["guardian_notes"] = [
        "家长端当前只读开放课时、班级、排课和学习记录。",
        "不会展示同班其他学生姓名、联系方式、签到或课时信息。",
        "后续可接入学习报告、作业点评和通知确认。",
    ]
    return context


def resolve_lesson_teacher_id(value):
    requested_teacher_id = parse_int(value)
    user = g.get("current_user")
    if user and user["role"] == "teacher":
        if not user["teacher_id"]:
            abort(403)
        if requested_teacher_id and requested_teacher_id != user["teacher_id"]:
            abort(403)
        return user["teacher_id"]
    return requested_teacher_id


def empty_lesson_detail(lesson_id: int) -> dict:
    detail = {
        "id": None,
        "lesson_id": lesson_id,
        "updated_by": "",
        "created_at": "",
        "updated_at": "",
    }
    detail.update({field: "" for field in LESSON_DETAIL_FIELDS})
    return detail


def normalize_lesson_detail_payload(data) -> dict:
    return {field: (data.get(field) or "").strip() for field in LESSON_DETAIL_FIELDS}


def get_lesson_detail(lesson_id: int) -> dict:
    row = fetch_one("SELECT * FROM lesson_details WHERE lesson_id = ?", (lesson_id,))
    return dict(row) if row else empty_lesson_detail(lesson_id)


def save_lesson_detail(lesson_id: int, data, updated_by: str) -> dict:
    payload = normalize_lesson_detail_payload(data)
    timestamp = now_text()
    get_db().execute(
        """
        INSERT INTO lesson_details
            (lesson_id, teaching_content, learning_goal, class_performance, homework,
             next_plan, materials, updated_by, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(lesson_id) DO UPDATE SET
            teaching_content = excluded.teaching_content,
            learning_goal = excluded.learning_goal,
            class_performance = excluded.class_performance,
            homework = excluded.homework,
            next_plan = excluded.next_plan,
            materials = excluded.materials,
            updated_by = excluded.updated_by,
            updated_at = excluded.updated_at
        """,
        (
            lesson_id,
            payload["teaching_content"],
            payload["learning_goal"],
            payload["class_performance"],
            payload["homework"],
            payload["next_plan"],
            payload["materials"],
            updated_by,
            timestamp,
            timestamp,
        ),
    )
    get_db().commit()
    return get_lesson_detail(lesson_id)


def query_lesson_attendance(lesson_id: int, class_id: int):
    return fetch_all(
        """
        SELECT s.id AS student_id, s.name, s.phone, s.parent_phone,
               COALESCE(a.status, '未确认') AS status,
               COALESCE(a.deduct_hours, 0) AS deduct_hours,
               a.checkin_time,
               COALESCE(a.operator, '') AS operator,
               COALESCE(a.remark, '') AS remark
        FROM class_students cs
        JOIN students s ON s.id = cs.student_id
        LEFT JOIN attendance a ON a.student_id = s.id AND a.lesson_id = ?
        WHERE cs.class_id = ?
          AND cs.status = 'active'
          AND s.status = 'active'
        ORDER BY s.name
        """,
        (lesson_id, class_id),
    )


def summarize_attendance(rows):
    summary = {status: 0 for status in ATTENDANCE_STATUS}
    summary["deduct_hours"] = 0
    for row in rows:
        summary[row["status"]] = summary.get(row["status"], 0) + 1
        summary["deduct_hours"] += float(row["deduct_hours"] or 0)
    return summary


def save_attendance(lesson):
    student_ids = request.form.getlist("student_ids")
    operator = request.form.get("operator", "").strip() or "系统"
    timestamp = now_text()
    count_in_statistics = int(lesson["count_in_statistics"] or 0)
    for student_id in student_ids:
        status = request.form.get(f"status_{student_id}", "未确认")
        if status not in ATTENDANCE_STATUS:
            status = "未确认"
        deduct_hours = parse_float(
            request.form.get(f"deduct_hours_{student_id}"),
            default_deduct_hours(status, lesson["lesson_hours"], count_in_statistics),
        )
        if not count_in_statistics:
            deduct_hours = 0
        checkin_time = timestamp if status != "未确认" else None
        remark = request.form.get(f"remark_{student_id}", "").strip()
        get_db().execute(
            """
            INSERT INTO attendance
                (lesson_id, student_id, status, checkin_time, deduct_hours, operator,
                 remark, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(lesson_id, student_id) DO UPDATE SET
                status = excluded.status,
                checkin_time = excluded.checkin_time,
                deduct_hours = excluded.deduct_hours,
                operator = excluded.operator,
                remark = excluded.remark,
                updated_at = excluded.updated_at
            """,
            (
                lesson["id"],
                parse_int(student_id),
                status,
                checkin_time,
                deduct_hours,
                operator,
                remark,
                timestamp,
                timestamp,
            ),
        )
    if lesson["status"] != "cancelled":
        get_db().execute("UPDATE lessons SET status = 'completed' WHERE id = ?", (lesson["id"],))
    get_db().commit()


def save_attendance_records(lesson, records: list[dict], operator: str):
    timestamp = now_text()
    count_in_statistics = int(lesson["count_in_statistics"] or 0)
    for record in records:
        student_id = parse_int(record.get("student_id"))
        if not student_id:
            continue
        status = record.get("status", "未确认")
        if status not in ATTENDANCE_STATUS:
            status = "未确认"
        deduct_hours = parse_float(
            record.get("deduct_hours"),
            default_deduct_hours(status, lesson["lesson_hours"], count_in_statistics),
        )
        if not count_in_statistics:
            deduct_hours = 0
        checkin_time = timestamp if status != "未确认" else None
        remark = (record.get("remark") or "").strip()
        get_db().execute(
            """
            INSERT INTO attendance
                (lesson_id, student_id, status, checkin_time, deduct_hours, operator,
                 remark, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(lesson_id, student_id) DO UPDATE SET
                status = excluded.status,
                checkin_time = excluded.checkin_time,
                deduct_hours = excluded.deduct_hours,
                operator = excluded.operator,
                remark = excluded.remark,
                updated_at = excluded.updated_at
            """,
            (
                lesson["id"],
                student_id,
                status,
                checkin_time,
                deduct_hours,
                operator,
                remark,
                timestamp,
                timestamp,
            ),
        )
    if lesson["status"] != "cancelled":
        get_db().execute("UPDATE lessons SET status = 'completed' WHERE id = ?", (lesson["id"],))
    get_db().commit()


def count_attendance_by_status(start_date: str, end_date: str, filters: dict | None = None):
    conditions = ["l.lesson_date BETWEEN ? AND ?"]
    params = [start_date, end_date]
    if filters and filters.get("teacher_id"):
        conditions.append("l.teacher_id = ?")
        params.append(filters["teacher_id"])
    where_sql = " AND ".join(conditions)
    rows = fetch_all(
        f"""
        SELECT a.status, COUNT(*) AS count
        FROM attendance a
        JOIN lessons l ON l.id = a.lesson_id
        WHERE {where_sql}
        GROUP BY a.status
        """,
        tuple(params),
    )
    return {row["status"]: row["count"] for row in rows}


def get_active_teachers():
    user = g.get("current_user")
    if user and user["role"] == "teacher":
        return fetch_all(
            "SELECT * FROM teachers WHERE status = 'active' AND id = ? ORDER BY name",
            (user["teacher_id"] or 0,),
        )
    return fetch_all("SELECT * FROM teachers WHERE status = 'active' ORDER BY name")


def get_active_students():
    user = g.get("current_user")
    if user and normalize_role(user["role"]) == "student":
        return fetch_all(
            "SELECT * FROM students WHERE status = 'active' AND id = ? ORDER BY name",
            (user.get("student_id") or 0,),
        )
    return fetch_all("SELECT * FROM students WHERE status = 'active' ORDER BY name")


def get_active_classes():
    user = g.get("current_user")
    if user and user["role"] == "teacher":
        return fetch_all(
            """
            SELECT *
            FROM classes
            WHERE status != 'archived' AND teacher_id = ?
            ORDER BY class_name
            """,
            (user["teacher_id"] or 0,),
        )
    return fetch_all("SELECT * FROM classes WHERE status != 'archived' ORDER BY class_name")


def resolve_date_range(args):
    scope = args.get("scope", "day")
    today_value = date.today()
    if args.get("start_date") and args.get("end_date"):
        start = parse_date(args.get("start_date")) or today_value
        end = parse_date(args.get("end_date")) or start
        return {
            "scope": scope,
            "date": args.get("date", today_value.isoformat()),
            "month": args.get("month", today_value.strftime("%Y-%m")),
            "start_date": start.isoformat(),
            "end_date": end.isoformat(),
        }
    if scope == "week":
        selected = parse_date(args.get("date")) or today_value
        start = selected - timedelta(days=selected.weekday())
        end = start + timedelta(days=6)
    elif scope == "month":
        month_text = args.get("month") or today_value.strftime("%Y-%m")
        try:
            start = datetime.strptime(month_text + "-01", "%Y-%m-%d").date()
        except ValueError:
            start = today_value.replace(day=1)
            month_text = start.strftime("%Y-%m")
        if start.month == 12:
            next_month = start.replace(year=start.year + 1, month=1)
        else:
            next_month = start.replace(month=start.month + 1)
        end = next_month - timedelta(days=1)
        return {
            "scope": scope,
            "date": args.get("date", today_value.isoformat()),
            "month": month_text,
            "start_date": start.isoformat(),
            "end_date": end.isoformat(),
        }
    else:
        selected = parse_date(args.get("date")) or today_value
        start = end = selected
    return {
        "scope": scope,
        "date": args.get("date", today_value.isoformat()),
        "month": args.get("month", today_value.strftime("%Y-%m")),
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
    }


def query_attendance_rows(start_date: str, end_date: str, filters: dict):
    conditions = ["l.lesson_date BETWEEN ? AND ?"]
    params = [start_date, end_date]
    if filters.get("class_id"):
        conditions.append("l.class_id = ?")
        params.append(filters["class_id"])
    if filters.get("teacher_id"):
        conditions.append("l.teacher_id = ?")
        params.append(filters["teacher_id"])
    if filters.get("lesson_type_id"):
        conditions.append("l.lesson_type_id = ?")
        params.append(filters["lesson_type_id"])
    if filters.get("student_id"):
        conditions.append("a.student_id = ?")
        params.append(filters["student_id"])
    if filters.get("attendance_status"):
        conditions.append("a.status = ?")
        params.append(filters["attendance_status"])
    where_sql = " AND ".join(conditions)
    return fetch_all(
        f"""
        SELECT l.id AS lesson_id, l.lesson_date, l.weekday, l.start_time, l.end_time,
               l.lesson_topic, l.lesson_hours, l.classroom, l.status AS lesson_status,
               c.class_name, c.course_name, t.name AS teacher_name, lt.type_name,
               s.name AS student_name, s.parent_phone, a.status AS attendance_status,
               a.deduct_hours, a.checkin_time, a.operator, a.remark
        FROM attendance a
        JOIN lessons l ON l.id = a.lesson_id
        JOIN classes c ON c.id = l.class_id
        JOIN students s ON s.id = a.student_id
        LEFT JOIN teachers t ON t.id = l.teacher_id
        LEFT JOIN lesson_types lt ON lt.id = l.lesson_type_id
        WHERE {where_sql}
        ORDER BY l.lesson_date DESC, l.start_time, c.class_name, s.name
        """,
        tuple(params),
    )


def query_lessons_for_statistics(start_date: str, end_date: str, filters: dict):
    conditions = ["l.lesson_date BETWEEN ? AND ?", "l.status != 'cancelled'"]
    params = [start_date, end_date]
    if filters.get("class_id"):
        conditions.append("l.class_id = ?")
        params.append(filters["class_id"])
    if filters.get("teacher_id"):
        conditions.append("l.teacher_id = ?")
        params.append(filters["teacher_id"])
    if filters.get("lesson_type_id"):
        conditions.append("l.lesson_type_id = ?")
        params.append(filters["lesson_type_id"])
    where_sql = " AND ".join(conditions)
    return fetch_all(
        f"""
        SELECT l.id, l.lesson_hours, c.class_name, t.name AS teacher_name,
               (
                   SELECT COUNT(*)
                   FROM class_students cs
                   JOIN students s ON s.id = cs.student_id
                   WHERE cs.class_id = l.class_id AND cs.status = 'active' AND s.status = 'active'
               ) AS expected_count
        FROM lessons l
        JOIN classes c ON c.id = l.class_id
        LEFT JOIN teachers t ON t.id = l.teacher_id
        WHERE {where_sql}
        """,
        tuple(params),
    )


def build_statistics(attendance_rows, lesson_rows):
    status_counts = defaultdict(int)
    class_stats = {}
    teacher_stats = {}
    consumed = 0
    arrived_status = {"已到", "迟到", "补签", "试听"}
    for row in attendance_rows:
        status_counts[row["attendance_status"]] += 1
        consumed += float(row["deduct_hours"] or 0)
        class_name = row["class_name"] or "未命名班级"
        teacher_name = row["teacher_name"] or "未指定教师"
        class_stats.setdefault(
            class_name,
            {"class_name": class_name, "arrived": 0, "leave": 0, "absent": 0, "deduct_hours": 0},
        )
        teacher_stats.setdefault(
            teacher_name,
            {"teacher_name": teacher_name, "records": 0, "deduct_hours": 0},
        )
        if row["attendance_status"] in arrived_status:
            class_stats[class_name]["arrived"] += 1
        if row["attendance_status"] == "请假":
            class_stats[class_name]["leave"] += 1
        if row["attendance_status"] == "缺勤":
            class_stats[class_name]["absent"] += 1
        class_stats[class_name]["deduct_hours"] += float(row["deduct_hours"] or 0)
        teacher_stats[teacher_name]["records"] += 1
        teacher_stats[teacher_name]["deduct_hours"] += float(row["deduct_hours"] or 0)

    expected_total = sum(int(row["expected_count"] or 0) for row in lesson_rows)
    arrived_total = sum(status_counts[status] for status in arrived_status)
    summary = {
        "lesson_count": len(lesson_rows),
        "expected_total": expected_total,
        "arrived_total": arrived_total,
        "leave_total": status_counts["请假"],
        "absent_total": status_counts["缺勤"],
        "late_total": status_counts["迟到"],
        "trial_total": status_counts["试听"],
        "deduct_hours": consumed,
        "attendance_rate": round(arrived_total / expected_total * 100, 1)
        if expected_total
        else 0,
    }
    return summary, list(class_stats.values()), list(teacher_stats.values())


def export_attendance_workbook(rows, range_info: dict) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "签到明细"
    append_header(
        sheet,
        [
            "日期",
            "星期",
            "班级",
            "课程类型",
            "课程主题",
            "上课老师",
            "学生姓名",
            "签到状态",
            "扣课时",
            "签到时间",
            "备注",
        ],
    )
    for row in rows:
        sheet.append(
            [
                row["lesson_date"],
                WEEKDAY_MAP.get(row["weekday"], ""),
                row["class_name"],
                row["type_name"],
                row["lesson_topic"],
                row["teacher_name"],
                row["student_name"],
                row["attendance_status"],
                row["deduct_hours"],
                row["checkin_time"],
                row["remark"],
            ]
        )
    autosize(sheet)
    filename = f"attendance_{range_info['start_date']}_{range_info['end_date']}_{safe_stamp()}.xlsx"
    file_path = config.EXPORT_DIR / filename
    workbook.save(file_path)
    return file_path


def export_lesson_workbook(lesson, rows) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "本节课签到"
    append_header(
        sheet,
        ["日期", "班级", "课程主题", "学生姓名", "联系电话", "签到状态", "扣课时", "备注"],
    )
    for row in rows:
        sheet.append(
            [
                lesson["lesson_date"],
                lesson["class_name"],
                lesson["lesson_topic"],
                row["name"],
                row["phone"] or row["parent_phone"],
                row["status"],
                row["deduct_hours"],
                row["remark"],
            ]
        )
    autosize(sheet)
    filename = f"lesson_{lesson['id']}_{safe_stamp()}.xlsx"
    file_path = config.EXPORT_DIR / filename
    workbook.save(file_path)
    return file_path


def export_operation_logs_workbook(rows, filters: dict) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "操作日志"
    append_header(
        sheet,
        ["时间", "账号", "角色", "动作", "对象类型", "对象 ID", "详情", "IP 地址"],
    )
    for row in rows:
        sheet.append(
            [
                row["created_at"],
                row["username"],
                role_label(row["role"]),
                row["action"],
                row["target_type"],
                row["target_id"],
                row["detail"],
                row["ip_address"],
            ]
        )
    autosize(sheet)
    start = filters.get("start_date") or "all"
    end = filters.get("end_date") or "all"
    filename = f"operation_logs_{start}_{end}_{safe_stamp()}.xlsx"
    file_path = config.EXPORT_DIR / filename
    workbook.save(file_path)
    return file_path


def append_header(sheet, headings):
    sheet.append(headings)
    fill = PatternFill("solid", fgColor="1F2937")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autosize(sheet):
    for column_cells in sheet.columns:
        length = 8
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value) + 2)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(length, 32)


def safe_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def require_roles(*roles):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            user = g.get("current_user")
            required_permissions = endpoint_permissions(request.endpoint, request.method)
            if user is None:
                abort(403)
            if not matches_legacy_roles(user, roles):
                if not required_permissions or not has_any_permission(user, required_permissions):
                    abort(403)
            return view(*args, **kwargs)

        return wrapped

    return decorator


def ensure_roles(*roles):
    user = g.get("current_user")
    required_permissions = endpoint_permissions(request.endpoint, request.method)
    if user is None:
        abort(403)
    if not matches_legacy_roles(user, roles):
        if not required_permissions or not has_any_permission(user, required_permissions):
            abort(403)
        abort(403)


def log_operation(action: str, target_type: str = None, target_id: int = None, detail: str = ""):
    user = g.get("current_user")
    if user is None and session.get("user_id"):
        user = fetch_one("SELECT * FROM users WHERE id = ?", (session["user_id"],))
    get_db().execute(
        """
        INSERT INTO operation_logs
            (user_id, username, role, action, target_type, target_id, detail,
             ip_address, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            user["id"] if user else None,
            user["username"] if user else "",
            user["role"] if user else "",
            action,
            target_type,
            target_id,
            detail or "",
            request.remote_addr if request else "",
            now_text(),
        ),
    )
    get_db().commit()


def rows_to_dicts(rows) -> list[dict]:
    return [dict(row) for row in rows]


def asset_upload_root_for_scope(usage_scope: str) -> Path:
    scope = (usage_scope or "courseware").strip().lower()
    if scope == "scratch":
        return Path(current_app.config["SCRATCH_UPLOAD_DIR"])
    if scope == "materials":
        return Path(current_app.config["MATERIAL_UPLOAD_DIR"])
    if scope == "courseware":
        return Path(current_app.config["COURSEWARE_UPLOAD_DIR"])
    return Path(current_app.config["UPLOAD_DIR"]) / scope


def infer_asset_type(filename: str, fallback: str = "") -> str:
    suffix = Path(filename or "").suffix.lower()
    if fallback:
        return fallback.strip().lower()
    if suffix in {".sb2", ".sb3"}:
        return "scratch_project"
    if suffix in {".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp"}:
        return "image"
    if suffix in {".mp3", ".wav"}:
        return "audio"
    if suffix in {".ppt", ".pptx", ".pdf", ".doc", ".docx", ".xlsx", ".xlsm"}:
        return "courseware"
    return "file"


def validate_upload_filename(filename: str) -> str:
    suffix = Path(filename or "").suffix.lower()
    if not suffix:
        raise ValueError("上传文件缺少扩展名。")
    if suffix not in current_app.config["ALLOWED_UPLOAD_EXTENSIONS"]:
        raise ValueError(f"不支持的文件类型：{suffix}")
    safe_name = secure_filename(filename)
    if not safe_name:
        safe_name = "upload" + suffix
    return safe_name


def validate_scratch_project_bytes(filename: str, content: bytes):
    suffix = Path(filename or "").suffix.lower()
    if suffix not in {".sb2", ".sb3"}:
        return
    try:
        with zipfile.ZipFile(BytesIO(content)) as archive:
            if "project.json" not in archive.namelist():
                raise ValueError("Scratch 模板缺少 project.json。")
            json.loads(archive.read("project.json").decode("utf-8"))
    except ValueError:
        raise
    except (zipfile.BadZipFile, json.JSONDecodeError, UnicodeDecodeError) as error:
        raise ValueError("Scratch 模板文件无法解析，请上传有效的 .sb3/.sb2 文件。") from error


def validate_scratch_project_upload(upload):
    suffix = Path(upload.filename or "").suffix.lower()
    if suffix not in {".sb2", ".sb3"}:
        return
    position = upload.stream.tell()
    content = upload.read()
    upload.stream.seek(position)
    validate_scratch_project_bytes(upload.filename, content)


def save_uploaded_asset(upload, usage_scope: str = "courseware", asset_type: str = "", metadata: dict | None = None) -> dict:
    safe_name = validate_upload_filename(upload.filename)
    suffix = Path(safe_name).suffix.lower()
    target_dir = asset_upload_root_for_scope(usage_scope)
    target_dir.mkdir(parents=True, exist_ok=True)
    stored_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:10]}{suffix}"
    target_path = target_dir / stored_name
    upload.save(target_path)
    file_size = target_path.stat().st_size
    mime_type = upload.mimetype or mimetypes.guess_type(safe_name)[0] or "application/octet-stream"
    metadata_json = json.dumps(metadata or {}, ensure_ascii=False)
    cursor = get_db().execute(
        """
        INSERT INTO uploaded_assets
            (owner_user_id, asset_type, usage_scope, original_filename, storage_path,
             public_path, mime_type, file_size, status, metadata_json, created_at)
        VALUES (?, ?, ?, ?, ?, '', ?, ?, 'uploaded', ?, ?)
        """,
        (
            g.current_user["id"] if g.get("current_user") else None,
            infer_asset_type(safe_name, asset_type),
            (usage_scope or "courseware").strip().lower(),
            upload.filename,
            str(target_path),
            mime_type,
            file_size,
            metadata_json,
            now_text(),
        ),
    )
    asset_id = cursor.lastrowid
    public_path = url_for("api_download_asset", asset_id=asset_id)
    get_db().execute("UPDATE uploaded_assets SET public_path = ? WHERE id = ?", (public_path, asset_id))
    get_db().commit()
    return get_uploaded_asset(asset_id)


def save_binary_asset(
    filename: str,
    content: bytes,
    usage_scope: str = "scratch",
    asset_type: str = "",
    mime_type: str | None = None,
    metadata: dict | None = None,
) -> dict:
    safe_name = validate_upload_filename(filename)
    suffix = Path(safe_name).suffix.lower()
    target_dir = asset_upload_root_for_scope(usage_scope)
    target_dir.mkdir(parents=True, exist_ok=True)
    stored_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:10]}{suffix}"
    target_path = target_dir / stored_name
    target_path.write_bytes(content)
    cursor = get_db().execute(
        """
        INSERT INTO uploaded_assets
            (owner_user_id, asset_type, usage_scope, original_filename, storage_path,
             public_path, mime_type, file_size, status, metadata_json, created_at)
        VALUES (?, ?, ?, ?, ?, '', ?, ?, 'uploaded', ?, ?)
        """,
        (
            g.current_user["id"] if g.get("current_user") else None,
            infer_asset_type(safe_name, asset_type),
            (usage_scope or "scratch").strip().lower(),
            filename,
            str(target_path),
            mime_type or mimetypes.guess_type(safe_name)[0] or "application/octet-stream",
            len(content),
            json.dumps(metadata or {}, ensure_ascii=False),
            now_text(),
        ),
    )
    asset_id = cursor.lastrowid
    public_path = url_for("api_download_asset", asset_id=asset_id)
    get_db().execute("UPDATE uploaded_assets SET public_path = ? WHERE id = ?", (public_path, asset_id))
    get_db().commit()
    return get_uploaded_asset(asset_id)


def decode_base64_payload(value: str | None, label: str) -> bytes:
    if not value:
        raise ValueError(f"缺少{label}数据。")
    raw = str(value)
    if "," in raw and raw.strip().lower().startswith("data:"):
        raw = raw.split(",", 1)[1]
    try:
        return base64.b64decode(raw, validate=True)
    except Exception as error:
        raise ValueError(f"{label}不是有效的 base64 数据。") from error


def get_uploaded_asset(asset_id: int) -> dict | None:
    row = fetch_one("SELECT * FROM uploaded_assets WHERE id = ?", (asset_id,))
    return dict(row) if row else None


def query_uploaded_assets(filters: dict | None = None):
    filters = filters or {}
    conditions = []
    params = []
    if filters.get("asset_type"):
        conditions.append("a.asset_type = ?")
        params.append(filters["asset_type"])
    if filters.get("usage_scope"):
        conditions.append("a.usage_scope = ?")
        params.append(filters["usage_scope"])
    if filters.get("status"):
        conditions.append("a.status = ?")
        params.append(filters["status"])
    scope = current_access_scope()
    if scope["role"] in {"teacher", "student", "parent"}:
        conditions.append("a.owner_user_id = ?")
        params.append(g.current_user["id"])
    where_sql = "WHERE " + " AND ".join(conditions) if conditions else ""
    return fetch_all(
        f"""
        SELECT a.*, u.display_name AS owner_name
        FROM uploaded_assets a
        LEFT JOIN users u ON u.id = a.owner_user_id
        {where_sql}
        ORDER BY a.created_at DESC, a.id DESC
        LIMIT 200
        """,
        tuple(params),
    )


def query_lesson_assets(lesson_id: int):
    return fetch_all(
        """
        SELECT la.id AS lesson_asset_id, la.lesson_id, la.asset_id, la.title, la.note,
               la.status AS binding_status, la.created_at AS bound_at,
               a.owner_user_id, a.asset_type, a.usage_scope, a.original_filename,
               a.public_path, a.mime_type, a.file_size, a.status AS asset_status,
               u.display_name AS owner_name
        FROM lesson_assets la
        JOIN uploaded_assets a ON a.id = la.asset_id
        LEFT JOIN users u ON u.id = a.owner_user_id
        WHERE la.lesson_id = ? AND la.status = 'active' AND a.status = 'uploaded'
        ORDER BY la.created_at DESC, la.id DESC
        """,
        (lesson_id,),
    )


def get_lesson_asset(lesson_asset_id: int) -> dict | None:
    row = fetch_one(
        """
        SELECT la.id AS lesson_asset_id, la.lesson_id, la.asset_id, la.title, la.note,
               la.status AS binding_status, la.created_at AS bound_at,
               a.owner_user_id, a.asset_type, a.usage_scope, a.original_filename,
               a.public_path, a.mime_type, a.file_size, a.status AS asset_status,
               u.display_name AS owner_name
        FROM lesson_assets la
        JOIN uploaded_assets a ON a.id = la.asset_id
        LEFT JOIN users u ON u.id = a.owner_user_id
        WHERE la.id = ?
        """,
        (lesson_asset_id,),
    )
    return dict(row) if row else None


def create_lesson_asset(lesson_id: int, req) -> dict:
    upload = req.files.get("file")
    if upload is None or not upload.filename:
        raise ValueError("请上传课件文件。")
    title = (req.form.get("title") or "").strip()
    note = (req.form.get("note") or "").strip()
    asset = save_uploaded_asset(
        upload,
        usage_scope="courseware",
        asset_type=req.form.get("asset_type") or "courseware",
        metadata={
            "lesson_id": lesson_id,
            "title": title,
            "note": note,
            "purpose": "lesson_courseware",
        },
    )
    now = now_text()
    get_db().execute(
        """
        INSERT INTO lesson_assets
            (lesson_id, asset_id, title, note, status, created_by, created_at)
        VALUES (?, ?, ?, ?, 'active', ?, ?)
        ON CONFLICT(lesson_id, asset_id) DO UPDATE SET
            title = excluded.title,
            note = excluded.note,
            status = 'active'
        """,
        (
            lesson_id,
            asset["id"],
            title or asset["original_filename"],
            note,
            g.current_user["id"] if g.get("current_user") else None,
            now,
        ),
    )
    get_db().commit()
    row = fetch_one(
        "SELECT id FROM lesson_assets WHERE lesson_id = ? AND asset_id = ?",
        (lesson_id, asset["id"]),
    )
    return get_lesson_asset(row["id"]) if row else asset


def deactivate_lesson_asset(lesson_id: int, lesson_asset_id: int) -> bool:
    cursor = get_db().execute(
        """
        UPDATE lesson_assets
        SET status = 'inactive'
        WHERE id = ? AND lesson_id = ? AND status = 'active'
        """,
        (lesson_asset_id, lesson_id),
    )
    get_db().commit()
    return cursor.rowcount > 0


def can_access_uploaded_asset(asset: dict) -> bool:
    user = g.get("current_user")
    if user is None:
        return False
    role = normalize_role(user["role"])
    asset_id = asset.get("id")
    if not asset_id:
        return False
    owner_id = asset.get("owner_user_id")
    if owner_id and int(owner_id) == int(user["id"]):
        return True
    if role not in {"teacher", "student", "parent"} and has_permission(user, "uploads.manage"):
        return True
    if has_permission(user, "scratch.materials.view"):
        material = fetch_one(
            "SELECT 1 FROM scratch_materials WHERE asset_id = ? AND status = 'active'",
            (asset_id,),
        )
        if material:
            return True
    if role == "teacher" and user.get("teacher_id"):
        row = fetch_one(
            """
            SELECT 1
            FROM lessons l
            LEFT JOIN lesson_assets la
                ON la.lesson_id = l.id AND la.status = 'active'
            LEFT JOIN lesson_scratch_templates lst
                ON lst.lesson_id = l.id AND lst.status = 'active'
            LEFT JOIN scratch_templates st
                ON st.id = lst.template_id AND st.status = 'active'
            LEFT JOIN scratch_works sw
                ON sw.lesson_id = l.id
            WHERE l.teacher_id = ?
              AND l.status != 'cancelled'
              AND (la.asset_id = ?
                   OR st.asset_id = ? OR st.thumbnail_asset_id = ?
                   OR sw.asset_id = ? OR sw.thumbnail_asset_id = ?)
            LIMIT 1
            """,
            (user["teacher_id"], asset_id, asset_id, asset_id, asset_id, asset_id),
        )
        return row is not None
    if role in {"student", "parent"} and user.get("student_id"):
        row = fetch_one(
            """
            SELECT 1
            FROM class_students cs
            JOIN lessons l ON l.class_id = cs.class_id
            LEFT JOIN lesson_assets la
                ON la.lesson_id = l.id AND la.status = 'active'
            LEFT JOIN lesson_scratch_templates lst
                ON lst.lesson_id = l.id AND lst.status = 'active'
            LEFT JOIN scratch_templates st
                ON st.id = lst.template_id AND st.status = 'active'
            WHERE cs.student_id = ?
              AND cs.status = 'active'
              AND l.status != 'cancelled'
              AND (la.asset_id = ?
                   OR st.asset_id = ? OR st.thumbnail_asset_id = ?)
            LIMIT 1
            """,
            (user["student_id"], asset_id, asset_id, asset_id),
        )
        if row:
            return True
        row = fetch_one(
            """
            SELECT 1
            FROM scratch_works
            WHERE student_id = ?
              AND (asset_id = ? OR thumbnail_asset_id = ?)
            LIMIT 1
            """,
            (user["student_id"], asset_id, asset_id),
        )
        return row is not None
    return False


def send_asset_file(asset: dict):
    path = Path(asset["storage_path"]).resolve()
    upload_root = Path(current_app.config["UPLOAD_DIR"]).resolve()
    if upload_root not in path.parents and path != upload_root:
        abort(403)
    if not path.exists():
        abort(404)
    return send_from_directory(
        path.parent,
        path.name,
        as_attachment=True,
        download_name=asset["original_filename"],
        mimetype=asset.get("mime_type") or None,
    )


def scratch_editor_config() -> dict:
    return {
        "mode": current_app.config["HYDRO_INTEGRATION_MODE"],
        "hydro_base_url": current_app.config["HYDRO_BASE_URL"],
        "editor_url": current_app.config["SCRATCH_EDITOR_URL"],
        "oj_api_url": current_app.config["SCRATCH_OJ_API_URL"],
        "editor_project_dir": current_app.config["SCRATCH_EDITOR_PROJECT_DIR"],
    }


def absolute_app_url(path: str | None) -> str:
    if not path:
        return ""
    value = str(path)
    if value.startswith(("http://", "https://")):
        return value
    if not value.startswith("/"):
        value = "/" + value
    return request.host_url.rstrip("/") + value


def origin_from_url(url: str | None) -> str:
    if not url:
        return "*"
    parsed = urlparse(url)
    if not parsed.scheme or not parsed.netloc:
        return "*"
    return f"{parsed.scheme}://{parsed.netloc}"


def query_scratch_templates(filters: dict | None = None):
    filters = filters or {}
    conditions = []
    params = []
    if filters.get("status"):
        conditions.append("st.status = ?")
        params.append(filters["status"])
    if filters.get("q"):
        conditions.append("(st.title LIKE ? OR st.description LIKE ?)")
        term = f"%{filters['q']}%"
        params.extend([term, term])
    where_sql = "WHERE " + " AND ".join(conditions) if conditions else ""
    return fetch_all(
        f"""
        SELECT st.*, a.original_filename AS asset_filename, a.public_path AS asset_url,
               ta.public_path AS thumbnail_url, u.display_name AS creator_name
        FROM scratch_templates st
        LEFT JOIN uploaded_assets a ON a.id = st.asset_id
        LEFT JOIN uploaded_assets ta ON ta.id = st.thumbnail_asset_id
        LEFT JOIN users u ON u.id = st.created_by
        {where_sql}
        ORDER BY st.updated_at DESC, st.id DESC
        """,
        tuple(params),
    )


def get_scratch_template(template_id: int) -> dict | None:
    row = fetch_one(
        """
        SELECT st.*, a.original_filename AS asset_filename, a.public_path AS asset_url,
               ta.public_path AS thumbnail_url
        FROM scratch_templates st
        LEFT JOIN uploaded_assets a ON a.id = st.asset_id
        LEFT JOIN uploaded_assets ta ON ta.id = st.thumbnail_asset_id
        WHERE st.id = ?
        """,
        (template_id,),
    )
    return dict(row) if row else None


def ensure_scratch_template_preview_access(template_id: int):
    user = g.get("current_user")
    if not user:
        abort(403)
    role = normalize_role(user["role"])
    if role in {"student", "parent"}:
        student_id = user.get("student_id")
        if not student_id:
            abort(403)
        row = fetch_one(
            """
            SELECT 1
            FROM lesson_scratch_templates lst
            JOIN lessons l ON l.id = lst.lesson_id
            JOIN class_students cs ON cs.class_id = l.class_id
            WHERE lst.template_id = ?
              AND lst.status = 'active'
              AND l.status != 'cancelled'
              AND cs.student_id = ?
              AND cs.status = 'active'
            LIMIT 1
            """,
            (template_id, student_id),
        )
        if row is None:
            abort(403)


def ensure_scratch_template_update_access(template: dict):
    user = g.get("current_user")
    role = normalize_role(user["role"]) if user else ""
    if role == "teacher":
        if not template.get("created_by") or int(template["created_by"]) != int(user["id"]):
            abort(403)
    elif role in {"student", "parent", "readonly"}:
        abort(403)


def create_scratch_template(req) -> dict:
    data = req.form if req.files else (req.get_json(silent=True) or {})
    file_asset = None
    thumb_asset = None
    if req.files.get("file"):
        validate_scratch_project_upload(req.files["file"])
        file_asset = save_uploaded_asset(
            req.files["file"],
            usage_scope="scratch",
            asset_type="scratch_template",
            metadata={"purpose": "scratch_template"},
        )
    if req.files.get("thumbnail"):
        thumb_asset = save_uploaded_asset(
            req.files["thumbnail"],
            usage_scope="scratch",
            asset_type="image",
            metadata={"purpose": "scratch_template_thumbnail"},
        )
    title = (data.get("title") or (Path(file_asset["original_filename"]).stem if file_asset else "")).strip()
    if not title:
        raise ValueError("请填写模板名称或上传模板文件。")
    asset_id = parse_int(data.get("asset_id")) or (file_asset["id"] if file_asset else None)
    thumbnail_asset_id = parse_int(data.get("thumbnail_asset_id")) or (thumb_asset["id"] if thumb_asset else None)
    now = now_text()
    cursor = get_db().execute(
        """
        INSERT INTO scratch_templates
            (title, description, asset_id, thumbnail_asset_id, source_type, editor_url,
             status, created_by, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            title,
            (data.get("description") or "").strip(),
            asset_id,
            thumbnail_asset_id,
            (data.get("source_type") or "uploaded").strip(),
            (data.get("editor_url") or "").strip(),
            data.get("status") or "active",
            g.current_user["id"] if g.get("current_user") else None,
            now,
            now,
        ),
    )
    get_db().commit()
    return get_scratch_template(cursor.lastrowid)


def update_scratch_template(template_id: int, data) -> dict | None:
    template = get_scratch_template(template_id)
    if template is None:
        return None
    ensure_scratch_template_update_access(template)
    title = (data.get("title") or template["title"]).strip()
    if not title:
        raise ValueError("请填写模板名称。")
    get_db().execute(
        """
        UPDATE scratch_templates
        SET title = ?, description = ?, asset_id = ?, thumbnail_asset_id = ?,
            source_type = ?, editor_url = ?, status = ?, updated_at = ?
        WHERE id = ?
        """,
        (
            title,
            (data.get("description") if data.get("description") is not None else template.get("description") or "").strip(),
            parse_int(data.get("asset_id")) or template.get("asset_id"),
            parse_int(data.get("thumbnail_asset_id")) or template.get("thumbnail_asset_id"),
            (data.get("source_type") or template.get("source_type") or "uploaded").strip(),
            (data.get("editor_url") if data.get("editor_url") is not None else template.get("editor_url") or "").strip(),
            data.get("status") or template.get("status") or "active",
            now_text(),
            template_id,
        ),
    )
    get_db().commit()
    return get_scratch_template(template_id)


def bind_lesson_scratch_template(lesson_id: int, template_id: int, bind_note: str = "") -> dict:
    template = get_scratch_template(template_id)
    if template is None or template["status"] != "active":
        raise ValueError("Scratch 模板不存在或未启用。")
    now = now_text()
    get_db().execute(
        """
        INSERT INTO lesson_scratch_templates
            (lesson_id, template_id, bind_note, status, created_by, created_at)
        VALUES (?, ?, ?, 'active', ?, ?)
        ON CONFLICT(lesson_id, template_id) DO UPDATE SET
            bind_note = excluded.bind_note,
            status = 'active'
        """,
        (
            lesson_id,
            template_id,
            (bind_note or "").strip(),
            g.current_user["id"] if g.get("current_user") else None,
            now,
        ),
    )
    get_db().commit()
    return dict(fetch_one("SELECT * FROM lesson_scratch_templates WHERE lesson_id = ? AND template_id = ?", (lesson_id, template_id)))


def parse_json_payload_field(data, key: str):
    raw = data.get(key)
    if raw in (None, ""):
        return None
    if isinstance(raw, (dict, list)):
        return raw
    try:
        return json.loads(str(raw))
    except json.JSONDecodeError as error:
        raise ValueError(f"{key} 不是有效 JSON。") from error


def normalize_judge_config(config_payload, title: str = "", total_score: float = 100) -> dict | None:
    if not config_payload:
        return None
    if not isinstance(config_payload, dict):
        raise ValueError("judge_config 必须是 JSON 对象。")
    config_payload = dict(config_payload)
    config_payload.setdefault("title", title)
    config_payload["totalScore"] = parse_float(config_payload.get("totalScore"), total_score)
    for key in ("staticChecks", "dynamicChecks"):
        if config_payload.get(key) is None:
            continue
        if not isinstance(config_payload[key], list):
            raise ValueError(f"{key} 必须是数组。")
    return config_payload


def assignment_payload_from_data(data, fallback_title: str = "") -> dict:
    title = (data.get("assignment_title") or data.get("title") or fallback_title or "").strip()
    statement_md = (data.get("statement_md") or data.get("description") or data.get("bind_note") or "").strip()
    bind_note = (data.get("bind_note") or data.get("description") or "").strip()
    due_at = (data.get("due_at") or "").strip()
    test_point_spec = parse_json_payload_field(data, "test_point_spec_json") or parse_json_payload_field(data, "test_point_spec")
    judge_config = parse_json_payload_field(data, "judge_config_json") or parse_json_payload_field(data, "judge_config")
    if test_point_spec and not judge_config:
        judge_config = generate_judge_config_from_spec(test_point_spec)
    max_score = parse_float(data.get("max_score"), None)
    if max_score is None and isinstance(judge_config, dict):
        max_score = parse_float(judge_config.get("totalScore"), 100)
    if max_score is None:
        max_score = 100
    judge_config = normalize_judge_config(judge_config, title=title, total_score=max_score)
    return {
        "assignment_title": title,
        "statement_md": statement_md,
        "bind_note": bind_note,
        "due_at": due_at,
        "judge_config_json": json.dumps(judge_config, ensure_ascii=False) if judge_config else "",
        "test_point_spec_json": json.dumps(test_point_spec, ensure_ascii=False) if test_point_spec else "",
        "auto_judge": 1 if parse_bool(data.get("auto_judge"), False) else 0,
        "max_score": max_score,
    }


def create_lesson_scratch_assignment(lesson_id: int, req) -> dict:
    data = req.form if req.files else (req.get_json(silent=True) or {})
    template_id = parse_int(data.get("template_id"))
    if template_id is None:
        if not req.files.get("file"):
            raise ValueError("请选择已有模板或上传 .sb3 模板文件。")
        template = create_scratch_template(req)
        template_id = template["id"]
    template = get_scratch_template(template_id)
    if template is None or template["status"] != "active":
        raise ValueError("Scratch 模板不存在或未启用。")
    bind_lesson_scratch_template(lesson_id, template_id, data.get("bind_note", ""))
    return update_lesson_scratch_assignment(lesson_id, template_id, data, template_title=template["title"])


def update_lesson_scratch_assignment(
    lesson_id: int,
    template_id: int,
    data,
    template_title: str | None = None,
) -> dict | None:
    binding = fetch_one(
        "SELECT * FROM lesson_scratch_templates WHERE lesson_id = ? AND template_id = ?",
        (lesson_id, template_id),
    )
    if binding is None:
        return None
    if template_title is None:
        template = get_scratch_template(template_id)
        template_title = template["title"] if template else ""
    payload = assignment_payload_from_data(data, fallback_title=template_title)
    now = now_text()
    get_db().execute(
        """
        UPDATE lesson_scratch_templates
        SET assignment_title = ?, statement_md = ?, bind_note = ?, due_at = ?,
            judge_config_json = ?, test_point_spec_json = ?, auto_judge = ?,
            max_score = ?, published_at = COALESCE(published_at, ?), status = 'active'
        WHERE lesson_id = ? AND template_id = ?
        """,
        (
            payload["assignment_title"],
            payload["statement_md"],
            payload["bind_note"],
            payload["due_at"],
            payload["judge_config_json"],
            payload["test_point_spec_json"],
            payload["auto_judge"],
            payload["max_score"],
            now,
            lesson_id,
            template_id,
        ),
    )
    get_db().commit()
    return get_lesson_scratch_assignment(lesson_id, template_id)


def get_lesson_scratch_assignment(lesson_id: int, template_id: int) -> dict | None:
    row = fetch_one(
        """
        SELECT lst.*, st.title, st.description, st.asset_id, st.thumbnail_asset_id,
               st.editor_url, a.public_path AS asset_url, ta.public_path AS thumbnail_url
        FROM lesson_scratch_templates lst
        JOIN scratch_templates st ON st.id = lst.template_id
        LEFT JOIN uploaded_assets a ON a.id = st.asset_id
        LEFT JOIN uploaded_assets ta ON ta.id = st.thumbnail_asset_id
        WHERE lst.lesson_id = ? AND lst.template_id = ?
        """,
        (lesson_id, template_id),
    )
    return dict(row) if row else None


def query_lesson_scratch_templates(lesson_id: int):
    return fetch_all(
        """
        SELECT lst.*, st.title, st.description, st.asset_id, st.thumbnail_asset_id,
               st.editor_url, a.public_path AS asset_url, ta.public_path AS thumbnail_url,
               COUNT(sw.id) AS work_count,
               SUM(CASE WHEN sw.status IN ('submitted', 'reviewed') THEN 1 ELSE 0 END) AS submitted_count
        FROM lesson_scratch_templates lst
        JOIN scratch_templates st ON st.id = lst.template_id
        LEFT JOIN uploaded_assets a ON a.id = st.asset_id
        LEFT JOIN uploaded_assets ta ON ta.id = st.thumbnail_asset_id
        LEFT JOIN scratch_works sw
          ON sw.lesson_id = lst.lesson_id
         AND sw.template_id = lst.template_id
        WHERE lst.lesson_id = ? AND lst.status = 'active' AND st.status = 'active'
        GROUP BY lst.id
        ORDER BY lst.created_at DESC, lst.id DESC
        """,
        (lesson_id,),
    )


def student_can_access_lesson(student_id: int, lesson_id: int) -> bool:
    row = fetch_one(
        """
        SELECT 1
        FROM lessons l
        JOIN class_students cs ON cs.class_id = l.class_id
        WHERE l.id = ?
          AND cs.student_id = ?
          AND cs.status = 'active'
          AND l.status != 'cancelled'
        """,
        (lesson_id, student_id),
    )
    return row is not None


def create_student_scratch_work(lesson_id: int, template_id: int, student_id: int, owner_user_id: int) -> dict:
    if not student_can_access_lesson(student_id, lesson_id):
        raise ValueError("当前学生不在该课次班级中。")
    binding = fetch_one(
        """
        SELECT 1
        FROM lesson_scratch_templates
        WHERE lesson_id = ? AND template_id = ? AND status = 'active'
        """,
        (lesson_id, template_id),
    )
    if binding is None:
        raise ValueError("该课次未绑定这个模板。")
    template = get_scratch_template(template_id)
    if template is None:
        raise ValueError("Scratch 模板不存在。")
    now = now_text()
    get_db().execute(
        """
        INSERT OR IGNORE INTO scratch_works
            (lesson_id, template_id, student_id, owner_user_id, title, status,
             asset_id, thumbnail_asset_id, editor_url, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, 'draft', ?, ?, ?, ?, ?)
        """,
        (
            lesson_id,
            template_id,
            student_id,
            owner_user_id,
            template["title"],
            template.get("asset_id"),
            template.get("thumbnail_asset_id"),
            template.get("editor_url") or url_for("scratch_editor", template_id=template_id),
            now,
            now,
        ),
    )
    get_db().commit()
    work = fetch_one(
        """
        SELECT *
        FROM scratch_works
        WHERE lesson_id = ? AND template_id = ? AND student_id = ?
        """,
        (lesson_id, template_id, student_id),
    )
    work_id = work["id"]
    get_db().execute(
        "UPDATE scratch_works SET editor_url = ? WHERE id = ? AND (editor_url = '' OR editor_url IS NULL)",
        (url_for("scratch_editor", work_id=work_id), work_id),
    )
    get_db().commit()
    return get_scratch_work(work_id)


def get_scratch_work(work_id: int) -> dict | None:
    row = fetch_one(
        """
        SELECT sw.*, s.name AS student_name, st.title AS template_title,
               l.lesson_date, l.start_time, l.end_time, c.class_name,
               a.public_path AS asset_url, ta.public_path AS thumbnail_url,
               reviewer.display_name AS reviewer_name
        FROM scratch_works sw
        JOIN students s ON s.id = sw.student_id
        LEFT JOIN scratch_templates st ON st.id = sw.template_id
        LEFT JOIN lessons l ON l.id = sw.lesson_id
        LEFT JOIN classes c ON c.id = l.class_id
        LEFT JOIN uploaded_assets a ON a.id = sw.asset_id
        LEFT JOIN uploaded_assets ta ON ta.id = sw.thumbnail_asset_id
        LEFT JOIN users reviewer ON reviewer.id = sw.reviewed_by
        WHERE sw.id = ?
        """,
        (work_id,),
    )
    return dict(row) if row else None


def query_scratch_works(filters: dict | None = None):
    filters = filters or {}
    conditions = []
    params = []
    if filters.get("lesson_id"):
        conditions.append("sw.lesson_id = ?")
        params.append(filters["lesson_id"])
    if filters.get("student_id"):
        conditions.append("sw.student_id = ?")
        params.append(filters["student_id"])
    if filters.get("status"):
        conditions.append("sw.status = ?")
        params.append(filters["status"])
    where_sql = "WHERE " + " AND ".join(conditions) if conditions else ""
    return fetch_all(
        f"""
        SELECT sw.*, s.name AS student_name, st.title AS template_title,
               l.lesson_date, l.start_time, l.end_time, c.class_name,
               a.public_path AS asset_url, ta.public_path AS thumbnail_url,
               reviewer.display_name AS reviewer_name
        FROM scratch_works sw
        JOIN students s ON s.id = sw.student_id
        LEFT JOIN scratch_templates st ON st.id = sw.template_id
        LEFT JOIN lessons l ON l.id = sw.lesson_id
        LEFT JOIN classes c ON c.id = l.class_id
        LEFT JOIN uploaded_assets a ON a.id = sw.asset_id
        LEFT JOIN uploaded_assets ta ON ta.id = sw.thumbnail_asset_id
        LEFT JOIN users reviewer ON reviewer.id = sw.reviewed_by
        {where_sql}
        ORDER BY sw.updated_at DESC, sw.id DESC
        """,
        tuple(params),
    )


def query_student_scratch_tasks(student_id: int):
    return fetch_all(
        """
        SELECT lst.lesson_id, lst.template_id,
               COALESCE(NULLIF(lst.assignment_title, ''), st.title) AS title,
               COALESCE(NULLIF(lst.statement_md, ''), st.description) AS description,
               lst.bind_note, lst.due_at, lst.max_score, lst.auto_judge,
               l.lesson_date, l.start_time, l.end_time, c.class_name,
               sw.id AS work_id, sw.status AS work_status, sw.score, sw.review_comment,
               sw.judge_status, sw.judge_score, sw.visibility, sw.like_count, sw.favorite_count,
               sw.submitted_at, sw.reviewed_at
        FROM lesson_scratch_templates lst
        JOIN scratch_templates st ON st.id = lst.template_id
        JOIN lessons l ON l.id = lst.lesson_id
        JOIN classes c ON c.id = l.class_id
        JOIN class_students cs ON cs.class_id = c.id
        LEFT JOIN scratch_works sw
          ON sw.lesson_id = lst.lesson_id
         AND sw.template_id = lst.template_id
         AND sw.student_id = cs.student_id
        WHERE cs.student_id = ?
          AND cs.status = 'active'
          AND lst.status = 'active'
          AND st.status = 'active'
          AND l.status != 'cancelled'
        ORDER BY l.lesson_date DESC, l.start_time DESC
        LIMIT 12
        """,
        (student_id,),
    )


def ensure_scratch_work_access(work: dict, write: bool = False):
    user = g.get("current_user")
    role = normalize_role(user["role"]) if user else ""
    if role == "student":
        if not user.get("student_id") or int(user["student_id"]) != int(work["student_id"]):
            abort(403)
        if write and work["status"] == "reviewed":
            abort(403)
        return
    if role == "parent":
        if not user.get("student_id") or int(user["student_id"]) != int(work["student_id"]) or write:
            abort(403)
        return
    if role == "teacher" and work.get("lesson_id"):
        lesson = get_lesson(work["lesson_id"])
        ensure_lesson_access(lesson)


def ensure_scratch_work_review_access(work: dict):
    lesson = get_lesson(work["lesson_id"]) if work.get("lesson_id") else None
    if lesson is not None:
        ensure_lesson_access(lesson)


def save_scratch_work_progress(work: dict, req) -> dict:
    data = req.form if req.files else (req.get_json(silent=True) or {})
    asset_id = parse_int(data.get("asset_id")) or work.get("asset_id")
    thumbnail_asset_id = parse_int(data.get("thumbnail_asset_id")) or work.get("thumbnail_asset_id")
    if req.files.get("file"):
        validate_scratch_project_upload(req.files["file"])
        asset = save_uploaded_asset(
            req.files["file"],
            usage_scope="scratch",
            asset_type="scratch_work",
            metadata={"work_id": work["id"], "purpose": "scratch_work"},
        )
        asset_id = asset["id"]
    if req.files.get("thumbnail"):
        thumbnail = save_uploaded_asset(
            req.files["thumbnail"],
            usage_scope="scratch",
            asset_type="image",
            metadata={"work_id": work["id"], "purpose": "scratch_work_thumbnail"},
        )
        thumbnail_asset_id = thumbnail["id"]
    get_db().execute(
        """
        UPDATE scratch_works
        SET title = ?, asset_id = ?, thumbnail_asset_id = ?, editor_url = ?,
            submit_note = ?, updated_at = ?
        WHERE id = ?
        """,
        (
            (data.get("title") or work["title"]).strip(),
            asset_id,
            thumbnail_asset_id,
            (data.get("editor_url") or work.get("editor_url") or "").strip(),
            (data.get("submit_note") or work.get("submit_note") or "").strip(),
            now_text(),
            work["id"],
        ),
    )
    get_db().commit()
    return get_scratch_work(work["id"])


def save_scratch_work_from_editor(work: dict, data: dict) -> dict:
    project_base64 = data.get("project_base64") or data.get("sb3_base64") or data.get("project")
    project_bytes = decode_base64_payload(project_base64, "Scratch 作品")
    title = (data.get("title") or work["title"] or f"scratch-work-{work['id']}").strip()
    validate_scratch_project_bytes(f"work-{work['id']}.sb3", project_bytes)
    asset = save_binary_asset(
        f"work-{work['id']}.sb3",
        project_bytes,
        usage_scope="scratch",
        asset_type="scratch_work",
        mime_type="application/octet-stream",
        metadata={"work_id": work["id"], "purpose": "editor_save"},
    )
    thumbnail_asset_id = work.get("thumbnail_asset_id")
    thumbnail_base64 = data.get("thumbnail_base64") or data.get("thumbnail")
    if thumbnail_base64:
        thumbnail = save_binary_asset(
            f"work-{work['id']}-thumbnail.png",
            decode_base64_payload(thumbnail_base64, "Scratch 缩略图"),
            usage_scope="scratch",
            asset_type="image",
            mime_type="image/png",
            metadata={"work_id": work["id"], "purpose": "editor_thumbnail"},
        )
        thumbnail_asset_id = thumbnail["id"]
    get_db().execute(
        """
        UPDATE scratch_works
        SET title = ?, asset_id = ?, thumbnail_asset_id = ?, editor_url = ?,
            submit_note = ?, updated_at = ?
        WHERE id = ?
        """,
        (
            title,
            asset["id"],
            thumbnail_asset_id,
            (data.get("editor_url") or work.get("editor_url") or "").strip(),
            (data.get("submit_note") or work.get("submit_note") or "").strip(),
            now_text(),
            work["id"],
        ),
    )
    get_db().commit()
    return get_scratch_work(work["id"])


def submit_scratch_work(work_id: int, data) -> dict:
    get_db().execute(
        """
        UPDATE scratch_works
        SET status = 'submitted',
            submit_note = ?,
            submitted_at = COALESCE(submitted_at, ?),
            updated_at = ?
        WHERE id = ?
        """,
        ((data.get("submit_note") or "").strip(), now_text(), now_text(), work_id),
    )
    get_db().commit()
    work = get_scratch_work(work_id)
    award_student_points(work["student_id"], "scratch_submit", work_id, 5, "提交 Scratch 作品")
    assignment = get_assignment_for_work(work)
    if assignment and assignment.get("auto_judge") and assignment.get("judge_config_json") and work.get("asset_id"):
        try:
            run_scratch_work_judge(work)
        except ValueError:
            pass
    return get_scratch_work(work_id)


def review_scratch_work(work_id: int, data) -> dict:
    score = parse_float(data.get("score"), None)
    work = get_scratch_work(work_id)
    get_db().execute(
        """
        UPDATE scratch_works
        SET status = 'reviewed',
            review_comment = ?,
            score = ?,
            reviewed_by = ?,
            reviewed_at = ?,
            updated_at = ?
        WHERE id = ?
        """,
        (
            (data.get("review_comment") or data.get("comment") or "").strip(),
            score,
            g.current_user["id"] if g.get("current_user") else None,
            now_text(),
            now_text(),
            work_id,
        ),
    )
    get_db().commit()
    if work and score is not None:
        award_student_points(work["student_id"], "scratch_review", work_id, max(1, round(score / 10)), "教师点评 Scratch 作品")
    return get_scratch_work(work_id)


def get_assignment_for_work(work: dict) -> dict | None:
    if not work.get("lesson_id") or not work.get("template_id"):
        return None
    return get_lesson_scratch_assignment(work["lesson_id"], work["template_id"])


def query_scratch_judge_runs(work_id: int):
    return fetch_all(
        """
        SELECT *
        FROM scratch_judge_runs
        WHERE work_id = ?
        ORDER BY created_at DESC, id DESC
        """,
        (work_id,),
    )


def read_scratch_project(asset: dict) -> dict:
    path = Path(asset["storage_path"]).resolve()
    upload_root = Path(current_app.config["UPLOAD_DIR"]).resolve()
    if upload_root not in path.parents and path != upload_root:
        abort(403)
    if not path.exists():
        raise ValueError("作品文件不存在，请先保存作品。")
    try:
        with zipfile.ZipFile(path) as archive:
            with archive.open("project.json") as project_file:
                return json.loads(project_file.read().decode("utf-8"))
    except KeyError as error:
        raise ValueError(".sb3 文件中缺少 project.json。") from error
    except (zipfile.BadZipFile, json.JSONDecodeError, UnicodeDecodeError) as error:
        raise ValueError(".sb3 文件无法解析。") from error


def collect_scratch_project_meta(project: dict) -> dict:
    targets = project.get("targets") or []
    sprites = [target for target in targets if not target.get("isStage")]
    stage = next((target for target in targets if target.get("isStage")), None)
    variable_names = set()
    list_names = set()
    broadcast_names = set()
    block_opcodes = []
    for target in targets:
        for variable in (target.get("variables") or {}).values():
            if isinstance(variable, list) and variable:
                variable_names.add(str(variable[0]))
        for item in (target.get("lists") or {}).values():
            if isinstance(item, list) and item:
                list_names.add(str(item[0]))
        for broadcast in (target.get("broadcasts") or {}).values():
            if broadcast:
                broadcast_names.add(str(broadcast))
        for block in (target.get("blocks") or {}).values():
            if not isinstance(block, dict):
                continue
            opcode = block.get("opcode")
            if opcode:
                block_opcodes.append(str(opcode))
            fields = block.get("fields") or {}
            if opcode in {"event_broadcast", "event_broadcastandwait", "event_whenbroadcastreceived"}:
                for field in fields.values():
                    if isinstance(field, list) and field and isinstance(field[0], str):
                        broadcast_names.add(field[0])
    return {
        "stageName": stage.get("name") if stage else None,
        "spriteNames": [target.get("name") for target in sprites if target.get("name")],
        "variableNames": sorted(variable_names),
        "listNames": sorted(list_names),
        "broadcastNames": sorted(broadcast_names),
        "blockOpcodes": block_opcodes,
        "blockCount": len(block_opcodes),
    }


def summarize_opcodes(opcodes: list[str]) -> list[str]:
    counts = defaultdict(int)
    for opcode in opcodes:
        counts[opcode] += 1
    return [
        f"{opcode}: {count}"
        for opcode, count in sorted(counts.items(), key=lambda item: item[1], reverse=True)[:12]
    ]


def build_judge_detail(check: dict, passed: bool, message: str, evidence=None) -> dict:
    score = parse_float(check.get("score"), 0)
    return {
        "name": check.get("name") or check.get("type") or "check",
        "type": check.get("type"),
        "passed": passed,
        "score": score if passed else 0,
        "maxScore": score,
        "message": message,
        "hint": None if passed else check.get("hint"),
        "evidence": evidence or [],
    }


def evaluate_static_check(check: dict, meta: dict) -> dict:
    check_type = check.get("type")
    if check_type == "sprite_exists":
        sprite = str(check.get("sprite") or "")
        return build_judge_detail(check, sprite in meta["spriteNames"], f"角色检查：{sprite}", meta["spriteNames"])
    if check_type == "variable_exists":
        variable = str(check.get("variable") or "")
        return build_judge_detail(check, variable in meta["variableNames"], f"变量检查：{variable}", meta["variableNames"])
    if check_type == "list_exists":
        list_name = str(check.get("list") or "")
        return build_judge_detail(check, list_name in meta["listNames"], f"列表检查：{list_name}", meta["listNames"])
    if check_type == "broadcast_exists":
        broadcast = str(check.get("broadcast") or "")
        return build_judge_detail(check, broadcast in meta["broadcastNames"], f"广播检查：{broadcast}", meta["broadcastNames"])
    if check_type == "block_exists":
        opcode = str(check.get("opcode") or "")
        return build_judge_detail(check, opcode in meta["blockOpcodes"], f"积木检查：{opcode}", summarize_opcodes(meta["blockOpcodes"]))
    if check_type == "block_exists_any":
        opcodes = [str(item) for item in check.get("opcodes") or []]
        matched = [opcode for opcode in opcodes if opcode in meta["blockOpcodes"]]
        return build_judge_detail(check, bool(matched), f"任一积木检查：{', '.join(opcodes)}", matched or summarize_opcodes(meta["blockOpcodes"]))
    if check_type == "forbidden_block_absent":
        opcode = str(check.get("opcode") or "")
        return build_judge_detail(check, opcode not in meta["blockOpcodes"], f"禁用积木检查：{opcode}", summarize_opcodes(meta["blockOpcodes"]))
    if check_type == "min_block_count":
        opcode = str(check.get("opcode") or "")
        count = meta["blockOpcodes"].count(opcode)
        required = parse_int(check.get("count")) or 0
        return build_judge_detail(check, count >= required, f"积木数量 {opcode}: {count}/{required}", [f"{opcode}: {count}"])
    return build_judge_detail(check, False, f"暂不支持的静态检查：{check_type}")


def create_manual_required_detail(check: dict) -> dict:
    score = parse_float(check.get("score"), 0)
    return {
        "name": check.get("name") or check.get("type") or "dynamic_check",
        "type": check.get("type") or "dynamic_check",
        "passed": False,
        "score": 0,
        "maxScore": score,
        "message": "该测试点需要 scratch_oj 运行时测评或教师手动复核。",
        "hint": check.get("hint"),
        "evidence": [],
    }


def run_scratch_work_judge(work: dict) -> dict:
    assignment = get_assignment_for_work(work)
    if assignment is None:
        raise ValueError("该作品没有绑定的 Scratch 作业配置。")
    config_raw = assignment.get("judge_config_json") or ""
    if not config_raw:
        result = {
            "status": "manual_review",
            "totalScore": 0,
            "maxScore": parse_float(assignment.get("max_score"), 100),
            "passed": False,
            "details": [],
            "summary": {"passedChecks": 0, "totalChecks": 0, "rawScore": 0, "rawMaxScore": 0},
            "projectMeta": {},
            "message": "未配置自动测评规则，已进入手动复核。",
        }
        save_scratch_judge_result(work, result)
        return result
    try:
        judge_config = json.loads(config_raw)
    except json.JSONDecodeError as error:
        raise ValueError("作业测评配置不是有效 JSON。") from error
    if not work.get("asset_id"):
        raise ValueError("作品尚未保存 .sb3 文件，无法测评。")
    asset = get_uploaded_asset(work["asset_id"])
    if asset is None:
        raise ValueError("作品文件记录不存在。")
    project = read_scratch_project(asset)
    meta = collect_scratch_project_meta(project)
    static_checks = judge_config.get("staticChecks") or []
    dynamic_checks = judge_config.get("dynamicChecks") or []
    details = [evaluate_static_check(check, meta) for check in static_checks]
    details.extend(create_manual_required_detail(check) for check in dynamic_checks)
    raw_score = sum(parse_float(item.get("score"), 0) for item in details)
    raw_max_score = sum(parse_float(item.get("maxScore"), 0) for item in details)
    max_score = parse_float(judge_config.get("totalScore"), parse_float(assignment.get("max_score"), raw_max_score or 100))
    total_score = round((raw_score / raw_max_score) * max_score, 1) if raw_max_score else 0
    static_passed = all(item["passed"] for item in details[: len(static_checks)])
    status = "manual_review" if dynamic_checks else ("passed" if static_passed and total_score >= max_score else "failed")
    result = {
        "status": status,
        "totalScore": total_score,
        "maxScore": max_score,
        "passed": status == "passed",
        "details": details,
        "summary": {
            "passedChecks": len([item for item in details if item["passed"]]),
            "totalChecks": len(details),
            "rawScore": raw_score,
            "rawMaxScore": raw_max_score,
        },
        "projectMeta": meta,
    }
    save_scratch_judge_result(work, result)
    return result


def save_scratch_judge_result(work: dict, result: dict):
    now = now_text()
    detail_json = json.dumps(result, ensure_ascii=False)
    db = get_db()
    db.execute(
        """
        INSERT INTO scratch_judge_runs
            (work_id, lesson_id, template_id, student_id, status, total_score,
             max_score, passed, detail_json, created_by, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            work["id"],
            work.get("lesson_id"),
            work.get("template_id"),
            work["student_id"],
            result["status"],
            result["totalScore"],
            result["maxScore"],
            1 if result.get("passed") else 0,
            detail_json,
            g.current_user["id"] if g.get("current_user") else None,
            now,
        ),
    )
    db.execute(
        """
        UPDATE scratch_works
        SET judge_status = ?, judge_score = ?, judge_detail_json = ?,
            score = CASE WHEN score IS NULL THEN ? ELSE score END,
            updated_at = ?
        WHERE id = ?
        """,
        (result["status"], result["totalScore"], detail_json, result["totalScore"], now, work["id"]),
    )
    db.commit()
    if result.get("passed"):
        award_student_points(work["student_id"], "scratch_judge_passed", work["id"], 20, "Scratch 自动测评通过")


def award_student_points(student_id: int, source_type: str, source_id: int, points: int, reason: str):
    get_db().execute(
        """
        INSERT OR IGNORE INTO student_points_ledger
            (student_id, source_type, source_id, points, reason, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (student_id, source_type, source_id, int(points), reason, now_text()),
    )
    get_db().commit()


def generate_judge_config_from_spec(spec: dict) -> dict:
    if isinstance(spec, dict) and isinstance(spec.get("spec"), dict):
        spec = spec["spec"]
    if not isinstance(spec, dict):
        raise ValueError("测试点规格必须是 JSON 对象。")
    title = str(spec.get("title") or spec.get("problemId") or "Scratch 作业测评")
    total_score = parse_float(spec.get("totalScore"), 100)
    config = {
        "problemId": spec.get("problemId") or f"scratch-{uuid.uuid4().hex[:8]}",
        "title": title,
        "totalScore": total_score,
        "staticChecks": [],
        "dynamicChecks": [],
        "dynamicOptions": spec.get("dynamicOptions") or {},
    }
    raw_static_checks = spec.get("staticChecks") or spec.get("static_checks") or []
    raw_dynamic_checks = spec.get("dynamicChecks") or spec.get("dynamic_checks") or []
    if not isinstance(raw_static_checks, list) or not isinstance(raw_dynamic_checks, list):
        raise ValueError("staticChecks/dynamicChecks 必须是数组。")
    for check in raw_static_checks:
        config["staticChecks"].append(static_check_from_point(check))
    for check in raw_dynamic_checks:
        config["dynamicChecks"].append(dynamic_check_from_point(check))

    points = spec.get("testPoints") or spec.get("test_points") or spec.get("points") or []
    if not isinstance(points, list):
        raise ValueError("测试点列表必须是数组。")
    for point in points:
        expand_test_point(config, point)
    config["staticChecks"] = [item for item in config["staticChecks"] if item]
    config["dynamicChecks"] = [item for item in config["dynamicChecks"] if item]
    return config


def static_check_from_point(point: dict, check_type: str | None = None) -> dict:
    item = dict(point.get("check") or point)
    item["type"] = check_type or item.get("type") or item.get("kind")
    item.pop("kind", None)
    item.setdefault("name", item["type"])
    item.setdefault("score", 0)
    return item


def dynamic_check_from_point(point: dict, check_type: str | None = None) -> dict:
    item = dict(point.get("check") or point)
    item["type"] = check_type or item.get("type") or item.get("kind")
    item.pop("kind", None)
    item.setdefault("name", item["type"])
    item.setdefault("score", 0)
    return item


def expand_test_point(config: dict, point: dict):
    if not isinstance(point, dict):
        return
    kind = point.get("kind") or point.get("type")
    direct_static = {
        "sprite_exists",
        "variable_exists",
        "list_exists",
        "broadcast_exists",
        "block_exists",
        "block_exists_any",
        "forbidden_block_absent",
        "min_block_count",
    }
    direct_dynamic = {"runtime_runs", "variable_value"}
    if kind == "raw_static":
        config["staticChecks"].append(static_check_from_point(point.get("check") or {}))
    elif kind == "raw_dynamic":
        config["dynamicChecks"].append(dynamic_check_from_point(point.get("check") or {}))
    elif kind in direct_static:
        config["staticChecks"].append(static_check_from_point(point, kind))
    elif kind in direct_dynamic:
        config["dynamicChecks"].append(dynamic_check_from_point(point, kind))
    elif kind == "green_flag_sets_variable":
        variable = point.get("variable")
        config["staticChecks"].extend(
            [
                {"type": "variable_exists", "name": f"存在变量 {variable}", "variable": variable, "score": parse_float(point.get("variableScore"), 0), "hint": point.get("hint")},
                {"type": "block_exists", "name": "使用绿旗事件", "opcode": "event_whenflagclicked", "score": parse_float(point.get("greenFlagScore"), 0)},
                {"type": "block_exists", "name": "设置变量", "opcode": "data_setvariableto", "score": parse_float(point.get("setVariableScore"), 0)},
            ]
        )
        config["dynamicChecks"].append(
            {
                "type": "variable_value",
                "name": f"绿旗后 {variable} 达到目标值",
                "variable": variable,
                "expected": point.get("expected"),
                "score": parse_float(point.get("score"), 0),
                "steps": [{"action": "green_flag"}, {"action": "wait", "ms": 300}],
                "hint": point.get("hint"),
            }
        )
    elif kind == "key_changes_variable":
        variable = point.get("variable")
        key = point.get("key") or "space"
        config["staticChecks"].extend(
            [
                {"type": "variable_exists", "name": f"存在变量 {variable}", "variable": variable, "score": parse_float(point.get("variableScore"), 0), "hint": point.get("hint")},
                {"type": "block_exists", "name": f"使用 {key} 按键事件", "opcode": "event_whenkeypressed", "score": parse_float(point.get("keyEventScore"), 0)},
                {"type": "block_exists", "name": "改变变量", "opcode": "data_changevariableby", "score": parse_float(point.get("changeVariableScore"), 0)},
            ]
        )
        if "initialExpected" in point:
            config["dynamicChecks"].append(
                {
                    "type": "variable_value",
                    "name": f"绿旗后 {variable} 初始值正确",
                    "variable": variable,
                    "expected": point.get("initialExpected"),
                    "score": parse_float(point.get("initialScore"), 0),
                    "steps": [{"action": "green_flag"}, {"action": "wait", "ms": 300}],
                }
            )
        config["dynamicChecks"].append(
            {
                "type": "variable_value",
                "name": f"按键后 {variable} 改变",
                "variable": variable,
                "operator": point.get("operator") or "equals",
                "expected": point.get("expected"),
                "score": parse_float(point.get("score"), 0),
                "steps": [
                    {"action": "green_flag"},
                    {"action": "wait", "ms": 200},
                    {"action": "key_press", "key": key, "ms": 100},
                    {"action": "wait", "ms": 300},
                ],
                "hint": point.get("hint"),
            }
        )
    elif kind == "movement_counter":
        variable = point.get("variable") or "steps"
        sprite = point.get("sprite") or "Player"
        key = point.get("key") or "right arrow"
        config["staticChecks"].extend(
            [
                {"type": "sprite_exists", "name": f"存在角色 {sprite}", "sprite": sprite, "score": parse_float(point.get("spriteScore"), 0)},
                {"type": "variable_exists", "name": f"存在变量 {variable}", "variable": variable, "score": parse_float(point.get("variableScore"), 0)},
                {"type": "block_exists", "name": "使用按键事件", "opcode": "event_whenkeypressed", "score": parse_float(point.get("keyEventScore"), 0)},
                {"type": "block_exists_any", "name": "使用移动积木", "opcodes": ["motion_changexby", "motion_movesteps"], "score": parse_float(point.get("motionScore"), 0)},
                {"type": "block_exists", "name": "改变变量记录移动", "opcode": "data_changevariableby", "score": parse_float(point.get("changeVariableScore"), 0)},
            ]
        )
        steps = [{"action": "green_flag"}, {"action": "wait", "ms": 200}]
        for _index in range(parse_int(point.get("presses")) or 1):
            steps.extend([{"action": "key_press", "key": key, "ms": 100}, {"action": "wait", "ms": 200}])
        config["dynamicChecks"].append(
            {
                "type": "variable_value",
                "name": f"{variable} 移动计数达到目标",
                "variable": variable,
                "operator": point.get("operator") or "greater_or_equal",
                "expected": point.get("expected"),
                "score": parse_float(point.get("score"), 0),
                "steps": steps,
                "hint": point.get("hint"),
            }
        )


def db_permission_catalog() -> list[dict]:
    rows = rows_to_dicts(
        fetch_all(
            """
            SELECT permission_key, label, description, category, is_system, status
            FROM permission_definitions
            ORDER BY category, permission_key
            """
        )
    )
    if not rows:
        return permission_catalog()
    return [
        {
            "key": row["permission_key"],
            "label": row["label"],
            "description": row["description"],
            "category": row["category"],
            "is_system": bool(row["is_system"]),
            "status": row["status"],
        }
        for row in rows
    ]


def get_permission_definition(permission_key: str) -> dict | None:
    row = fetch_one(
        """
        SELECT permission_key, label, description, category, is_system, status
        FROM permission_definitions
        WHERE permission_key = ?
        """,
        (permission_key,),
    )
    if row is None:
        return None
    row = dict(row)
    return {
        "key": row["permission_key"],
        "label": row["label"],
        "description": row["description"],
        "category": row["category"],
        "is_system": bool(row["is_system"]),
        "status": row["status"],
    }


def db_permissions_for_role(role: str | None) -> tuple[str, ...]:
    role_code = normalize_role(role)
    rows = fetch_all(
        """
        SELECT r.permission_key
        FROM role_permission_assignments r
        JOIN permission_definitions p ON p.permission_key = r.permission_key
        WHERE r.role_code = ? AND p.status = 'active'
        ORDER BY r.permission_key
        """,
        (role_code,),
    )
    if rows:
        return tuple(row["permission_key"] for row in rows)
    return permissions_for_role(role_code)


def db_role_codes(active_only: bool = True) -> set[str]:
    sql = "SELECT role_code FROM role_definitions"
    params = ()
    if active_only:
        sql += " WHERE status = 'active'"
    rows = fetch_all(sql, params)
    if rows:
        return {normalize_role(row["role_code"]) for row in rows}
    return {option["value"] for option in role_options()}


def db_role_label(role: str | None) -> str:
    role_code = normalize_role(role)
    row = fetch_one("SELECT label FROM role_definitions WHERE role_code = ?", (role_code,))
    if row:
        return row["label"]
    return role_label(role)


def db_role_level(role: str | None) -> int:
    role_code = normalize_role(role)
    row = fetch_one("SELECT level FROM role_definitions WHERE role_code = ?", (role_code,))
    if row:
        return int(row["level"] or 0)
    for option in role_options():
        if option["value"] == role_code:
            return option["level"]
    return 0


def db_role_options(active_only: bool = True) -> list[dict]:
    rows = rows_to_dicts(
        fetch_all(
            """
            SELECT role_code, label, level, note, is_system, status
            FROM role_definitions
            ORDER BY level DESC, role_code
            """
        )
    )
    if rows:
        options = []
        for row in rows:
            if active_only and row["status"] != "active":
                continue
            role_code = normalize_role(row["role_code"])
            options.append(
                {
                    "value": role_code,
                    "label": row["label"],
                    "level": row["level"],
                    "note": row["note"] or "",
                    "is_system": bool(row["is_system"]),
                    "status": row["status"],
                    "permissions": db_permissions_for_role(role_code),
                }
            )
        return options

    options = []
    for option in role_options():
        item = dict(option)
        item["is_system"] = True
        item["status"] = "active"
        item["permissions"] = db_permissions_for_role(item["value"])
        options.append(item)
    return options


def normalize_role_code(value: str | None) -> str:
    return re.sub(r"[^a-z0-9_]", "_", str(value or "").strip().lower()).strip("_")


def save_role_definition(data) -> tuple[bool, str | dict]:
    role_code = normalize_role_code(data.get("role_code") or data.get("value"))
    label = str(data.get("label") or "").strip()
    if not role_code:
        return False, "请填写身份标识。"
    if not re.match(r"^[a-z][a-z0-9_]{1,31}$", role_code):
        return False, "身份标识需以小写字母开头，只能包含小写字母、数字和下划线，长度 2-32。"
    if role_code in {"admin", "staff"}:
        return False, "该身份标识为旧账号兼容保留，请换一个。"
    if not label:
        return False, "请填写身份名称。"
    if fetch_one("SELECT role_code FROM role_definitions WHERE role_code = ?", (role_code,)):
        return False, "身份标识已存在。"
    level = parse_int(data.get("level")) or 10
    now = now_text()
    get_db().execute(
        """
        INSERT INTO role_definitions
            (role_code, label, level, note, is_system, status, created_at, updated_at)
        VALUES (?, ?, ?, ?, 0, ?, ?, ?)
        """,
        (
            role_code,
            label,
            level,
            str(data.get("note") or "").strip(),
            data.get("status") or "active",
            now,
            now,
        ),
    )
    get_db().commit()
    log_operation("新增身份", "role", None, role_code)
    return True, next(role for role in db_role_options(active_only=False) if role["value"] == role_code)


def update_role_definition(role_code: str, data) -> tuple[bool, str | dict]:
    role_code = normalize_role(role_code)
    row = fetch_one("SELECT * FROM role_definitions WHERE role_code = ?", (role_code,))
    if row is None:
        return False, "身份不存在。"
    label = str(data.get("label") or "").strip()
    if not label:
        return False, "请填写身份名称。"
    status = data.get("status") or row["status"]
    if row["is_system"] and status != "active":
        return False, "系统内置身份不能停用。"
    get_db().execute(
        """
        UPDATE role_definitions
        SET label = ?, level = ?, note = ?, status = ?, updated_at = ?
        WHERE role_code = ?
        """,
        (
            label,
            parse_int(data.get("level")) or row["level"],
            str(data.get("note") or "").strip(),
            status,
            now_text(),
            role_code,
        ),
    )
    get_db().commit()
    log_operation("更新身份", "role", None, role_code)
    return True, next(role for role in db_role_options(active_only=False) if role["value"] == role_code)


def save_role_permissions(role_code: str, permission_keys) -> tuple[bool, str]:
    role_code = normalize_role(role_code)
    if role_code not in db_role_codes(active_only=False):
        return False, "身份不存在。"
    if role_code == "super_admin":
        return False, "超级管理员权限由系统强制保持完整，不能在页面中收窄。"
    permission_keys = sorted({str(item).strip() for item in permission_keys if item})
    valid_permissions = {
        row["permission_key"] for row in fetch_all(
            "SELECT permission_key FROM permission_definitions WHERE status = 'active'"
        )
    }
    invalid_permissions = [key for key in permission_keys if key not in valid_permissions]
    if invalid_permissions:
        return False, "包含不存在或未启用的权限点：" + "、".join(invalid_permissions)
    now = now_text()
    db = get_db()
    db.execute("DELETE FROM role_permission_assignments WHERE role_code = ?", (role_code,))
    for permission_key in permission_keys:
        db.execute(
            """
            INSERT INTO role_permission_assignments
                (role_code, permission_key, created_at, updated_at)
            VALUES (?, ?, ?, ?)
            """,
            (role_code, permission_key, now, now),
        )
    db.commit()
    log_operation("更新身份权限", "role", None, role_code)
    return True, "权限已保存。"


def default_landing_url(user) -> str:
    role = normalize_role(user["role"])
    if role == "student":
        return url_for("student_portal")
    if role == "parent":
        return url_for("parent_portal")
    return url_for("dashboard")


def should_honor_login_next(user, next_url: str | None) -> bool:
    if not next_url or not next_url.startswith("/") or next_url.startswith("//"):
        return False
    role = normalize_role(user["role"])
    if role == "student":
        return next_url.startswith("/student")
    if role == "parent":
        return next_url.startswith("/parent")
    classic_allow_prefixes = (
        "/permissions",
        "/students/",
    )
    return any(next_url.startswith(prefix) for prefix in classic_allow_prefixes)


def public_user(user) -> dict:
    role = normalize_role(user["role"])
    return {
        "id": user["id"],
        "username": user["username"],
        "display_name": user["display_name"],
        "role": role,
        "raw_role": user["role"],
        "role_label": db_role_label(user["role"]),
        "role_level": permissions_role_level(role),
        "permissions": list(user.get("permissions") or db_permissions_for_role(user["role"])),
        "teacher_id": user["teacher_id"],
        "teacher_name": user["teacher_name"] if "teacher_name" in user.keys() else None,
        "student_id": user.get("student_id"),
        "student_name": user.get("student_name"),
        "status": user["status"],
    }


def public_account(user) -> dict:
    row = dict(user)
    role = normalize_role(row.get("role"))
    row["raw_role"] = row.get("role")
    row["role"] = role
    row["role_label"] = db_role_label(row.get("raw_role"))
    row["role_level"] = permissions_role_level(role)
    row["permissions"] = list(db_permissions_for_role(row.get("raw_role")))
    return row


def permissions_role_level(role: str | None) -> int:
    return db_role_level(role)


def query_operation_logs(filters: dict, limit: int = 100):
    conditions = []
    params = []
    if filters.get("q"):
        conditions.append(
            "(username LIKE ? OR action LIKE ? OR target_type LIKE ? OR detail LIKE ?)"
        )
        term = f"%{filters['q']}%"
        params.extend([term, term, term, term])
    if filters.get("start_date"):
        conditions.append("created_at >= ?")
        params.append(filters["start_date"] + "T00:00:00")
    if filters.get("end_date"):
        conditions.append("created_at <= ?")
        params.append(filters["end_date"] + "T23:59:59")
    where_sql = "WHERE " + " AND ".join(conditions) if conditions else ""
    params.append(limit)
    return fetch_all(
        f"""
        SELECT *
        FROM operation_logs
        {where_sql}
        ORDER BY created_at DESC, id DESC
        LIMIT ?
        """,
        tuple(params),
    )


def query_low_hour_students(threshold: float = 2, filters: dict | None = None):
    conditions = [
        "s.status = 'active'",
        "(s.purchased_hours + s.gift_hours) > 0",
    ]
    params = []
    join_params = []
    class_join = "LEFT JOIN classes c ON c.id = cs.class_id"
    if filters and filters.get("teacher_id"):
        teacher_id = filters["teacher_id"]
        class_join = "LEFT JOIN classes c ON c.id = cs.class_id AND c.teacher_id = ?"
        join_params.append(teacher_id)
        conditions.append(
            """
            EXISTS (
                SELECT 1
                FROM class_students cs2
                JOIN classes c2 ON c2.id = cs2.class_id
                WHERE cs2.student_id = s.id
                  AND cs2.status = 'active'
                  AND c2.teacher_id = ?
            )
            """
        )
        params.append(teacher_id)
    where_sql = " AND ".join(conditions)
    return fetch_all(
        f"""
        SELECT s.*,
               COALESCE(GROUP_CONCAT(c.class_name, '、'), '') AS class_names,
               (
                   SELECT COALESCE(SUM(a.deduct_hours), 0)
                   FROM attendance a
                   WHERE a.student_id = s.id
               ) AS consumed_hours,
               (s.purchased_hours + s.gift_hours) AS total_hours,
               ((s.purchased_hours + s.gift_hours) - (
                   SELECT COALESCE(SUM(a2.deduct_hours), 0)
                   FROM attendance a2
                   WHERE a2.student_id = s.id
               )) AS remaining_hours
        FROM students s
        LEFT JOIN class_students cs ON cs.student_id = s.id AND cs.status = 'active'
        {class_join}
        WHERE {where_sql}
        GROUP BY s.id
        HAVING remaining_hours <= ?
        ORDER BY remaining_hours ASC, s.name
        LIMIT 8
        """,
        tuple(join_params + params + [threshold]),
    )


def query_consecutive_absences(limit: int = 8, filters: dict | None = None):
    if filters and filters.get("teacher_id"):
        students = fetch_all(
            """
            SELECT DISTINCT s.id, s.name, s.parent_phone
            FROM students s
            JOIN class_students cs ON cs.student_id = s.id AND cs.status = 'active'
            JOIN classes c ON c.id = cs.class_id
            WHERE s.status = 'active' AND c.teacher_id = ?
            """,
            (filters["teacher_id"],),
        )
    else:
        students = fetch_all("SELECT id, name, parent_phone FROM students WHERE status = 'active'")
    results = []
    for student in students:
        conditions = ["a.student_id = ?"]
        params = [student["id"]]
        if filters and filters.get("teacher_id"):
            conditions.append("l.teacher_id = ?")
            params.append(filters["teacher_id"])
        where_sql = " AND ".join(conditions)
        rows = fetch_all(
            f"""
            SELECT a.status, l.lesson_date, c.class_name
            FROM attendance a
            JOIN lessons l ON l.id = a.lesson_id
            JOIN classes c ON c.id = l.class_id
            WHERE {where_sql}
            ORDER BY l.lesson_date DESC, l.start_time DESC
            LIMIT 2
            """,
            tuple(params),
        )
        if len(rows) == 2 and all(row["status"] == "缺勤" for row in rows):
            results.append(
                {
                    "student_id": student["id"],
                    "name": student["name"],
                    "parent_phone": student["parent_phone"],
                    "last_date": rows[0]["lesson_date"],
                    "class_name": rows[0]["class_name"],
                }
            )
    return results[:limit]


def query_upcoming_lessons(start_date: str, days: int = 7, filters: dict | None = None):
    end_date = (parse_date(start_date) + timedelta(days=days)).isoformat()
    conditions = ["l.lesson_date > ?", "l.lesson_date <= ?", "l.status = 'planned'"]
    params = [start_date, end_date]
    if filters and filters.get("teacher_id"):
        conditions.append("l.teacher_id = ?")
        params.append(filters["teacher_id"])
    where_sql = " AND ".join(conditions)
    return fetch_all(
        f"""
        SELECT l.*, c.class_name, t.name AS teacher_name, lt.type_name,
               (
                   SELECT COUNT(*)
                   FROM class_students cs
                   JOIN students s ON s.id = cs.student_id
                   WHERE cs.class_id = l.class_id AND cs.status = 'active' AND s.status = 'active'
               ) AS expected_count
        FROM lessons l
        JOIN classes c ON c.id = l.class_id
        LEFT JOIN teachers t ON t.id = l.teacher_id
        LEFT JOIN lesson_types lt ON lt.id = l.lesson_type_id
        WHERE {where_sql}
        ORDER BY l.lesson_date, l.start_time
        LIMIT 10
        """,
        tuple(params),
    )


def resolve_backup_file(filename: str) -> Path:
    if filename != Path(filename).name:
        abort(404)
    backup_file = config.BACKUP_DIR / filename
    if not backup_file.exists() or backup_file.suffix.lower() not in {".db", ".sql"}:
        abort(404)
    return backup_file


def apply_teacher_scope(filters: dict) -> dict:
    user = g.get("current_user")
    scoped = dict(filters)
    if user and user["role"] == "teacher":
        scoped["teacher_id"] = str(user["teacher_id"] or 0)
    return scoped


app = create_app()


if __name__ == "__main__":
    app.run(host=config.HOST, port=config.PORT, debug=False)
